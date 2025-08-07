#!/usr/bin/env python 
# encoding: utf-8 
'''
@author: chao.li
@contact: chao.li@amlogic.com
@software: pycharm
@file: test_wifi_rvr_rvo.py
@time: 2024/12/4 15:19 
@desc: 
'''

import itertools
import logging
import os
import re
import threading
import time
from copy import copy
from src.test import get_testdata
from src.test.performance.test_wifi_rvr_rvo import corner_tool

import openpyxl
import pytest

from src.tools.connect_tool.TelnetInterface import TelnetInterface
from src.tools.router_tool.Router import Router
from src.tools.router_tool.router_factory import get_router
from src.tools.yamlTool import yamlTool

# 小米极限测试 记录
filename = 'XiaoMi-Rvr.xlsx'
rvr_xlsx = openpyxl.load_workbook(filename)
sheet = rvr_xlsx['Sheet1']
new_sheet = rvr_xlsx.create_sheet(title=f'{pytest.timestamp}')

for row in sheet.iter_rows(values_only=False):
    for cell in row:
        new_sheet[cell.coordinate].value = copy(cell.value)
        new_sheet[cell.coordinate].font = copy(cell.font)
        new_sheet[cell.coordinate].border = copy(cell.border)
        new_sheet[cell.coordinate].fill = copy(cell.fill)
        new_sheet[cell.coordinate].number_format = copy(cell.number_format)
        new_sheet[cell.coordinate].protection = copy(cell.protection)
        new_sheet[cell.coordinate].alignment = copy(cell.alignment)

for merged_range in sheet.merged_cells.ranges:
    new_sheet.merge_cells(str(merged_range))

rvr_xlsx.save(filename)


def writeInExcelArea(value, row_num, col_num):
    for i in range(0, len(value)):
        logging.info(f'execl write {row_num} {i + col_num}')
        new_sheet.cell(row=row_num, column=i + col_num, value=value[i])


wifi_yaml = yamlTool(os.getcwd() + '/config/config.yaml')
router_name = wifi_yaml.get_note('router')['name']

# 设置为True 时 开启 衰减测试流程
rf_needed = False
# 设置为True 时 开启 状态测试流程
corner_needed = False
# 设置为True 时 开启 路由相关配置
router_needed = True
# 实例路由器对象
router = get_router(router_name)
test_data = get_testdata(router)
# 设置是否需要push iperf
iperf_tool = False
bt_device = 'Mi Outdoor Bluetooth Speaker'
if pytest.connect_type == 'telnet':
    third_dut = True

sum_list_lock = threading.Lock()

rvr_tool = wifi_yaml.get_note('rvr')['tool']
# env_control = wifi_yaml.get_note('env_control')

# 初始化 衰减 & 转台 对象
if rf_needed:
    # 读取衰减 配置
    rf_step_list = []
    rf_ip = ''
    model = wifi_yaml.get_note('rf_solution')['model']
    if model != 'RADIORACK-4-220' and model != 'RC4DAT-8G-95':
        raise EnvironmentError("Doesn't support this model")

    rf_ip = wifi_yaml.get_note('rf_solution')[model]['ip_address']
    logging.info('test rf')
    rf_tool = TelnetInterface(rf_ip)
    logging.info(f'rf_ip {rf_ip}')
    rf_step_list = wifi_yaml.get_note('rf_solution')['step']
    rf_step_list = [i for i in range(*rf_step_list)][::8]
    logging.info(f'rf_step_list {rf_step_list}')

if corner_needed:
    corner_step_list = []
    # 配置衰减
    corner_ip = wifi_yaml.get_note('corner_angle')['ip_address']
    logging.info('test corner')
    corner_tool = TelnetInterface(corner_ip)
    logging.info(f'corner_ip {corner_ip}')
    corner_step_list = wifi_yaml.get_note('corner_angle')['step']
    corner_step_list = [i for i in range(*corner_step_list)][::45]
    logging.info(f'corner step_list {corner_step_list}')
else:
    corner_tool = None

step_list = [1]
if rf_needed and rf_step_list:
    step_list = rf_step_list
if corner_needed and corner_step_list:
    step_list = corner_step_list
if rf_needed and corner_needed and rf_step_list and corner_step_list:
    step_list = itertools.product(corner_step_list, rf_step_list)

logging.info(f'finally step_list {step_list}')

# 配置 测试报告
pytest.testResult.x_path = [] if (rf_needed and corner_needed) == 'both' else step_list
rx_result, tx_result = '', ''
tx_result_list, rx_result_list, tx_rssi_list, rx_rssi_list = [], [], [], []


@pytest.fixture(scope='session', autouse=True, params=test_data, ids=[str(i) for i in test_data])
def setup(request):
    global rx_result_list, tx_result_list
    logging.info('==== wifi env setup start')
    tx_result_list.clear()
    rx_result_list.clear()
    # 重置衰减&转台
    # 衰减器置0
    if rf_needed:
        logging.info('Reset rf value')
        rf_tool.execute_rf_cmd(0)
        logging.info(rf_tool.get_rf_current_value())
        time.sleep(30)

    # 转台置0
    if corner_needed:
        logging.info('Reset corner')
        corner_tool.set_turntable_zero()
        logging.info(corner_tool.get_turntanle_current_angle())
        time.sleep(3)

    # push_iperf()
    router_info = request.param
    if router_info.serial == '0':
        logging.info('disconnect bt')
        disconnect_bt()
    if router_needed:
        # 修改路由器配置
        assert router.change_setting(router_info), "Can't set ap , pls check first"
        band = '5 GHz' if '2' in router_info.band else '2.4 GHz'
        ssid = router_info.ssid + "_bat";
        router.change_setting(Router(band=band, ssid=ssid))
        time.sleep(10)

    logging.info('wifi env set done')
    with open(pytest.testResult.detail_file, 'a', encoding='utf-8') as f:
        f.write(f'Testing {router_info} \n')

    if pytest.connect_type == 'telnet':
        connect_status = True
        if router_needed:
            time.sleep(90)
    else:
        # 连接 网络 最多三次重试
        for _ in range(3):
            if not router_needed:
                break
            try:
                type = 'wpa3' if 'WPA3' in router_info.authentication else 'wpa2'
                if router_info.authentication.lower() in \
                        ['open', '不加密', '无', 'open system', '无加密(允许所有人连接)', 'none']:
                    logging.info('no passwd')
                    cmd = pytest.dut.CMD_WIFI_CONNECT.format(router_info.ssid, "open", "")
                else:
                    cmd = pytest.dut.CMD_WIFI_CONNECT.format(router_info.ssid, type,
                                                             router_info.wpa_passwd)
                if router_info.hide_ssid == '是':
                    cmd += pytest.dut.CMD_WIFI_HIDE

                pytest.dut.checkoutput(cmd)
                time.sleep(5)
                if pytest.dut.wait_for_wifi_address(target=re.findall(r'(\d+\.\d+\.\d+\.)', dut_ip)[0]):
                    connect_status = True
                    break
            except Exception as e:
                logging.info(e)
                connect_status = False

    if pytest.connect_type == 'telnet':
        dut_ip = pytest.dut.dut_ip
    connect_status = True
    logging.info(f'pc_ip:{pytest.dut.pc_ip}')
    logging.info('==== wifi env setup done')

    if rvr_tool == 'ixchariot':
        if '5' in router_info.band:
            pytest.dut.ix.modify_tcl_script("set script ",
                                            'set script "$ixchariot_installation_dir/Scripts/High_Performance_Throughput.scr"\n')
        else:
            pytest.dut.ix.modify_tcl_script("set script ",
                                            'set script "$ixchariot_installation_dir/Scripts/Throughput.scr"\n')
        pytest.dut.checkoutput(pytest.dut.IX_ENDPOINT_COMMAND)
        time.sleep(3)

    yield connect_status, router_info
    # 后置动作
    pytest.dut.kill_iperf()
    if rf_needed:
        logging.info('Reset rf value')
        rf_tool.execute_rf_cmd(0)
        logging.info(rf_tool.get_rf_current_value())
        time.sleep(10)
    logging.info(f'tx_result_list {tx_result_list}')
    logging.info(f'rx_result_list {rx_result_list}')
    # 重置结果
    if tx_result_list and router_info.data_row != '0':
        writeInExcelArea(tx_result_list, row_num=int(router_info.data_row) + 1, col_num=10)
        new_sheet.cell(row=int(router_info.data_row) + 1, column=16, value=str(tx_rssi_list)[1:-1])
    if rx_result_list and router_info.data_row != '0':
        writeInExcelArea(rx_result_list, row_num=int(router_info.data_row), col_num=10)
        new_sheet.cell(row=int(router_info.data_row), column=16, value=str(rx_rssi_list)[1:-1])
    rvr_xlsx.save(filename)
    if not router_needed:
        router.router_control.driver.quit()


# 生成 pdf
# if step_list != [0]:
#     pytest.testResult.write_to_excel()
#     if test_type == 'rf':
#         # 重置衰减
#         if not rf_debug:
#             rf_tool.execute_rf_cmd(0)
#         # 生成折线图
#         pytest.testResult.write_attenuation_data_to_pdf()
#     elif test_type == 'corner':
#         # 转台重置
#         if not rf_debug:
#             corner_tool.set_turntable_zero()
#         # 生成雷达图
#         pytest.testResult.write_corner_data_to_pdf()
#     else:
#         ...
def disconnect_bt():
    pytest.dut.start_activity(*('com.android.tv.settings', '.MainSettings'))
    for _ in range(10):
        pytest.dut.keyevent(20)
        pytest.dut.uiautomator_dump()
        if 'Pair accessory' in pytest.dut.get_dump_info():
            pytest.dut.keyevent(23)
            time.sleep(1)
            break
    else:
        assert False, "Can't find Remotes & Accessories"
    for _ in range(10):
        pytest.dut.keyevent(20)
        pytest.dut.uiautomator_dump()
        if f'text="{bt_device}" resource-id="com.android.tv.settings:id/decor_title"' in pytest.dut.get_dump_info():
            pytest.dut.keyevent(22)
            time.sleep(1)
            break
    else:
        assert False, "Can't find target bt device"
    pytest.dut.keyevent(20)
    pytest.dut.keyevent(23)
    pytest.dut.keyevent(19)
    pytest.dut.keyevent(23)


# 测试 iperf

@pytest.mark.parametrize("rf_value", step_list)
def test_xiaomi_rvr(setup, rf_value):
    global rx_result, tx_result
    # 判断板子是否存在  ip
    if not setup[0]:
        logging.info("Can't connect wifi ,input 0")
        # rx_result_list.append('0')
        # tx_result_list.append('0')
        with open(pytest.testResult.detail_file, 'a') as f:
            f.write("\n Can't connect wifi , skip this loop\n\n")
        return
    router_info = setup[1]

    # 执行 修改 步长
    # 修改衰减
    if rf_needed:
        logging.info(f'set rf value {rf_value}')
        value = rf_value[1] if type(rf_value) == tuple else rf_value
        rf_tool.execute_rf_cmd(value)
        # 获取当前衰减值
        logging.info(rf_tool.get_rf_current_value())

    if corner_needed:
        logging.info('set corner value')
        value = rf_value[0] if type(rf_value) == tuple else rf_value
        corner_tool.execute_turntable_cmd('rt', angle=value * 10)
        # 获取转台角度
        logging.info(corner_tool.get_turntanle_current_angle())

    with open(pytest.testResult.detail_file, 'a') as f:
        f.write('-' * 40 + '\n')
        info, corner_set = '', ''
        db_set = 0
        if rf_needed:
            db_set = rf_value[1] if type(rf_value) == tuple else rf_value
            info += 'db_set : ' + str(db_set) + '\n'

        if corner_needed:
            corner_set = rf_value[0] if type(rf_value) == tuple else rf_value
            info += 'corner_set : ' + str(corner_set) + '\n'

        f.write(info)
    # time.sleep(1)

    # 获取rssi
    rssi_num = pytest.dut.get_rssi()
    # handle iperf pair count
    logging.info(router_info)
    protocol = 'TCP' if 'TCP' in router_info.protocol_type else 'UDP'
    # iperf  打流
    if 'tx' in router_info.test_type:
        # 动态匹配 打流通道数
        pair = wifi_yaml.get_note('rvr')['pair']
        logging.info(f'rssi : {rssi_num} pair : {pair}')
        pytest.dut.get_tx_rate(router_info, rssi_num, protocol,
                               corner_tool=corner_tool,
                               db_set=db_set)
    # 获取rssi
    rssi_num = pytest.dut.get_rssi()
    if 'rx' in router_info.test_type:
        pair = wifi_yaml.get_note('rvr')['pair']
        logging.info(f'rssi : {rssi_num} pair : {pair}')
        pytest.dut.get_rx_rate(router_info, rssi_num, protocol,
                               corner_tool=corner_tool,
                               db_set=db_set)
