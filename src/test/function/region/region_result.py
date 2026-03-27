# src/test/function/region/region_result.py
"""
区域测试结果收集与报告生成模块
- 单例模式结果收集器
- 智能续写Excel报告（精确匹配标准格式）
- 移除标题行和Notes注释，修复TN国家显示
"""
import logging
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ========================================
# 【关键配置】预定义全量信道模板（严格按需求）
# ========================================
ALL_CHANNELS_2G = list(range(1, 15))  # 2.4G: 1~14
ALL_CHANNELS_5G = [
    36, 40, 44, 48, 52, 56, 60, 64,
    100, 104, 108, 112, 116, 120, 124, 128,
    132, 136, 140, 144,
    149, 153, 157, 161, 165
]
ALL_CHANNELS = sorted(ALL_CHANNELS_2G + ALL_CHANNELS_5G)  # 共39个信道（2.4G在前，5G在后）


# ========================================
# 【核心类】结果收集器（单例模式）
# ========================================
class ResultCollector:
    _instance = None

    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance._data = {}
            cls._instance._countries = set()
            logging.info("Intialized ResultCollector singleton")
        return cls._instance

    def add_result(self, channel, band, country, result):
        """
        添加单条测试结果

        :param channel: 信道号（整数或字符串）
        :param band: 频段 ("2.4G" 或 "5G")
        :param country: 国家代码（如 "PE"）
        :param result: 测试结果 ("Pass"/"Fail")
        """
        key = (str(channel), band, country)
        self._data[key] = result
        self._countries.add(country)
        logging.debug(f"📊 收集: {band} Ch{channel} @{country} = {result}")

    def get_all_results(self):
        """返回所有累积结果（深拷贝）"""
        return self._data.copy(), self._countries.copy()

    def clear_results(self):
        """清空所有结果（用于测试隔离）"""
        self._data = {}
        self._countries = set()
        logging.debug("🧹 ResultCollector 已清空")


# ========================================
# 【核心函数】智能续写报告（精确匹配标准格式）
# ========================================
def generate_region_report(report_dir, filename="Country_Code_Test_Report.xlsx", append=True):
    """
    生成/续写区域测试Excel报告（精确匹配标准报告格式）

    :param report_dir: 报告保存目录（字符串或Path对象）
    :param filename: 报告文件名（默认"Country_Code_Test_Report.xlsx"）
    :param append: True=续写（推荐），False=强制覆盖（仍使用固定模板）
    :return: 报告完整路径（Path）或None（失败时）
    """
    report_path = Path(report_dir) / filename
    collector = ResultCollector()
    current_results, _ = collector.get_all_results()

    # 安全检查：无数据且无历史文件时退出
    if not current_results and not (append and report_path.exists()):
        logging.warning("⚠️ 无测试结果且无历史报告，无法生成报告")
        return None

    # ========================
    # STEP 1: 读取历史数据（仅提取国家列表和有效数据）
    # ========================
    history_data = {}  # {(channel_int, country): result_str}
    history_countries = set()

    if append and report_path.exists():
        try:
            wb_hist = load_workbook(report_path, data_only=True)
            ws = wb_hist.active

            # 读取国家列表（第一行B列起）
            col = 2
            while True:
                country_val = ws.cell(row=1, column=col).value
                if not country_val or str(country_val).strip() == "":
                    break
                history_countries.add(str(country_val).strip())
                col += 1

            # 读取历史数据（仅限预定义信道）
            row = 2
            while True:
                ch_val = ws.cell(row=row, column=1).value
                if ch_val is None or str(ch_val).strip() == "":
                    break

                try:
                    ch_int = int(ch_val)
                    if ch_int not in ALL_CHANNELS:  # 跳过非模板信道
                        row += 1
                        continue
                except:
                    row += 1
                    continue

                # 读取该信道所有国家的结果
                for col_idx, country in enumerate(sorted(history_countries), start=2):
                    cell_val = ws.cell(row=row, column=col_idx).value
                    if cell_val and str(cell_val).strip().upper() not in ["", "NA"]:
                        history_data[(ch_int, country)] = str(cell_val).strip()
                row += 1

            logging.info(f"📂 读取历史: {len(history_data)} 条有效数据 | 国家: {sorted(history_countries)}")
        except Exception as e:
            logging.warning(f"⚠️ 读取历史报告异常（将新建）: {str(e)}")
            history_countries = set()

    # ========================
    # STEP 2: 合并当前测试结果（覆盖同(信道,国家)位置）
    # ========================
    merged_data = history_data.copy()
    current_countries = set()

    for (ch_str, band, country), result in current_results.items():
        try:
            ch_int = int(ch_str)
            if ch_int in ALL_CHANNELS:  # 仅处理模板内信道
                merged_data[(ch_int, country)] = result
                current_countries.add(country)
        except:
            continue

    # 所有国家 = 历史国家 ∪ 当前测试国家
    all_countries = sorted(history_countries | current_countries)

    # ========================
    # STEP 3: 生成固定模板报告（精确匹配标准格式）
    # ========================
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "WiFi Region Test Results"

        # ========== 样式定义（精确匹配标准报告） ==========
        # 边框样式
        BORDER_THIN = Side(border_style="thin", color="000000")
        CELL_BORDER = Border(
            left=BORDER_THIN,
            right=BORDER_THIN,
            top=BORDER_THIN,
            bottom=BORDER_THIN
        )

        # 表头样式（国家代码行）
        HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
        HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

        # 信道列样式
        CHANNEL_FONT = Font(bold=True, size=10, color="000000")
        CHANNEL_FILL_2G = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")  # 2.4G浅蓝
        CHANNEL_FILL_5G = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # 5G浅橙
        CHANNEL_ALIGN = Alignment(horizontal="center", vertical="center")

        # 结果单元格样式
        PASS_FONT = Font(color="006100", bold=True, size=10)
        PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        FAIL_FONT = Font(color="9C0006", bold=True, size=10)
        FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        NA_FONT = Font(color="000000", size=10)
        NA_FILL = PatternFill(fill_type=None)  # 无背景色

        RESULT_ALIGN = Alignment(horizontal="center", vertical="center")

        # ========== 设置列宽（标准报告列宽） ==========
        ws.column_dimensions['A'].width = 10  # Channel列
        for col_idx in range(2, len(all_countries) + 2):  # 国家列
            ws.column_dimensions[get_column_letter(col_idx)].width = 12.5

        row_idx = 1

        # ========== 第1行：列名（Channel + 国家代码） ==========
        ws.cell(row=row_idx, column=1, value="Channel").font = HEADER_FONT
        ws.cell(row=row_idx, column=1, value="Channel").fill = HEADER_FILL
        ws.cell(row=row_idx, column=1, value="Channel").alignment = HEADER_ALIGN
        ws.cell(row=row_idx, column=1, value="Channel").border = CELL_BORDER

        for col_idx, country in enumerate(all_countries, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=country)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = CELL_BORDER
        row_idx += 1

        # ========== 第2行起：数据行（39个信道） ==========
        for channel in ALL_CHANNELS:
            # --- 第1列：信道号 ---
            ch_cell = ws.cell(row=row_idx, column=1, value=channel)
            ch_cell.font = CHANNEL_FONT
            ch_cell.alignment = CHANNEL_ALIGN
            ch_cell.border = CELL_BORDER

            # 2.4G和5G使用不同背景色
            if channel in ALL_CHANNELS_2G:
                ch_cell.fill = CHANNEL_FILL_2G
            else:
                ch_cell.fill = CHANNEL_FILL_5G

            # --- 国家列：结果数据 ---
            for col_idx, country in enumerate(all_countries, start=2):
                result = merged_data.get((channel, country), "NA")
                cell = ws.cell(row=row_idx, column=col_idx, value=result)
                cell.border = CELL_BORDER
                cell.alignment = RESULT_ALIGN

                # 应用精确样式
                if result == "Pass":
                    cell.fill = PASS_FILL
                    cell.font = PASS_FONT
                elif result == "Fail":
                    cell.fill = FAIL_FILL
                    cell.font = FAIL_FONT
                else:  # "NA" 或其他
                    cell.font = NA_FONT
                    # 保持透明背景

            row_idx += 1

        # ========== 保存报告 ==========
        Path(report_dir).mkdir(parents=True, exist_ok=True)
        wb.save(report_path)

        # ========== 日志反馈 ==========
        total_cells = len(ALL_CHANNELS) * len(all_countries)
        non_na_count = sum(1 for v in merged_data.values() if v != "NA")
        logging.info(f"✅✅✅ 报告已生成: {report_path}")
        logging.info(
            f"📊 信道: {len(ALL_CHANNELS)} (固定模板) | 国家: {len(all_countries)} | 有效结果: {non_na_count}/{total_cells}")
        logging.info(f"📈 本次更新: {len(current_results)} 条（覆盖同位置历史数据）")
        return report_path

    except Exception as e:
        logging.error(f"❌ 报告生成失败: {str(e)}", exc_info=True)
        return None