# src/test/function/region/conftest.py
# src/test/function/region/conftest.py
import pytest  # 必须导入 pytest
import logging
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# 【关键】显式注册为 pytest 插件
#pytest_plugins = ["src.test.function.region.region_logic"]


# 【关键】包装根目录 conftest 的同名 hook
@pytest.hookimpl(wrapper=True)
def pytest_sessionfinish(session, exitstatus):
    # 1. 先执行根目录 conftest 的逻辑
    yield

    # 2. 再执行我们的区域报告逻辑
    try:
        # 检查是否区域测试
        test_path = str(session.config.invocation_dir).lower()
        report_dir = Path(session.config.getoption("--resultpath") or "./report").resolve()

        if "region" in test_path.lower():
            logging.info("📍【区域专用】开始生成区域测试报告...")

            # 导入结果收集器
            from .region_result import ResultCollector

            # 获取结果
            collector = ResultCollector()
            current_results, current_countries = collector.get_all_results()
            logging.info(f"📊【区域专用】收集到区域结果: {len(current_results)}条 | 国家: {sorted(current_countries)}")

            if current_results:
                _generate_region_report(report_dir, current_results, current_countries)
            else:
                logging.warning("⚠️【区域专用】无区域测试结果！检查测试用例是否调用了 collector.add_result()")

    except Exception as e:
        logging.error(f"🔥【区域专用】区域报告生成失败: {str(e)}", exc_info=True)


def _generate_region_report(report_dir, all_results, countries):
    """生成区域测试专用报告（不依赖任何模板）"""
    report_path = report_dir / "Country_Code_Test_Report.xlsx"

    try:
        # 创建全新工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "WiFi Region Test Results"

        # 准备数据
        all_channels = sorted(set(int(k[0]) for k in all_results.keys()))
        sorted_countries = sorted(countries)

        # 设置样式
        HEADER_FONT = Font(bold=True, color="FFFFFF")
        HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        NA_FILL = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")
        CELL_ALIGN = Alignment(horizontal="center", vertical="center")

        # 写入表头
        ws['A1'] = "Channel"
        ws['A1'].font = HEADER_FONT
        ws['A1'].fill = HEADER_FILL
        ws['A1'].alignment = CELL_ALIGN

        for col_idx, country in enumerate(sorted_countries, start=2):
            cell = ws.cell(row=1, column=col_idx, value=country)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CELL_ALIGN

        # 写入数据
        for row_idx, channel in enumerate(all_channels, start=2):
            ch_cell = ws.cell(row=row_idx, column=1, value=channel)
            ch_cell.alignment = CELL_ALIGN

            for col_idx, country in enumerate(sorted_countries, start=2):
                result = "NA"
                key_5g = (str(channel), "5G", country)
                key_2g = (str(channel), "2.4G", country)

                if key_5g in all_results:
                    result = all_results[key_5g]
                elif key_2g in all_results:
                    result = all_results[key_2g]

                cell = ws.cell(row=row_idx, column=col_idx, value=result)
                cell.alignment = CELL_ALIGN

                # 应用样式
                if result == "Pass":
                    cell.fill = PASS_FILL
                    cell.font = Font(color="006100", bold=True)
                elif result == "Fail":
                    cell.fill = FAIL_FILL
                    cell.font = Font(color="9C0006", bold=True)

        # 保存报告
        report_dir.mkdir(parents=True, exist_ok=True)
        wb.save(report_path)
        logging.info(f"✅ 区域报告已生成: {report_path}")
        logging.info(f"📈 信道: {len(all_channels)} | 国家: {len(sorted_countries)}")
        return True

    except Exception as e:
        logging.error(f"❌ 报告生成失败: {str(e)}", exc_info=True)
        return False