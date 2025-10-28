from __future__ import annotations

import base64
import csv
import logging
import math
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.legend import Legend, LegendEntry
from openpyxl.chart.marker import Marker
from openpyxl.drawing.image import Image
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

LOGGER = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Style constants derived from Demo.xlsx
# ---------------------------------------------------------------------------

COLUMN_WIDTHS: Dict[str, float] = {
    "A": 13.75,
    "B": 12.625,
    "C": 11.5,
    "D": 15.0,
    "E": 14.875,
    "F": 15.125,
    "G": 18.375,
    "H": 16.125,
    "I": 14.375,
    "J": 15.0,
    "K": 12.25,
    "L": 13.5,
    "M": 12.5,
    "N": 14.0,
    "O": 12.25,
}

ROW_HEIGHT_TITLE = 42.95

COL_AML_STANDARD = 6
COL_AML_RESULT = 7
COL_RIGHT_TABLE_ITEM = 13  # Column M
COL_RIGHT_ATT = COL_RIGHT_TABLE_ITEM + 1
COL_RIGHT_ANGLE = COL_RIGHT_TABLE_ITEM + 2
COL_RIGHT_RX_RSSI = COL_RIGHT_TABLE_ITEM + 3
COL_RIGHT_TX_RSSI = COL_RIGHT_TABLE_ITEM + 4
CHART_COLUMN_GAP = 1
CHART_TITLE_ROW = 4
CHART_MIN_TOP_ROW = 5
CHART_VERTICAL_HEIGHT_ROWS = 18

COLOR_BRAND_BLUE = "2D529F"
COLOR_SUBHEADER = "B4C6E7"
COLOR_RATE_PRIMARY = "FFF2CC"
COLOR_RATE_SECONDARY = "D9E1F2"
COLOR_RSSI_RX = "CFE2F3"
COLOR_RSSI_TX = "E2F0D9"
COLOR_GRIDLINE = "BFBFBF"

FONT_TITLE = Font(name="Arial", color="FFFFFF", bold=True, size=20)
FONT_SECTION = Font(name="Arial", color="1F4E78", bold=True, size=16)
FONT_HEADER = Font(name="Arial", color="FFFFFF", bold=True, size=11)
FONT_SUBHEADER = Font(name="Arial", color="1F4E78", bold=True, size=11)
FONT_BODY = Font(name="Arial", color="333333", size=11)
FONT_STANDARD = Font(name="Arial", color="333333", size=11)

ALIGN_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

BORDER_THIN = Border(
    left=Side(style="thin", color="8A8A8A"),
    right=Side(style="thin", color="8A8A8A"),
    top=Side(style="thin", color="8A8A8A"),
    bottom=Side(style="thin", color="8A8A8A"),
)

LOGO_PNG_BASE64 = (
    "iVBORw0KGgoAAAANSUhEUgAAAWwAAABoCAYAAADVecobAAAACXBIWXMAABcSAAAXEgFnn9JSAAAAGXRFWHRTb2Z0"
    "d2FyZQBBZG9iZSBJbWFnZVJlYWR5ccllPAAADvFJREFUeNrsnetx4soWhbcp/4cbgXUisCYCayIYTgTGEZiJwHIE"
    "gyMwjmBwBBYRjIhgIIIrIuDSx7svGo4kpH6pJa2vqssuG/Ro7V69+qmrw+FAnjM5pu0xjVs6/+6YEk6rY8oIAABa"
    "4KoDgh0f05Mn17Ln61kgdJwSKuZ5hKwDEOzhuOsy1sc0hdt2hhDeD5X4RtaBPjHy/PrmHoq14I4+u0gAAACCfSRg"
    "wfaVW/rsHgEAgMELduypu84j+tYnCCMAwJAFW7jr+47k4QxhBAAYsmDHHcpDCDYAYLCC3SV3LbhFGAEAhirYcQfz"
    "MUQoAQCGJthhx9y1BAOPALgnOaZDwxR3+YZ9E2ysIAQAgA4IdkSfC1IAAAB4LtgxHgcAAPgv2HDXAADQEcGGuwYA"
    "gA4I9hTuGgAAuiHYmBkCAAAdEOzZMd3gMQAAgP+CHeMRAACA/4INdw0AAB0RbLhrAABowHWLYg13DQDQQbyRquk+"
    "PlsIdjMm5PervwAA3SAd2g230SXi64t1AQAAgg13DQAA3RNsuGsAAOiAYMNdAwCABi4HHWO4a9BjQjrNWMj/vqXT"
    "zIT872B4TOj0OsH87wIxgJrx7xmVDKi6EuzgmB7xvEBPEPEc5VLTKaprLpDJMa06JjQ2KRWqjpKPkVDBsG7O4iS7"
    "OhwOLi58Sd18V2NdvnKmAnuB/6HwvSvDoiV2lhTdercGj7vnwrjwRKzyAhMYvtc6FVnUg5bWnGPFdI/CmwuHHTgW"
    "602ups5y1yBTmwt25DX0lT5WWnLsxdaA+ZjLxz0LVtxCPgqRnFkSGdvx1nRr5meys8o64uPa3Cr63oVgx5aP/84P"
    "Lm0Q6Pmmisu9uEWheBqAI17nmnFdbuJOuXXoSsTuuCXxwuUmcyDUC8cuuo8V+sKVKbU9S8SWuxYu+uGY/sOFatHQ"
    "lSRcICI+xvdj2iH2jAqPqJh+0ecg26yDhVBUNj9bcpyPXNGFFstlwpUDxFqvQt+67EGwLdhLC87tKwfy0pADyVjw"
    "A64EINxmEV1Qr5YFyIaYffMg3xIWBdMikxLe8qTLvI0K3aZgm+xu2LOYRmS3f2/JovKCeDTOLT87n0U7ZDHzxXWO"
    "WRRMtVBmLbYa+oTQiR9tnNhmH3Zs6DibXNPjUmGL6DS6HbBL2dOpHzWlU1/3tsJxz/lzr4hN4wKU8HPyrW97wtfm"
    "o5i95mJXR6wRz2bEurUZb7YE25S7fmPxzCqar3IKzU2FSMhryV/TjjN/UXL8ZU7c4UjMivaKK9bMo+vy/TnL1slW"
    "sTxCrM20UFqdnmyrS8SEu37nDMpK3JAQ1N/0OUCjMlVPfOeJC0DZ9aaE5fQ2uPEsX7swU2JMamNCcgAV6BH6UOnZ"
    "EGwT7npD5f12pkdmxyzcKRXPkRaFBH3a5lHZfN5Wa9D0Ktx1Lu0NHveOmvdnL9BCNIKtSm+Xi5VNG4K90Pz+nkU5"
    "K2mS2Bo0uaXymQxzwuwRG45x6sF1LAwVOrEg4wt9rq6McklUSn/R56D52tD11q3oAoPGZsfG5e/cfZal557Fakzm"
    "FtwJUf6ey8OA/lxdesX/+14UL6b7sGcGmpYxFffTzWo2Sd7p34topvz9Mf05CHnuXmSzMyqoMOQ0HtdsNLoPVJr6"
    "bxea3hEHmYlVcVMyP/XTdbw+U/k4iGTL9ylja6WRd2OOh7hmWTIh1PMBd6uY2mW0ySpWOcAspxvH/694xV4iBtP2"
    "oMe25LjRhe9lxxQf06Ti2sT/0mNKCv4Xnh1v0fD+opr5EyvkSaLxPBKF88U1jz1RvJ9z6pwrsnTsVPPaZ4rPZaJ5"
    "7m2NcwQGns1S8f5U4yLxKL517yXP3IC2Co1KR4bdim6zIS6p4ZYXHLWshapcTkbl08nOm5iPJf3ZMYF8fsbcfNPp"
    "p21rXnao6a4fNFoHGbcuVPPtpkZ3km530zN1b4WqDXTd9YOhbrd/umtNCraumO1KCsC8oiJ4q+jvLiso85Jmfp37"
    "wWh7cSDpBHVbA486YvRuoCtnq1lmppbvD+bk1I3aRqVeyMjgjem662XD/qM3gw5gUrNAZGRm4KhvLDXyJWrpmlUd"
    "6N5g3C1IfTB7eiGebz24v66j00oxUalbEeyJIcu/LMmwohquatpfRKd9Rg6ckgtOrqhZXjaLIUEc135+vhJoGIwV"
    "mV3wo5pvYyrvTtKpBBfk14KmNlHNR2uVngnBNrFP8IbKZ4Y0ae6JYBM7kN2fXdPdBaENGgg5BLuYZAAFkQyZk/MK"
    "QJWw4d/7VvHaROUNMfk8tFLp6Qq2qSkvScmxixbgiK6QooHDmKoXQNxWNHFuGhTsFLFcyJbMLhKx7bBVnZPp559q"
    "5FtguEIqM05w1+1W6sYE29RbOJIGGbYsqQ3rvBggNOBG0FysFp8+F8bEs3wLK4xU31tJvlbqVis9HcE25a7LArYo"
    "GHclQWVrlsJdxUMBqJDaPu6kojUJwdYj9CxGtAXb5DvutjUzbFUStHWX32aG3BZc9jALoy3n5Es8Ia71WyleCnZA"
    "5t5NuG6QYYmm4KYNH8zecIEHfqBqNLae3UdoUGjgsM20UrwU7NhBht3VDKhQMzNDhczH7mfA14oHZqLHqAi2cNcm"
    "N/Gu61r2JU22ugG6Kfl+0LCJGCFsAABdEWzT7jrQbGrUbQKWzXe9MejIAXAJBr8h2E7d9SVHbZKlgviu4LCBx2CQ"
    "EILt1F1XOWyTnfdisc22oTvfFVzDBII9aNC6Aq3GSBPBjiy567p7Oqi+dkw49XnFPTWpnExs2g/aR3WjKls7C6qa"
    "gG3Nv6FCckfgi2DHFq+j7p4dQc3P5cU6Umg6igK9dJwHYLjuSbWQmxbsCR6xdqVutQU+anARdxavo+gm05qfW1Jx"
    "f/eayl9YUHWONRXvOTInc+91A+2SeFQYA424ShyUw6GiOj5wa9Nl1xVs284yqhmMsxJHITJIbBb+TKeXoZ6LddnK"
    "yRcWafHza4kjD+Gue4WqCx1bcNk6IpkadodThMbFvK3D3NZFjWoG1J3lzIlKarj3s7/dVXx2yaIaF2S2KGSLisyN"
    "+GdS0gxOCH3XcNjlpoFaON6mwgWqVkhW3eHAYsRK91IdwXbhLMclgbsoccVNXI4U3DuFTJxCrHvrsHceFEYdM5RY"
    "codoSeoL9pgs7Ss+qhGcd44yaFaSaeuCzEhqNN8mHHy/coK7rFnYQq4YfkKse8tKozCa2u9YRxyXlsTmntCXLXnX"
    "+O43C60xuvaotpXdHUmBkKdnwjlmMd1xwUtyzcOQjxMViO03PtaKTq96Svmzk9z3bhGrvUcI3qOGqCWaLirWMEM7"
    "ujyYLj6j8xq0iPCyjhVrhiqvNSrXRi2y0QXH63pWRFzSfC3rxL/hQifE+4PTD87k8YXvfLD7PvDvP/nvebHeE+gr"
    "Kem9UPlVw0GJeH4yXE6aOPA6rYiEMC97aUADXrlFptONFnDl8THyxF3nXfa8JOMeHF/LM2fUjkCfC6RuYVxR/YG6"
    "gIXwh8Y5dzWvW/fexmxoYhr2/GwT3V+PbBCaVvBTfo6/pdO/9shdS35wUKcFASi7M2xe25orjTSXaQmhL7uvgj0j"
    "vXGab5w2HJsibs5nb0QcRya62uoaKdEyFVsy6K5OfuL0xuUgIX/2BXexl8rCkB7e5Ny21Lek4HMhp8JV1dceues8"
    "CRX3oaXsUmLDlcqeC9uyIBNlzfgT+tYaNgVizk5Sl1uyP/axbuicYzK3ncL9mfhXvZA4cBQXKen1MdetFOYGy/84"
    "V8k37hYblTzktlf0XepDizko/qbPBS8bxeB/4WNMWJSTkuBruwKDYNvrmhKF/rkDebBXaFJvLcbumFsmRcmVfiSO"
    "ziPM3LsPQXDusE2+WNeUaM8rXIWc7ZFvTkxqFNC6Tak5Bzy6Q/zovniydOyY3CwQ02Gm2NJY8L1962FMJKQ3G6Zp"
    "/ifU8gyy6wKB8kmcxLW8crNuXiNgTU1DijjQMb3PHxaW43PqQ4Es4YHU5417IzYWK9tXB+fJcvnYmkaOPHXX5wh3"
    "8Jtdls2pRjMW/Q+ItXdkZLdrKuOK2re3uDyQ/owPeW/rHsbF0uEzk2s2WpvuO/LYXRchBj1+ccbNDYj3hE5TZzKu"
    "qSHUfrvstwGJtgmxPr+3lx7GhbgvV9NvU8fn+4PrDrjrIoSoyrmscrQ6odPKRZmxGd9bmLtP8XvAPyHO3WPGz/XR"
    "0vEzOm0W9tjSPe7YSNhYaSjK+Yr0Vlr62PqSewa5KNMpn29JjscGrnPOpasDa/nRajAM5M6KS4txK4XN9VjGM5/T"
    "5hzjhE7bMPRFuLPc/Tw6Op8cW4td6eeI3L5YFwBTyBWGz2SvTzFhJ/VguQksrl909fzFhd/Vy3WlcIvzfqfu93HL"
    "OdNfHN7LwkEcSt6uCXOMQbcLaMyFZsrJRhN1yUmI24x/mphKJhfCrKjdN6BvOQ/lMmy52i7IpSYE1O5aDtnPHNBp"
    "v/vbjsbhJhd/2TXf1HpgBT0kzK3um3DLoCYuoDZmEyV0WqwRnqVJhSjIcRY5xpJQs/UAbQie7p7aTx7cx5ZOY3MB"
    "P6fMURxOcnEY8f+rup52fL0yFcbI1eFwQHF3h4pbyTQKT52FREVBvu3I+YCfqAr2O+E1ZZVcIwuc1/guxcn1fsYp"
    "HjFA/NhjhCwAAFhoSQIINgCgA6iOH8BhQ7ABAI7dteqMDAg2BBsA4JCZ4vfkLAkAwQYAOEBni4sE2QfBBgC4Iyb1"
    "9Q0rZN9lMA8bAGCCGanvS72nYb/oFw4bAPDPirvAwXnk7oaqwF3DYQMweA7sXuU+ITaWZes4a4nYfGqLxwWHDcDQ"
    "EX3KYpn4f8nsG5sCdsa6Yv0GsYbDBgB8OuwiNnR6gXXTuc9yNzoTWzLvuQKBYEOwAYBg1xRNuTtfRv/eIS7K/TS9"
    "y6XYQzrGY4JgAwDqCXZbrHOVAYBgAwDB9vS6dmR/b+pegkFHAIBLRBfMFGINwQYA+C/WEWGTJwg2AMBrxMyUEGIN"
    "wQYA+M0LO+stskIPvCIMAGDTVYvd+xJkBRw2AKCaXYtC/UCfXSAQawg2AKAGAQvnu4NziQFFscz8Cwv1EtlvHszD"
    "BmAYiO1LIzqtWLwz4N7lCkmVJe5Agf8JMADx3l232RGuYgAAAABJRU5ErkJggg=="
)


# ---------------------------------------------------------------------------
# Scenario model
# ---------------------------------------------------------------------------


DEFAULT_ATTENUATIONS = list(range(0, 78, 3))


@dataclass
class RvrScenario:
    freq: str = "5G"
    standard: str = "Auto"
    bandwidth: str = "20/40/80 MHz"
    channel_label: str = "CH36"
    angle_label: str = "0°"
    attenuation_steps: List[float] = field(default_factory=lambda: DEFAULT_ATTENUATIONS.copy())
    rx_values: Dict[float, float] = field(default_factory=dict)
    tx_values: Dict[float, float] = field(default_factory=dict)
    rssi_rx: Dict[float, float] = field(default_factory=dict)
    rssi_tx: Dict[float, float] = field(default_factory=dict)

    @property
    def title(self) -> str:
        return f"{self.freq.upper()} {self.standard.upper()} {self.bandwidth.upper()}"

    @property
    def subtitle(self) -> str:
        return f"{self.freq} {self.standard.upper()} {self.bandwidth}"

    @property
    def channel(self) -> str:
        return self.channel_label

def _sanitize_number(value: Optional[str]) -> Optional[float]:
    if value in (None, "", "NULL"):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _format_angle(value: Optional[str]) -> str:
    text = str(value or "").strip()
    if not text:
        return "0?"
    return text if text.endswith("?") else f"{text}?"


def _format_bandwidth(value: Optional[str]) -> str:
    text = str(value or "").strip()
    if not text:
        return "20/40/80 MHz"
    normalized = text.replace("_", " ").replace("MHz", "").strip()
    if normalized.lower().endswith("mhz"):
        return normalized
    return f"{normalized} MHz"


def _format_channel(value: Optional[str]) -> str:
    text = str(value or "").strip()
    if not text:
        return "CH36"
    upper = text.upper()
    if upper.startswith("CH"):
        return text.title()
    if upper.replace(".", "").isdigit():
        return f"CH{upper}"
    return f"CH {text.title()}"



# ---------------------------------------------------------------------------
# Utility functions
# ---------------------------------------------------------------------------


def _configure_sheet(ws: Worksheet) -> None:
    for column, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[column].width = width
    ws.row_dimensions[1].height = ROW_HEIGHT_TITLE
    # Allow the remaining rows to auto-size based on their wrapped content so
    # long descriptions do not overlap neighbouring cells.
    for row in range(2, 100):
        ws.row_dimensions[row].height = None


def _merge(ws: Worksheet, range_string: str) -> None:
    ws.merge_cells(range_string)


def _set_cell(
    ws: Worksheet,
    row: int,
    column: int,
    value,
    font: Optional[Font] = None,
    alignment: Optional[Alignment] = None,
    fill: Optional[str] = None,
    border: bool = False,
) -> None:
    cell = ws.cell(row=row, column=column, value=value)
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    if border:
        cell.border = BORDER_THIN


# ---------------------------------------------------------------------------
# Standard text helpers
# ---------------------------------------------------------------------------


def _aml_standard_text(att: float) -> Optional[str]:
    if att <= 15:
        return "RX Tput>=503.5\nTX Tput>=475"
    if att <= 21:
        return "RX Tput>=320\nTX Tput>=300"
    if 30 <= att <= 45:
        return "RX Tput>100\nTX Tput>95"
    if att >= 48:
        return "N/A"
def _aml_threshold(att: float) -> Optional[Tuple[int, int]]:
    if att <= 12:
        return 400, 300
    if att <= 21:
        return 320, 200
    if att <= 45:
        return 100, 80
    return None


def _group_runs(values: Sequence[float], formatter) -> List[Tuple[int, int, Optional[str]]]:
    if not values:
        return []
    groups: List[Tuple[int, int, Optional[str]]] = []
    start = 0
    current = formatter(values[0])
    for idx in range(1, len(values)):
        text = formatter(values[idx])
        if text != current:
            groups.append((start, idx - 1, current))
            start = idx
            current = text
    groups.append((start, len(values) - 1, current))
    return groups


def _apply_grouped_texts(ws: Worksheet, column: int, start_row: int, values: Sequence[float], formatter) -> None:
    for start_idx, end_idx, text in _group_runs(values, formatter):
        if not text:
            continue
        top = start_row + start_idx
        bottom = start_row + end_idx
        if top != bottom:
            _merge(ws, f"{get_column_letter(column)}{top}:{get_column_letter(column)}{bottom}")
        _set_cell(ws, top, column, text, font=FONT_STANDARD, alignment=ALIGN_LEFT_WRAP, border=True)
        for r in range(top + 1, bottom + 1):
            cell = ws.cell(row=r, column=column)
            cell.border = BORDER_THIN
            cell.alignment = ALIGN_LEFT_WRAP


def _set_result_cell(ws: Worksheet, row: int, column: int, threshold: Optional[Tuple[int, int]]) -> None:
    if threshold is None:
        value = "Pass"
    else:
        rx, tx = threshold
        value = f'=IF(AND(D{row}>{rx},E{row}>{tx}),"Pass","Fail")'
    _set_cell(ws, row, column, value, font=FONT_BODY, alignment=ALIGN_CENTER, border=True)


def _apply_result_formatting(ws: Worksheet, start_row: int, end_row: int) -> None:
    if start_row > end_row:
        return
    column_letter = get_column_letter(COL_AML_RESULT)
    fail_rule = FormulaRule(
        formula=[f"={column_letter}{start_row}=\"Fail\""],
        font=Font(name="Arial", color="FF0000", bold=True),
    )
    pass_rule = FormulaRule(
        formula=[f"={column_letter}{start_row}=\"Pass\""],
        font=Font(name="Arial", color="008000", bold=True),
    )
    ws.conditional_formatting.add(f"{column_letter}{start_row}:{column_letter}{end_row}", fail_rule)
    ws.conditional_formatting.add(f"{column_letter}{start_row}:{column_letter}{end_row}", pass_rule)


# ---------------------------------------------------------------------------
# Layout writer
# ---------------------------------------------------------------------------


def _write_title(ws: Worksheet, scenario: RvrScenario) -> None:
    _merge(ws, "A1:W1")
    _merge(ws, "A2:W2")
    _merge(ws, "A3:W3")
    _merge(ws, "A4:W4")

    _set_cell(ws, 1, 1, "RVR Test Report", font=FONT_TITLE, alignment=ALIGN_CENTER, fill=COLOR_BRAND_BLUE, border=True)
    _set_cell(ws, 2, 1, "1. Throughput: 2.4G", font=FONT_SECTION, alignment=ALIGN_CENTER, border=True)
    _set_cell(ws, 3, 1, "2. Throughput: 5G", font=FONT_SECTION, alignment=ALIGN_CENTER, border=True)
    subtitle = f"{scenario.freq} {scenario.standard.upper()} {scenario.bandwidth}"
    _set_cell(ws, 4, 1, subtitle, font=FONT_SECTION, alignment=ALIGN_CENTER, border=True)

    logo_bytes = base64.b64decode(LOGO_PNG_BASE64)
    image = Image(BytesIO(logo_bytes))
    image.width = 240
    image.height = 70
    image.anchor = "A1"
    ws.add_image(image)


def _write_headers(ws: Worksheet, scenario: RvrScenario) -> None:
    merges = [
        "A5:A6",
        "B5:B6",
        "C5:C6",
        "F5:F6",
        "G5:G6",
    ]
    right_item_letter = get_column_letter(COL_RIGHT_TABLE_ITEM)
    merges.append(f"{right_item_letter}5:{right_item_letter}6")
    for rng in merges:
        _merge(ws, rng)

    headers = [
        (5, 1, "Item"),
        (5, 2, "ATT\n(Unit:dB)"),
        (5, 3, "Angle"),
        (5, 4, "RX(Unit:Mbps)"),
        (5, 5, "TX(Unit:Mbps)"),
        (5, COL_AML_STANDARD, "AML_Standard"),
        (5, COL_AML_RESULT, "AML_Result"),
        (5, COL_RIGHT_TABLE_ITEM, "Item"),
        (5, COL_RIGHT_ATT, "ATT\n(Unit:dB)"),
        (5, COL_RIGHT_ANGLE, "Angle"),
        (5, COL_RIGHT_RX_RSSI, "RX_RSSI (Unit:dBm)"),
        (5, COL_RIGHT_TX_RSSI, "TX_RSSI (Unit:dBm)"),
    ]
    for row, col, text in headers:
        _set_cell(ws, row, col, text, font=FONT_HEADER, alignment=ALIGN_CENTER_WRAP, fill=COLOR_BRAND_BLUE, border=True)

    _set_cell(ws, 6, 4, scenario.channel, font=FONT_SUBHEADER, alignment=ALIGN_CENTER, fill=COLOR_SUBHEADER, border=True)
    _set_cell(ws, 6, 5, scenario.channel, font=FONT_SUBHEADER, alignment=ALIGN_CENTER, fill=COLOR_SUBHEADER, border=True)
    for column in (1, 2, 3, COL_AML_STANDARD, COL_AML_RESULT):
        _set_cell(ws, 6, column, None, font=FONT_SUBHEADER, alignment=ALIGN_CENTER, fill=COLOR_SUBHEADER, border=True)

    _set_cell(ws, 6, COL_RIGHT_RX_RSSI, scenario.channel, font=FONT_SUBHEADER, alignment=ALIGN_CENTER, fill=COLOR_SUBHEADER, border=True)
    _set_cell(ws, 6, COL_RIGHT_TX_RSSI, scenario.channel, font=FONT_SUBHEADER, alignment=ALIGN_CENTER, fill=COLOR_SUBHEADER, border=True)
    for column in (COL_RIGHT_TABLE_ITEM, COL_RIGHT_ATT, COL_RIGHT_ANGLE):
        _set_cell(ws, 6, column, None, font=FONT_SUBHEADER, alignment=ALIGN_CENTER, fill=COLOR_SUBHEADER, border=True)


def _write_data(ws: Worksheet, scenario: RvrScenario, start_row: int = 7) -> int:
    attenuations = sorted(scenario.attenuation_steps) or DEFAULT_ATTENUATIONS
    end_row = start_row + len(attenuations) - 1

    _merge(ws, f"A{start_row}:A{end_row}")
    right_item_letter = get_column_letter(COL_RIGHT_TABLE_ITEM)
    _merge(ws, f"{right_item_letter}{start_row}:{right_item_letter}{end_row}")
    _set_cell(ws, start_row, 1, scenario.subtitle, font=FONT_BODY, alignment=ALIGN_CENTER_WRAP, border=True)
    _set_cell(ws, start_row, COL_RIGHT_TABLE_ITEM, f"{scenario.standard.upper()} {scenario.bandwidth}", font=FONT_BODY, alignment=ALIGN_CENTER_WRAP, border=True)

    for r in range(start_row + 1, end_row + 1):
        for c in (1, COL_RIGHT_TABLE_ITEM):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER_THIN
            cell.alignment = ALIGN_CENTER

    for index, attenuation in enumerate(attenuations):
        row = start_row + index
        highlight = COLOR_RATE_PRIMARY if index < 6 else COLOR_RATE_SECONDARY

        _set_cell(ws, row, 2, attenuation, font=FONT_BODY, alignment=ALIGN_CENTER, border=True)
        _set_cell(ws, row, 3, scenario.angle_label, font=FONT_BODY, alignment=ALIGN_CENTER, border=True)

        _set_cell(ws, row, 4, scenario.rx_values.get(attenuation), font=FONT_BODY, alignment=ALIGN_CENTER, fill=highlight, border=True)
        _set_cell(ws, row, 5, scenario.tx_values.get(attenuation), font=FONT_BODY, alignment=ALIGN_CENTER, fill=highlight, border=True)

        for column in (COL_AML_STANDARD,):
            _set_cell(ws, row, column, None, font=FONT_STANDARD, alignment=ALIGN_LEFT_WRAP, border=True)

        _set_result_cell(ws, row, COL_AML_RESULT, _aml_threshold(attenuation))

        _set_cell(ws, row, COL_RIGHT_ATT, attenuation, font=FONT_BODY, alignment=ALIGN_CENTER, border=True)
        _set_cell(ws, row, COL_RIGHT_ANGLE, scenario.angle_label, font=FONT_BODY, alignment=ALIGN_CENTER, border=True)
        _set_cell(ws, row, COL_RIGHT_RX_RSSI, scenario.rssi_rx.get(attenuation), font=FONT_BODY, alignment=ALIGN_CENTER, fill=COLOR_RSSI_RX, border=True)
        _set_cell(ws, row, COL_RIGHT_TX_RSSI, scenario.rssi_tx.get(attenuation), font=FONT_BODY, alignment=ALIGN_CENTER, fill=COLOR_RSSI_TX, border=True)

    _apply_grouped_texts(ws, COL_AML_STANDARD, start_row, attenuations, _aml_standard_text)
    _apply_result_formatting(ws, start_row, end_row)
    return end_row

def _nice_number(value: float, *, round_up: bool = True) -> float:
    if value <= 0:
        return 0.0
    exponent = math.floor(math.log10(value))
    fraction = value / (10 ** exponent)
    candidates = (1, 2, 2.5, 5, 10)
    if round_up:
        for candidate in candidates:
            if fraction <= candidate:
                fraction = candidate
                break
        else:
            fraction = candidates[-1]
    else:
        for candidate in reversed(candidates):
            if fraction >= candidate:
                fraction = candidate
                break
        else:
            fraction = candidates[0]
    return fraction * (10 ** exponent)


def _resolve_throughput_axis(values: Sequence[Optional[float]]) -> Tuple[float, float]:
    numeric_values: list[float] = []
    for value in values:
        if value is None:
            continue
        try:
            numeric = float(value)
        except (TypeError, ValueError):
            continue
        if not math.isfinite(numeric):
            continue
        numeric_values.append(numeric)
    if not numeric_values:
        return 10.0, 2.0
    max_value = max(numeric_values)
    if max_value <= 0:
        return 10.0, 2.0
    padded = max_value * 1.1
    upper = _nice_number(padded, round_up=True)
    if upper <= 0:
        upper = 10.0
    major = _nice_number(upper / 5.0, round_up=True)
    if major <= 0:
        major = upper
    LOGGER.info(
        "Resolved throughput axis | max=%.2f padded=%.2f upper=%.2f major=%.2f",
        max_value,
        padded,
        upper,
        major,
    )
    return upper, major


def _style_chart(
    chart: ScatterChart,
    *,
    data_points: Sequence[Optional[float]],
    show_markers: bool,
    step_values: Sequence[float],
) -> None:
    chart.width = 13.8
    chart.height = 7.5
    chart.varyColors = False
    LOGGER.info("Configured scatter chart | vary_colors=%s", chart.varyColors)

    if chart.legend is None:
        chart.legend = Legend()
    chart.legend.position = "b"
    chart.legend.overlay = False
    legend_entries = [LegendEntry(idx=0, delete=False)]
    if data_points:
        for idx in range(1, len(data_points)):
            legend_entries.append(LegendEntry(idx=idx, delete=True))
    chart.legend.legendEntry = legend_entries
    legend_summary = [(entry.idx, entry.delete) for entry in legend_entries]
    LOGGER.info(
        "Chart legend configured | position=%s overlay=%s kept_entries=%d removed_entries=%d entries=%s",
        chart.legend.position,
        chart.legend.overlay,
        1,
        max(len(legend_entries) - 1, 0),
        legend_summary,
    )

    chart.x_axis.majorGridlines = None
    chart.y_axis.majorGridlines = ChartLines()
    chart.y_axis.majorGridlines.spPr = GraphicalProperties(
        ln=LineProperties(solidFill=COLOR_GRIDLINE, w=12000)
    )
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.title = None
    chart.y_axis.title = None
    chart.x_axis.majorTickMark = "cross"
    chart.y_axis.majorTickMark = "out"
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "nextTo"
    chart.x_axis.crosses = "min"
    chart.y_axis.crosses = "min"
    chart.x_axis.number_format = "0"
    chart.y_axis.number_format = "0"
    chart.x_axis.tickLblSkip = 1
    chart.x_axis.tickMarkSkip = 1
    if step_values:
        x_min = float(step_values[0])
        x_max = float(step_values[-1])
        if len(step_values) > 1:
            raw_step = float(step_values[1] - step_values[0])
            major_step = raw_step if raw_step > 0 else 1.0
        else:
            major_step = 1.0
    else:
        x_min = 0.0
        x_max = float(max(len(data_points) - 1, 0))
        major_step = 1.0 if x_max == 0 else x_max / max(len(data_points) - 1, 1)
        if major_step <= 0:
            major_step = 1.0

    chart.x_axis.scaling.min = x_min
    chart.x_axis.scaling.max = x_max
    chart.x_axis.majorUnit = max(major_step, 1.0)

    upper, major = _resolve_throughput_axis(data_points)
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = upper
    chart.y_axis.majorUnit = major
    chart.y_axis.minorTickMark = "none"
    chart.x_axis.minorTickMark = "none"

    chart.plot_area.layout = Layout(
        manualLayout=ManualLayout(x=0.06, y=0.12, w=0.8, h=0.7)
    )
    LOGGER.info(
        "Chart layout + axis | width=%.1f height=%.1f y_max=%.2f major=%.2f x_min=%.2f x_max=%.2f x_step=%.2f layout=(x=%.2f y=%.2f w=%.2f h=%.2f)",
        chart.width,
        chart.height,
        upper,
        major,
        x_min,
        x_max,
        chart.x_axis.majorUnit,
        chart.plot_area.layout.manualLayout.x,
        chart.plot_area.layout.manualLayout.y,
        chart.plot_area.layout.manualLayout.w,
        chart.plot_area.layout.manualLayout.h,
    )

    for series in chart.series:
        if hasattr(series, "graphicalProperties") and hasattr(series.graphicalProperties, "line"):
            series.graphicalProperties.line.width = 20000  # 2pt
            series.graphicalProperties.line.solidFill = COLOR_BRAND_BLUE
        series.smooth = False
        if show_markers:
            marker = Marker(symbol="circle")
            marker.graphicalProperties = GraphicalProperties(solidFill=COLOR_BRAND_BLUE)
            marker.size = 6
        else:
            marker = Marker(symbol="none")
        series.marker = marker
        LOGGER.info(
            "Styled chart series | title=%s line_width=%s marker_symbol=%s marker_size=%s smooth=%s",
            getattr(series, "title", None),
            getattr(series.graphicalProperties.line, "width", None) if hasattr(series, "graphicalProperties") else None,
            marker.symbol,
            getattr(marker, "size", None),
            series.smooth,
        )


def _build_throughput_chart(
    *,
    title: str,
    series_title: str,
    categories: Reference,
    values: Reference,
    data_points: Sequence[Optional[float]],
    show_markers: bool,
    step_values: Sequence[float],
) -> ScatterChart:
    chart = ScatterChart()
    chart.scatterStyle = "line"
    chart.varyColors = False
    chart.title = title
    series = Series(
        values,
        xvalues=categories,
        title=series_title,
    )
    series.smooth = False
    chart.series.append(series)
    _style_chart(
        chart,
        data_points=data_points,
        show_markers=show_markers,
        step_values=step_values,
    )
    LOGGER.info(
        "Built throughput chart | title=%s series=%s point_count=%d markers=%s",
        title,
        series_title,
        len(data_points),
        show_markers,
    )
    return chart


def _add_charts(ws: Worksheet, scenario: RvrScenario, start_row: int, end_row: int) -> None:
    if start_row > end_row:
        return

    attenuations = sorted(scenario.attenuation_steps) or DEFAULT_ATTENUATIONS
    rx_series = [scenario.rx_values.get(att) for att in attenuations]
    tx_series = [scenario.tx_values.get(att) for att in attenuations]
    LOGGER.info(
        "Preparing chart data | attenuations=%s rx_sample=%s tx_sample=%s",
        attenuations[:10],
        rx_series[:10],
        tx_series[:10],
    )

    categories = Reference(ws, min_col=2, min_row=start_row, max_row=end_row)
    rx_values = Reference(ws, min_col=4, min_row=start_row, max_row=end_row)
    tx_values = Reference(ws, min_col=5, min_row=start_row, max_row=end_row)

    rx_title = f"{scenario.title} RVR Throughput_RX"
    tx_title = f"{scenario.title} RVR Throughput_TX"

    rx_chart = _build_throughput_chart(
        title=rx_title,
        series_title=scenario.channel,
        categories=categories,
        values=rx_values,
        data_points=rx_series,
        show_markers=False,
        step_values=attenuations,
    )
    tx_chart = _build_throughput_chart(
        title=tx_title,
        series_title=scenario.channel,
        categories=categories,
        values=tx_values,
        data_points=tx_series,
        show_markers=False,
        step_values=attenuations,
    )

    point_count = end_row - start_row + 1
    first_anchor_row = max(CHART_TITLE_ROW + 1, CHART_MIN_TOP_ROW)
    chart_spacing = 0

    chart_start_col_index = COL_AML_RESULT + CHART_COLUMN_GAP
    left_anchor_col = get_column_letter(chart_start_col_index)
    LOGGER.info(
        "Chart anchor resolved | start_col_index=%d (%s) gap=%d first_row=%d spacing=%d",
        chart_start_col_index,
        left_anchor_col,
        CHART_COLUMN_GAP,
        first_anchor_row,
        chart_spacing,
    )

    rx_anchor = f"{left_anchor_col}{first_anchor_row}"
    rx_chart.anchor = rx_anchor
    ws.add_chart(rx_chart)

    # position the second chart directly beneath the first with configurable spacing
    bottom_spacing = chart_spacing
    tx_top_row = first_anchor_row + CHART_VERTICAL_HEIGHT_ROWS  - 5 + bottom_spacing
    tx_anchor = f"{left_anchor_col}{tx_top_row}"
    tx_chart.anchor = tx_anchor
    ws.add_chart(tx_chart)

    LOGGER.info(
        "Placed throughput charts | rx_anchor=%s tx_anchor=%s spacing=%d points=%d chart_col=%s",
        rx_anchor,
        tx_anchor,
        chart_spacing,
        point_count,
        left_anchor_col,
    )


# ---------------------------------------------------------------------------
# Data extraction
# ---------------------------------------------------------------------------


def _load_scenario(result_file: Path | str) -> RvrScenario:
    scenario = RvrScenario()
    path = Path(result_file)
    if not path.exists():
        LOGGER.warning("Result CSV not found for RVR report: %s", path)
        return scenario

    attenuation_values: set[float] = set()
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                if not row:
                    continue
                scenario.freq = row.get("Freq_Band") or scenario.freq
                scenario.standard = row.get("Standard") or scenario.standard
                scenario.bandwidth = _format_bandwidth(row.get("BW") or scenario.bandwidth)
                scenario.channel_label = _format_channel(row.get("CH_Freq_MHz") or row.get("Channel") or scenario.channel_label)
                scenario.angle_label = _format_angle(row.get("Angel") or row.get("Angle") or scenario.angle_label)

                attenuation = _sanitize_number(row.get("DB") or row.get("Total_Path_Loss"))
                if attenuation is None:
                    continue
                attenuation_values.add(attenuation)

                throughput = _sanitize_number(row.get("Throughput"))
                direction = str(row.get("Direction") or "").upper()
                if throughput is not None:
                    if direction in {"DL", "RX"}:
                        scenario.rx_values[attenuation] = throughput
                    elif direction in {"UL", "TX"}:
                        scenario.tx_values[attenuation] = throughput

                rssi = _sanitize_number(row.get("RSSI"))
                if rssi is not None:
                    if direction in {"DL", "RX"}:
                        scenario.rssi_rx[attenuation] = rssi
                    elif direction in {"UL", "TX"}:
                        scenario.rssi_tx[attenuation] = rssi
    except Exception:
        LOGGER.exception("Failed to parse RVR CSV: %s", path)

    if attenuation_values:
        scenario.attenuation_steps = sorted(attenuation_values)
    return scenario


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def generate_xiaomi_report(
    result_file: Path | str,
    output_path: Path | str,
    forced_test_type: str | None = None,
) -> Path:
    if forced_test_type and forced_test_type.upper() != "RVR":
        LOGGER.info("Forced test type %s ignored (only RVR supported).", forced_test_type)

    scenario = _load_scenario(result_file)
    LOGGER.info(
        "Building Xiaomi-style RVR report | freq=%s | standard=%s | bandwidth=%s | channel=%s",
        scenario.freq,
        scenario.standard,
        scenario.bandwidth,
        scenario.channel_label,
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Coffey RVR"

    _configure_sheet(sheet)
    _write_title(sheet, scenario)
    _write_headers(sheet, scenario)
    start_row = 7
    end_row = _write_data(sheet, scenario, start_row=start_row)
    _add_charts(sheet, scenario, start_row=start_row, end_row=end_row)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    LOGGER.info("RVR report saved to %s", output)
    return output.resolve()


__all__ = ["generate_xiaomi_report", "RvrScenario"]

