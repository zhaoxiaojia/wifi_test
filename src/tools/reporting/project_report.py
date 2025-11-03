from __future__ import annotations

import csv
import logging
import math
from collections import Counter
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Sequence, Tuple, Union

from openpyxl import Workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.legend import Legend
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from src.test.performance import get_rvo_static_db_list, get_rvo_target_rssi_list
from src.tools.performance.rvr_chart_generator import PerformanceRvrChartGenerator
LOGGER = logging.getLogger(__name__)

# Path to embedded report logo placed alongside this module.
REPORT_LOGO_PATH = Path(__file__).with_name("report_logo.png")
try:
    REPORT_LOGO_BYTES = REPORT_LOGO_PATH.read_bytes()
except FileNotFoundError:
    REPORT_LOGO_BYTES = None
    LOGGER.warning("Report logo not found at %s; skipping logo embedding.", REPORT_LOGO_PATH)

# ---------------------------------------------------------------------------
# Style constants derived from Demo.xlsx
# ---------------------------------------------------------------------------

COLUMN_WIDTHS: Dict[str, float] = {
    "A": 13.75,
    "B": 12.625,
    "C": 11.5,
    "D": 15.0,
    "E": 14.875,
    "F": 15.125,
    "G": 18.375,
    "H": 16.125,
    "I": 14.375,
    "J": 15.0,
    "K": 12.25,
    "L": 13.5,
    "M": 12.5,
    "N": 14.0,
    "O": 12.25,
    "P": 12.25,
    "Q": 13.5,
    "R": 11.5,
    "S": 12.0,
    "T": 12.0,
    "U": 12.0,
    "V": 12.0,
    "W": 12.0,
    "X": 12.0,
    "Y": 12.0,
    "Z": 12.0,
    "AA": 12.0,
    "AB": 12.0,
    "AC": 12.0,
    "AD": 12.0,
    "AE": 12.0,
    "AF": 12.0,
    "AG": 12.0,
    "AH": 12.0,
    "AI": 12.0,
    "AJ": 12.0,
    "AK": 12.0,
    "AL": 12.0,
    "AM": 12.0,
    "AN": 12.0,
}

REPORT_LAST_COLUMN = "AN"

ROW_HEIGHT_TITLE = 42.95

CHART_COLUMN_GAP = 1
CHART_TITLE_ROW = 4
CHART_MIN_TOP_ROW = 5
CHART_VERTICAL_HEIGHT_ROWS = 18
POLAR_CHART_HEIGHT_ROWS = 20
POLAR_IMAGE_WIDTH = 525
POLAR_IMAGE_HEIGHT = 400
POLAR_ROW_SPACING = 6

COLOR_BRAND_BLUE = "2D529F"
COLOR_SUBHEADER = "B4C6E7"
COLOR_RATE_PRIMARY = "FFF2CC"
COLOR_RATE_SECONDARY = "D9E1F2"
COLOR_HEADER_GENERAL = "E7E6E6"
COLOR_RSSI_RX = "CFE2F3"
COLOR_RSSI_TX = "E2F0D9"
COLOR_GRIDLINE = "BFBFBF"
SERIES_COLORS = [
    "2D529F",  # brand blue
    "FF8C00",  # orange
    "2CA02C",  # green
    "9467BD",  # purple
    "D62728",  # red
    "17BECF",  # teal
    "8C564B",  # brown
]

FONT_TITLE = Font(name="Arial", color="FFFFFF", bold=True, size=20)
FONT_SECTION = Font(name="Arial", color="1F4E78", bold=True, size=16)
FONT_HEADER = Font(name="Arial", color="FFFFFF", bold=True, size=11)
FONT_SUBHEADER = Font(name="Arial", color="1F4E78", bold=True, size=11)
FONT_HEADER_DARK = Font(name="Arial", color="333333", bold=True, size=11)
FONT_BODY = Font(name="Arial", color="333333", size=11)
FONT_STANDARD = Font(name="Arial", color="333333", size=11)

ALIGN_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

BORDER_THIN = Border(
    left=Side(style="thin", color="8A8A8A"),
    right=Side(style="thin", color="8A8A8A"),
    top=Side(style="thin", color="8A8A8A"),
    bottom=Side(style="thin", color="8A8A8A"),
)

BORDER_THIN_DIAGONAL_DOWN = Border(
    left=Side(style="thin", color="8A8A8A"),
    right=Side(style="thin", color="8A8A8A"),
    top=Side(style="thin", color="8A8A8A"),
    bottom=Side(style="thin", color="8A8A8A"),
    diagonal=Side(style="thin", color="8A8A8A"),
    diagonalDown=True,
)

BORDER_THIN_DIAGONAL_UP = Border(
    left=Side(style="thin", color="8A8A8A"),
    right=Side(style="thin", color="8A8A8A"),
    top=Side(style="thin", color="8A8A8A"),
    bottom=Side(style="thin", color="8A8A8A"),
    diagonal=Side(style="thin", color="8A8A8A"),
    diagonalUp=True,
)


# ---------------------------------------------------------------------------
# Scenario model
# ---------------------------------------------------------------------------


DEFAULT_ATTENUATIONS = list(range(0, 78, 3))


@dataclass
class ProjectScenario:
    key: str = "SCENARIO|DEFAULT"
    freq: str = "5G"
    standard: str = "Auto"
    bandwidth: str = "20/40/80 MHz"
    channel_label: str = "CH36"
    angle_label: str = "0deg"
    attenuation_steps: List[float] = field(default_factory=lambda: DEFAULT_ATTENUATIONS.copy())
    step_summary: str = ""
    rx_values: Dict[float, float] = field(default_factory=dict)
    tx_values: Dict[float, float] = field(default_factory=dict)
    rssi_rx: Dict[float, float] = field(default_factory=dict)
    rssi_tx: Dict[float, float] = field(default_factory=dict)
    angle_order: List[str] = field(default_factory=list)
    angle_rx_matrix: Dict[float, Dict[str, float]] = field(default_factory=dict)
    angle_tx_matrix: Dict[float, Dict[str, float]] = field(default_factory=dict)
    angle_rssi_rx_matrix: Dict[float, Dict[str, float]] = field(default_factory=dict)
    angle_rssi_tx_matrix: Dict[float, Dict[str, float]] = field(default_factory=dict)

    @property
    def title(self) -> str:
        return f"{self.freq.upper()} {self.standard.upper()} {self.bandwidth.upper()}"

    @property
    def subtitle(self) -> str:
        parts = [self.freq, self.standard.upper(), self.bandwidth]
        return " ".join(part for part in parts if part)

    @property
    def channel(self) -> str:
        return self.channel_label


@dataclass
class ScenarioGroup:
    key: str
    freq: str
    standard: str
    bandwidth: str
    angle_label: str
    attenuation_steps: List[float] = field(default_factory=list)
    step_summary: str = ""
    channels: List[ProjectScenario] = field(default_factory=list)
    raw_keys: set[str] = field(default_factory=set)
    test_type: str = "RVR"
    angle_order: List[str] = field(default_factory=list)

    @property
    def title(self) -> str:
        return f"{self.freq.upper()} {self.standard.upper()} {self.bandwidth.upper()}"

    @property
    def summary_label(self) -> str:
        parts = [self.freq, self.standard.upper(), self.bandwidth]
        return " ".join(part for part in parts if part)


@dataclass
class GroupLayout:
    rx_cols: List[int]
    tx_cols: List[int]
    aml_standard_col: int
    aml_result_col: int
    rssi_start_col: int
    right_rx_cols: List[int]
    right_tx_cols: List[int]



def _normalize_scenario_key(value: Optional[str]) -> str:
    text = str(value or "").strip()
    if not text:
        return "SCENARIO|DEFAULT"
    return text


def _group_base_key(raw_key: Optional[str]) -> str:
    normalized = _normalize_scenario_key(raw_key)
    if not normalized:
        return "SCENARIO|DEFAULT"
    parts = normalized.split("|")
    filtered = [part for part in parts if not part.upper().startswith("CHANNEL=")]
    if not filtered:
        return normalized
    return "|".join(filtered)


def _normalize_test_type(value: Optional[str]) -> str:
    """Normalize the provided test type string.

    Empty inputs intentionally return an empty string so the caller can decide
    whether to fall back to a default (usually "RVR").
    """

    text = str(value or "").strip()
    if not text:
        return ""
    return text.upper()


def _detect_row_test_type(row: dict[str, object]) -> str:
    candidates = (
        row.get("__test_type_display__"),
        row.get("Test_Type"),
        row.get("Test Type"),
        row.get("TestType"),
        row.get("Test"),
    )
    for candidate in candidates:
        text = str(candidate or "").strip()
        if not text:
            continue
        normalized = _normalize_test_type(text)
        if normalized:
            return normalized
    has_profile = any(
        str(row.get(name) or "").strip()
        for name in ("Profile_Mode", "Profile Mode", "RVO_Profile_Mode")
    )
    if has_profile:
        return "RVO"
    return "RVR"


def _build_group_layout(channel_count: int) -> GroupLayout:
    base_col = 3  # Column C
    rx_cols: list[int] = [base_col + index for index in range(channel_count)]
    tx_cols: list[int] = [base_col + channel_count + index for index in range(channel_count)]

    aml_standard_col = base_col + channel_count * 2
    aml_result_col = aml_standard_col + 1

    rssi_start_col = aml_result_col + 1

    right_rx_cols: list[int] = [
        rssi_start_col + index for index in range(channel_count)
    ]
    right_tx_cols: list[int] = [
        rssi_start_col + channel_count + index for index in range(channel_count)
    ]

    return GroupLayout(
        rx_cols=rx_cols,
        tx_cols=tx_cols,
        aml_standard_col=aml_standard_col,
        aml_result_col=aml_result_col,
        rssi_start_col=rssi_start_col,
        right_rx_cols=right_rx_cols,
        right_tx_cols=right_tx_cols,
    )


def _most_common(counter: Counter[str], default: str) -> str:
    for candidate, _count in counter.most_common():
        text = str(candidate or "").strip()
        if text:
            return text
    return default


def _format_numeric_label(value: Optional[float]) -> str:
    if value is None:
        return ""
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return ""
    if math.isclose(numeric, round(numeric), abs_tol=1e-6):
        return str(int(round(numeric)))
    return f"{numeric:.2f}".rstrip('0').rstrip('.')


def _summarize_attenuation_steps(values: Sequence[float]) -> str:
    if not values:
        return ""
    unique_values = sorted({float(v) for v in values})
    if not unique_values:
        return ""
    start = unique_values[0]
    end = unique_values[-1]
    if len(unique_values) == 1:
        return f"{_format_numeric_label(start)} dB"
    deltas = [
        round(unique_values[idx + 1] - unique_values[idx], 4)
        for idx in range(len(unique_values) - 1)
    ]
    positive_deltas = [delta for delta in deltas if delta > 0]
    step_value = 0.0
    if positive_deltas:
        delta_counter = Counter(positive_deltas)
        step_value = float(delta_counter.most_common(1)[0][0])
    start_label = _format_numeric_label(start)
    end_label = _format_numeric_label(end)
    if step_value > 0:
        step_label = _format_numeric_label(step_value)
        return f"{start_label} - {end_label} (step {step_label} dB)"
    return f"{start_label} - {end_label}"


def _resolve_rvo_att_steps() -> List[Tuple[Optional[float], str, str]]:
    """Return RVO attenuation display entries as (value, item_label, att_label)."""

    def _safe_values(loader) -> list[Optional[int]]:
        try:
            return [value for value in loader() if value is not None]
        except Exception:
            LOGGER.exception("Failed to load RVO configuration values from %s", loader)
            return []

    static_values = _safe_values(get_rvo_static_db_list)
    target_values = _safe_values(get_rvo_target_rssi_list)

    if static_values and target_values:
        LOGGER.warning(
            "Both static_db and target_rssi configured; prefer target_rssi for ATT rows."
        )
        static_values = []

    entries: List[Tuple[Optional[float], str, str]] = []
    if target_values:
        for value in target_values:
            numeric = float(value) if value is not None else None
            if value is not None:
                label = f"target rssi : RSSI {value} dBm"
                att_text = f"RSSI {value} dBm"
            else:
                label = "target rssi"
                att_text = "RSSI"
            entries.append((numeric, label, att_text))
    elif static_values:
        for value in static_values:
            numeric = float(value) if value is not None else None
            if value is not None:
                label = f"static db : ATT {value} dB"
                att_text = f"ATT {value} dB"
            else:
                label = "static db"
                att_text = "ATT"
            entries.append((numeric, label, att_text))

    return entries

def _sorted_unique_angles(angles: Sequence[str]) -> List[str]:
    """Return unique angles sorted numerically when possible."""

    def _angle_key(text: str) -> Tuple[int, float, str]:
        normalized = str(text or "").strip().lower().replace("°", "").replace("deg", "")
        try:
            numeric = float(normalized)
            return (0, numeric, text)
        except (TypeError, ValueError):
            return (1, 0.0, text)

    seen = set()
    ordered: List[str] = []
    for angle in angles:
        if angle not in seen:
            ordered.append(angle)
            seen.add(angle)
    return sorted(ordered, key=_angle_key)


def _prepare_rvo_table_entries(
    group: ScenarioGroup,
    override_steps: Sequence[Optional[float]],
    att_labels: Sequence[str],
    item_labels: Sequence[str],
) -> tuple[List[str], List[dict[str, object]]]:
    """Build the axis definition for an RVO matrix table."""

    # Aggregate candidate angles from group and channel scopes.
    angle_candidates: List[str] = []
    if group.angle_order:
        angle_candidates.extend(group.angle_order)
    for scenario in group.channels:
        angle_candidates.extend(scenario.angle_order)
    angles = _sorted_unique_angles(angle_candidates)

    # Determine attenuation steps (vertical axis) in preferred order.
    if override_steps:
        steps: List[Optional[float]] = list(override_steps)
    else:
        collected: List[float] = []
        for scenario in group.channels:
            collected.extend(scenario.attenuation_steps)
        unique = sorted({float(step) for step in collected}) if collected else []
        steps = [float(step) for step in unique]

    entries: List[dict[str, object]] = []
    for scenario in group.channels:
        scenario_key = getattr(scenario, "key", "")
        if not scenario_key:
            LOGGER.warning(
                "Skipped RVO entry due to missing scenario key | group=%s", group.key
            )
            continue

        for index, step in enumerate(steps):
            numeric = float(step) if isinstance(step, (int, float)) else None
            label_idx = index if override_steps else index
            att_display = ""
            item_display = ""
            if override_steps:
                if label_idx < len(att_labels):
                    att_display = att_labels[label_idx]
                if label_idx < len(item_labels):
                    item_display = item_labels[label_idx]
                if not att_display and numeric is not None:
                    att_display = f"{_format_numeric_label(numeric)} dB"
                if not item_display and att_display:
                    item_display = att_display
            else:
                att_display = f"{_format_numeric_label(numeric)} dB" if numeric is not None else ""
                item_display = att_display
            entries.append(
                {
                    "scenario_key": scenario_key,
                    "attenuation": numeric,
                    "att_display": att_display,
                    "item_display": item_display,
                }
            )

    return angles, entries


def _estimate_rvo_last_column(groups: Sequence[ScenarioGroup], rvo_att_entries: Sequence[Tuple[Optional[float], str, str]]) -> int:
    if not groups:
        return column_index_from_string(REPORT_LAST_COLUMN)

    override_steps = [entry[0] for entry in rvo_att_entries]
    att_labels = [entry[2] for entry in rvo_att_entries]
    item_labels = [entry[1] for entry in rvo_att_entries]

    max_col = 8
    for group in groups:
        angles, entries = _prepare_rvo_table_entries(
            group,
            override_steps,
            att_labels,
            item_labels,
        )
        angle_headers = list(angles) if angles else [group.angle_label or "0deg"]
        angle_count = len(angle_headers)
        throughput_start_col = 4
        throughput_end_col = throughput_start_col + angle_count - 1 if angle_count else throughput_start_col - 1
        average_col = throughput_end_col + 1
        ovality_col = average_col + 1
        aml_standard_col = ovality_col + 1
        aml_result_col = aml_standard_col + 1
        right_att_col = aml_result_col + 1
        rssi_start_col = right_att_col + 1
        if angle_count:
            rssi_end_col = rssi_start_col + angle_count - 1
            used_last_col = rssi_end_col
        else:
            used_last_col = ovality_col
        max_col = max(max_col, used_last_col)
    return max_col


def _estimate_standard_last_column(groups: Sequence[ScenarioGroup]) -> int:
    if not groups:
        return column_index_from_string(REPORT_LAST_COLUMN)
    max_col = 8
    for group in groups:
        layout = _build_group_layout(len(group.channels))
        candidate_cols = [layout.aml_result_col]
        if layout.right_tx_cols:
            candidate_cols.append(layout.right_tx_cols[-1])
        elif layout.tx_cols:
            candidate_cols.append(layout.tx_cols[-1])
        max_col = max(max_col, max(candidate_cols))
    return max_col


def _determine_report_last_column(
    groups: Sequence[ScenarioGroup],
    test_type: str,
    *,
    rvo_att_entries: Sequence[Tuple[Optional[float], str, str]] | None = None,
) -> int:
    if test_type == "RVO":
        entries = rvo_att_entries or []
        return _estimate_rvo_last_column(groups, entries)
    return _estimate_standard_last_column(groups)


def _sanitize_number(value: Optional[str]) -> Optional[float]:
    if value in (None, "", "NULL"):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _format_angle(value: Optional[str]) -> str:
    text = str(value or '').strip()
    if not text:
        return '0deg'
    normalized = text.replace('\u00B0', 'deg').replace('?', '').strip()
    if not normalized:
        return '0deg'
    if normalized.lower().endswith('deg'):
        return normalized
    return f"{normalized}deg"

def _format_bandwidth(value: Optional[str]) -> str:
    text = str(value or "").strip()
    if not text:
        return "20/40/80 MHz"
    normalized = text.replace("_", " ").replace("MHz", "").strip()
    if normalized.lower().endswith("mhz"):
        return normalized
    return f"{normalized} MHz"


def _format_channel(value: Optional[str]) -> str:
    text = str(value or "").strip()
    if not text:
        return "CH36"
    upper = text.upper()
    if upper.startswith("CH"):
        return text.title()
    if upper.replace(".", "").isdigit():
        return f"CH{upper}"
    return f"CH {text.title()}"



# ---------------------------------------------------------------------------
# Utility functions
# ---------------------------------------------------------------------------


def _configure_sheet(ws: Worksheet) -> None:
    for column, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[column].width = width


def _merge(ws: Worksheet, range_string: str) -> None:
    ws.merge_cells(range_string)


def _set_cell(
    ws: Worksheet,
    row: int,
    column: int,
    value,
    font: Optional[Font] = None,
    alignment: Optional[Alignment] = None,
    fill: Optional[str] = None,
    border: Union[bool, Border] = False,
    number_format: Optional[str] = None,
) -> None:
    cell = ws.cell(row=row, column=column, value=value)
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    if border:
        cell.border = BORDER_THIN if border is True else border
    if number_format:
        cell.number_format = number_format


# ---------------------------------------------------------------------------
# Standard text helpers
# ---------------------------------------------------------------------------


def _aml_standard_text(att: float) -> Optional[str]:
    if att <= 15:
        return "RX Tput>=503.5\nTX Tput>=475"
    if att <= 21:
        return "RX Tput>=320\nTX Tput>=300"
    if 30 <= att <= 45:
        return "RX Tput>100\nTX Tput>95"
    if att >= 48:
        return "N/A"


def _rvo_standard_text(att: float) -> Optional[str]:
    if not math.isfinite(att):
        return None
    if att <= -55:
        return "MIN Tput: 63dB>0"
    return "AVG Tput: 13dB≥85.5\n33dB≥66.5\n53dB≥30"
def _aml_threshold(att: float) -> Optional[Tuple[int, int]]:
    if att <= 12:
        return 400, 300
    if att <= 21:
        return 320, 200
    if att <= 45:
        return 100, 80
    return None


def _group_runs(values: Sequence[float], formatter) -> List[Tuple[int, int, Optional[str]]]:
    if not values:
        return []
    groups: List[Tuple[int, int, Optional[str]]] = []
    start = 0
    current = formatter(values[0])
    for idx in range(1, len(values)):
        text = formatter(values[idx])
        if text != current:
            groups.append((start, idx - 1, current))
            start = idx
            current = text
    groups.append((start, len(values) - 1, current))
    return groups


def _apply_grouped_texts(ws: Worksheet, column: int, start_row: int, values: Sequence[float], formatter) -> None:
    for start_idx, end_idx, text in _group_runs(values, formatter):
        if not text:
            continue
        top = start_row + start_idx
        bottom = start_row + end_idx
        if top != bottom:
            _merge(ws, f"{get_column_letter(column)}{top}:{get_column_letter(column)}{bottom}")
        _set_cell(ws, top, column, text, font=FONT_STANDARD, alignment=ALIGN_LEFT_WRAP, border=True)
        for r in range(top + 1, bottom + 1):
            cell = ws.cell(row=r, column=column)
            cell.border = BORDER_THIN
            cell.alignment = ALIGN_LEFT_WRAP


def _set_result_cell(ws: Worksheet, row: int, column: int, threshold: Optional[Tuple[int, int]], rx_column: int, tx_column: int) -> None:
    if threshold is None:
        value = "Pass"
    else:
        rx, tx = threshold
        rx_letter = get_column_letter(rx_column)
        tx_letter = get_column_letter(tx_column)
        value = f'=IF(AND({rx_letter}{row}>{rx},{tx_letter}{row}>{tx}),"Pass","Fail")'
    _set_cell(ws, row, column, value, font=FONT_BODY, alignment=ALIGN_CENTER, border=True)


def _apply_result_formatting(ws: Worksheet, start_row: int, end_row: int, result_column: int) -> None:
    if start_row > end_row:
        return
    column_letter = get_column_letter(result_column)
    fail_rule = FormulaRule(
        formula=[f"={column_letter}{start_row}=\"Fail\""],
        font=Font(name="Arial", color="FF0000", bold=True),
    )
    pass_rule = FormulaRule(
        formula=[f"={column_letter}{start_row}=\"Pass\""],
        font=Font(name="Arial", color="008000", bold=True),
    )
    ws.conditional_formatting.add(f"{column_letter}{start_row}:{column_letter}{end_row}", fail_rule)
    ws.conditional_formatting.add(f"{column_letter}{start_row}:{column_letter}{end_row}", pass_rule)


# ---------------------------------------------------------------------------
# Layout writer
# ---------------------------------------------------------------------------


TitleSummaryBuilder = Callable[[Sequence[ScenarioGroup]], str]


def _build_frequency_summary(
    groups: Sequence[ScenarioGroup], *, label: Optional[str] = None
) -> str:
    if not groups:
        return f"{label or 'Frequency Summary'}: None"

    frequency_order: list[str] = []
    frequency_map: dict[str, Counter[str]] = {}

    for group in groups:
        freq = (group.freq or "").strip().upper() or "UNKNOWN"
        if freq not in frequency_map:
            frequency_map[freq] = Counter()
            frequency_order.append(freq)

        detail_parts: list[str] = []
        if group.standard:
            detail_parts.append(group.standard.upper())
        if group.bandwidth:
            detail_parts.append(group.bandwidth)
        detail_label = " ".join(part for part in detail_parts if part) or "UNKNOWN"

        frequency_map[freq][detail_label] += len(group.channels)

    segments: list[str] = []
    for freq in frequency_order:
        detail_counter = frequency_map[freq]
        total = sum(detail_counter.values())
        if not detail_counter:
            segments.append(f"{freq}: {total} scenarios")
            continue

        detail_texts = [f"{detail_label} x{count}" for detail_label, count in sorted(detail_counter.items())]
        segments.append(f"{freq}: {total} scenarios ({'; '.join(detail_texts)})")

    summary = " / ".join(segments)
    return f"{label}: {summary}" if label else summary

def _build_throughput_title_summary(groups: Sequence[ScenarioGroup]) -> str:
    if not groups:
        return "1、Throughput:None"

    frequency_order: list[str] = []
    frequency_details: dict[str, dict[str, int]] = {}

    for group in groups:
        freq = (group.freq or "").strip()
        freq_label = freq.upper() if freq else "UNKNOWN"
        if freq_label not in frequency_details:
            frequency_details[freq_label] = {}
            frequency_order.append(freq_label)

        detail_map = frequency_details[freq_label]
        summary_label = group.summary_label or freq_label
        scenario_count = len(group.channels) or 1
        detail_map[summary_label] = detail_map.get(summary_label, 0) + scenario_count

    summary_lines: list[str] = []
    for index, freq_label in enumerate(frequency_order, start=1):
        summary_lines.append(f"{index}、Throughput:{freq_label}")
        detail_map = frequency_details[freq_label]
        if detail_map:
            detail_segments: list[str] = []
            for label, count in detail_map.items():
                if count == 1:
                    detail_segments.append(label)
                else:
                    detail_segments.append(f"{label} x{count}")
            summary_lines.append(" / ".join(detail_segments))

    return "\n".join(summary_lines)


def _build_rvo_title_summary(groups: Sequence[ScenarioGroup]) -> str:
    return _build_throughput_title_summary(groups)


def _build_rvr_title_summary(groups: Sequence[ScenarioGroup]) -> str:
    return _build_throughput_title_summary(groups)


_TITLE_SUMMARY_BUILDERS: Dict[str, TitleSummaryBuilder] = {
    "RVO": _build_rvo_title_summary,
    "RVR": _build_rvr_title_summary,
}


def _resolve_title_summary(groups: Sequence[ScenarioGroup], test_type: str) -> str:
    builder = _TITLE_SUMMARY_BUILDERS.get(test_type.upper(), _build_frequency_summary)
    return builder(groups)


def _write_report_title(
    ws: Worksheet,
    *,
    groups: Sequence[ScenarioGroup],
    test_type: str,
    start_row: int = 1,
    remarks: Optional[str] = None,
    last_column: str,
) -> int:
    top_row = start_row
    last_col = last_column or REPORT_LAST_COLUMN
    _merge(ws, f"A{top_row}:{last_col}{top_row}")
    ws.row_dimensions[top_row].height = ROW_HEIGHT_TITLE
    title_text = f"WiFi {test_type.upper()} Test Report"
    _set_cell(
        ws,
        top_row,
        1,
        title_text,
        font=FONT_TITLE,
        alignment=ALIGN_CENTER,
        fill=COLOR_BRAND_BLUE,
        border=True,
    )

    if REPORT_LOGO_BYTES:
        image = Image(BytesIO(REPORT_LOGO_BYTES))
        image.width = 240
        image.height = 70
        image.anchor = f"A{top_row}"
        ws.add_image(image)

    remark_row = top_row + 1
    _merge(ws, f"A{remark_row}:{last_col}{remark_row}")
    default_remark = "Remarks:Ovality=Min Tup/AVG Tup*100%"
    remark_text = default_remark if remarks is None else remarks
    _set_cell(
        ws,
        remark_row,
        1,
        remark_text,
        font=FONT_SUBHEADER,
        alignment=ALIGN_CENTER_WRAP,
        border=True,
    )

    summary_row = remark_row + 1
    _merge(ws, f"A{summary_row}:{last_col}{summary_row}")
    summary_text = _resolve_title_summary(groups, test_type)
    _set_cell(
        ws,
        summary_row,
        1,
        summary_text,
        font=FONT_BODY,
        alignment=ALIGN_CENTER_WRAP,
        border=True,
    )

    LOGGER.info(
        "Report title written | top_row=%d remark_row=%d summary_row=%d", top_row, remark_row, summary_row
    )
    return summary_row


def _write_group_header(ws: Worksheet, group: ScenarioGroup, start_row: int, *, last_column: str) -> int:
    last_col = last_column or REPORT_LAST_COLUMN
    _merge(ws, f"A{start_row}:{last_col}{start_row}")
    _set_cell(
        ws,
        start_row,
        1,
        group.summary_label,
        font=FONT_SECTION,
        alignment=ALIGN_CENTER,
        border=True,
    )
    LOGGER.info("Group header written | row=%d label=%s", start_row, group.summary_label)
    return start_row


def _write_headers(
    ws: Worksheet,
    group: ScenarioGroup,
    channels: Sequence[ProjectScenario],
    layout: GroupLayout,
    *,
    header_row: int,
) -> int:
    sub_row = header_row + 1
    step_text = group.step_summary or ""
    base_headers: list[tuple[int, int, str]] = [
        (header_row, 1, "Item"),
        (header_row, 2, "ATT\n(Unit:dB)"),
        (header_row, layout.aml_standard_col, "AML_Standard"),
        (header_row, layout.aml_result_col, "AML_Result"),
    ]
    for row, col, text in base_headers:
        _set_cell(
            ws,
            row,
            col,
            text,
            font=FONT_HEADER,
            alignment=ALIGN_CENTER_WRAP,
            fill=COLOR_BRAND_BLUE,
            border=True,
        )

    def _write_header_block(columns: Sequence[int], label: str) -> None:
        if not columns:
            return
        start = columns[0]
        end = columns[-1]
        if end > start:
            start_letter = get_column_letter(start)
            end_letter = get_column_letter(end)
            _merge(ws, f"{start_letter}{header_row}:{end_letter}{header_row}")
        _set_cell(
            ws,
            header_row,
            start,
            label,
            font=FONT_HEADER,
            alignment=ALIGN_CENTER_WRAP,
            fill=COLOR_BRAND_BLUE,
            border=True,
        )

    _write_header_block(layout.rx_cols, "RX(Unit:Mbps)")
    _write_header_block(layout.tx_cols, "TX(Unit:Mbps)")

    rssi_columns = layout.right_rx_cols + layout.right_tx_cols
    if rssi_columns:
        _write_header_block(rssi_columns, "RSSI")
    else:
        _set_cell(
            ws,
            header_row,
            layout.rssi_start_col,
            "RSSI",
            font=FONT_HEADER,
            alignment=ALIGN_CENTER_WRAP,
            fill=COLOR_BRAND_BLUE,
            border=True,
        )

    _set_cell(
        ws,
        sub_row,
        1,
        step_text,
        font=FONT_SUBHEADER,
        alignment=ALIGN_CENTER,
        fill=COLOR_SUBHEADER,
        border=True,
    )
    _set_cell(
        ws,
        sub_row,
        layout.rssi_start_col,
        step_text,
        font=FONT_SUBHEADER,
        alignment=ALIGN_CENTER,
        fill=COLOR_SUBHEADER,
        border=True,
    )

    for idx, scenario in enumerate(channels):
        channel_text = scenario.channel
        if idx < len(layout.rx_cols):
            _set_cell(
                ws,
                sub_row,
                layout.rx_cols[idx],
                channel_text,
                font=FONT_SUBHEADER,
                alignment=ALIGN_CENTER,
                fill=COLOR_SUBHEADER,
                border=True,
            )
        if idx < len(layout.tx_cols):
            _set_cell(
                ws,
                sub_row,
                layout.tx_cols[idx],
                channel_text,
                font=FONT_SUBHEADER,
                alignment=ALIGN_CENTER,
                fill=COLOR_SUBHEADER,
                border=True,
            )

    _set_cell(ws, sub_row, layout.aml_standard_col, None, font=FONT_SUBHEADER, alignment=ALIGN_CENTER, fill=COLOR_SUBHEADER, border=True)
    _set_cell(ws, sub_row, layout.aml_result_col, None, font=FONT_SUBHEADER, alignment=ALIGN_CENTER, fill=COLOR_SUBHEADER, border=True)

    for idx, scenario in enumerate(channels):
        channel_text = scenario.channel
        if idx < len(layout.right_rx_cols):
            _set_cell(
                ws,
                sub_row,
                layout.right_rx_cols[idx],
                channel_text,
                font=FONT_SUBHEADER,
                alignment=ALIGN_CENTER,
                fill=COLOR_SUBHEADER,
                border=True,
            )
        if idx < len(layout.right_tx_cols):
            _set_cell(
                ws,
                sub_row,
                layout.right_tx_cols[idx],
                channel_text,
                font=FONT_SUBHEADER,
                alignment=ALIGN_CENTER,
                fill=COLOR_SUBHEADER,
                border=True,
            )
    LOGGER.info(
        "Header row configured | start_row=%d channel_count=%d step_summary=%s",
        header_row,
        len(channels),
        step_text or "N/A",
    )
    return sub_row


def _write_data(
    ws: Worksheet,
    group: ScenarioGroup,
    channels: Sequence[ProjectScenario],
    layout: GroupLayout,
    start_row: int = 7,
    *,
    override_steps: Optional[Sequence[Optional[float]]] = None,
    att_labels: Optional[Sequence[str]] = None,
    item_labels: Optional[Sequence[str]] = None,
) -> int:
    base_steps = list(group.attenuation_steps or DEFAULT_ATTENUATIONS)
    if override_steps is not None:
        override = [step for step in override_steps if step is not None]
        steps = list(override) if override else base_steps
    else:
        steps = base_steps

    if not steps:
        steps = DEFAULT_ATTENUATIONS.copy()

    step_count = len(steps)
    end_row = start_row + step_count - 1

    if steps and not item_labels:
        _merge(ws, f"A{start_row}:A{end_row}")
    _set_cell(
        ws,
        start_row,
        1,
        group.summary_label,
        font=FONT_HEADER,
        alignment=ALIGN_CENTER,
        fill=COLOR_BRAND_BLUE,
        border=True,
    )

    for index, attenuation in enumerate(steps):
        row = start_row + index
        highlight = COLOR_RATE_PRIMARY if index < 6 else COLOR_RATE_SECONDARY

        if item_labels:
            label = item_labels[index] if index < len(item_labels) else ""
            _set_cell(ws, row, 1, label, font=FONT_BODY, alignment=ALIGN_CENTER_WRAP, border=True)

        att_value = att_labels[index] if att_labels and index < len(att_labels) else attenuation
        _set_cell(ws, row, 2, att_value, font=FONT_BODY, alignment=ALIGN_CENTER, border=True)

        for channel_idx, scenario in enumerate(channels):
            rx_col = layout.rx_cols[channel_idx]
            tx_col = layout.tx_cols[channel_idx]
            rx_value = scenario.rx_values.get(attenuation)
            tx_value = scenario.tx_values.get(attenuation)
            _set_cell(
                ws,
                row,
                rx_col,
                rx_value,
                font=FONT_BODY,
                alignment=ALIGN_CENTER,
                fill=highlight,
                border=True,
            )
            _set_cell(
                ws,
                row,
                tx_col,
                tx_value,
                font=FONT_BODY,
                alignment=ALIGN_CENTER,
                fill=highlight,
                border=True,
            )

        _set_cell(ws, row, layout.aml_standard_col, None, font=FONT_STANDARD, alignment=ALIGN_LEFT_WRAP, border=True)
        _set_cell(
            ws,
            row,
            layout.aml_result_col,
            None,
            font=FONT_BODY,
            alignment=ALIGN_CENTER,
            border=True,
        )

        for channel_idx, scenario in enumerate(channels):
            rx_col = layout.right_rx_cols[channel_idx]
            tx_col = layout.right_tx_cols[channel_idx]
            _set_cell(
                ws,
                row,
                rx_col,
                scenario.rssi_rx.get(attenuation),
                font=FONT_BODY,
                alignment=ALIGN_CENTER,
                fill=COLOR_RSSI_RX,
                border=True,
            )
            _set_cell(
                ws,
                row,
                tx_col,
                scenario.rssi_tx.get(attenuation),
                font=FONT_BODY,
                alignment=ALIGN_CENTER,
                fill=COLOR_RSSI_TX,
                border=True,
            )

    # AML 标准与结果暂时留空，后续再补逻辑，因此不应用自动填充/格式化。
    LOGGER.info(
        'Data rows populated | group=%s start_row=%d end_row=%d points=%d channels=%d',
        group.key,
        start_row,
        end_row,
        step_count,
        len(channels),
    )
    return end_row

def _write_rvo_table(
    ws: Worksheet,
    group: ScenarioGroup,
    *,
    start_row: int,
    angles: Sequence[str],
    entries: Sequence[dict[str, object]],
) -> tuple[int, int]:
    """Render an RVO summary table with angle throughput and RSSI blocks."""

    def _display_angle(value: str) -> str:
        text = str(value or "").strip()
        if text.lower().endswith("deg"):
            return text[:-3] + "\u00B0"
        return text

    header_row_rx = start_row
    sub_header_row_rx = header_row_rx + 1
    raw_angle_count = len(angles)
    angle_headers = list(angles) if angles else [group.angle_label or "0deg"]
    angle_count = len(angle_headers)

    item_col = 1
    channel_col = 2
    left_att_col = 3
    throughput_start_col = 4
    throughput_end_col = (
        throughput_start_col + angle_count - 1 if angle_count else throughput_start_col - 1
    )
    average_col = throughput_end_col + 1
    ovality_col = average_col + 1
    aml_standard_col = ovality_col + 1
    aml_result_col = aml_standard_col + 1
    right_att_col = aml_result_col + 1
    rssi_start_col = right_att_col + 1
    rssi_end_col = (
        rssi_start_col + angle_count - 1 if angle_count else rssi_start_col - 1
    )

    throughput_start_letter = get_column_letter(throughput_start_col)
    throughput_end_letter = (
        get_column_letter(throughput_end_col) if angle_count else throughput_start_letter
    )
    average_letter = get_column_letter(average_col)
    rssi_start_letter = get_column_letter(rssi_start_col)
    rssi_end_letter = get_column_letter(rssi_end_col) if angle_count else rssi_start_letter

    def _set_widths(columns: Sequence[int], width: float) -> None:
        for col in columns:
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = width

    merged_columns = [
        item_col,
        channel_col,
        left_att_col,
        average_col,
        ovality_col,
        aml_standard_col,
        aml_result_col,
        right_att_col,
    ]
    header_cells_rx = [
        (item_col, "Item"),
        (channel_col, "CH"),
        (left_att_col, "Angle\n\nATT"),
        (average_col, "Average\n(Unit:Mb)"),
        (ovality_col, "Ovality(%)"),
        (aml_standard_col, "AML_Standard"),
        (aml_result_col, "AML_Result"),
        (right_att_col, "Angle\n\nATT"),
    ]
    header_cells_tx = [
        (channel_col, "CH"),
        (left_att_col, "Angle\n\nATT"),
        (average_col, "Average\n(Unit:Mb)"),
        (ovality_col, "Ovality(%)"),
        (aml_standard_col, "AML_Standard"),
        (aml_result_col, "AML_Result"),
        (right_att_col, "Angle\n\nATT"),
    ]

    _set_widths([item_col], 14)
    _set_widths([channel_col], 12)
    _set_widths([left_att_col, right_att_col], 11)
    _set_widths([average_col, ovality_col, aml_result_col], 12)
    _set_widths([aml_standard_col], 22)
    if angle_count:
        _set_widths(range(throughput_start_col, throughput_end_col + 1), 11)
        _set_widths(range(rssi_start_col, rssi_end_col + 1), 11)

    def _merge_header_block(top_row: int) -> None:
        bottom_row = top_row + 1
        for col in merged_columns:
            letter = get_column_letter(col)
            _merge(ws, f"{letter}{top_row}:{letter}{bottom_row}")

    def _write_header_block(top_row: int, *, throughput_title: str, rssi_title: str, include_item: bool) -> None:
        sub_row = top_row + 1
        _merge_header_block(top_row)
        headers = header_cells_rx if include_item else header_cells_tx
        for col, text in headers:
            border_style = (
                BORDER_THIN_DIAGONAL_DOWN
                if col == left_att_col
                else BORDER_THIN_DIAGONAL_UP
                if col == right_att_col
                else True
            )
            _set_cell(
                ws,
                top_row,
                col,
                text,
                font=FONT_HEADER,
                alignment=ALIGN_CENTER_WRAP,
                fill=COLOR_BRAND_BLUE,
                border=border_style,
            )
        if angle_count:
            _merge(ws, f"{throughput_start_letter}{top_row}:{throughput_end_letter}{top_row}")
        _set_cell(
            ws,
            top_row,
            throughput_start_col,
            throughput_title,
            font=FONT_HEADER,
            alignment=ALIGN_CENTER_WRAP,
            fill=COLOR_BRAND_BLUE,
            border=True,
        )
        for idx, angle in enumerate(angle_headers):
            _set_cell(
                ws,
                sub_row,
                throughput_start_col + idx,
                _display_angle(angle),
                font=FONT_SUBHEADER,
                alignment=ALIGN_CENTER,
                fill=COLOR_RATE_PRIMARY,
                border=True,
            )
        if angle_count:
            _merge(ws, f"{rssi_start_letter}{top_row}:{rssi_end_letter}{top_row}")
        _set_cell(
            ws,
            top_row,
            rssi_start_col,
            rssi_title,
            font=FONT_HEADER,
            alignment=ALIGN_CENTER_WRAP,
            fill=COLOR_BRAND_BLUE,
            border=True,
        )
        for idx, angle in enumerate(angle_headers):
            _set_cell(
                ws,
                sub_row,
                rssi_start_col + idx,
                _display_angle(angle),
                font=FONT_SUBHEADER,
                alignment=ALIGN_CENTER,
                fill=COLOR_RATE_SECONDARY,
                border=True,
            )

    _write_header_block(
        header_row_rx,
        throughput_title="RX (Unit:Mbps)",
        rssi_title="RX_RSSI (Unit:dBm)",
        include_item=True,
    )

    rows_by_scenario: Dict[str, List[dict[str, object]]] = {}
    for entry in entries:
        scenario_key = entry.get("scenario_key")
        if isinstance(scenario_key, str) and scenario_key:
            rows_by_scenario.setdefault(scenario_key, []).append(entry)

    def _attenuation_value(entry: dict[str, object]) -> Optional[float]:
        value = entry.get("attenuation")
        if isinstance(value, (int, float)):
            return float(value)
        try:
            return float(value)  # type: ignore[arg-type]
        except (TypeError, ValueError):
            return None

    def _att_display_text(entry: dict[str, object]) -> str:
        return str(entry.get("att_display") or "")

    def _item_label_rx(entry: dict[str, object], scenario: ProjectScenario) -> str:
        primary = str(entry.get("item_display") or "").strip()
        if primary:
            return primary
        secondary = str(entry.get("att_display") or "").strip()
        if secondary:
            return secondary
        return scenario.channel or ""

    def _item_label_tx(entry: dict[str, object], fallback: str) -> str:
        for key in ("item_display_tx", "item_display_rx"):
            candidate = str(entry.get(key) or "").strip()
            if candidate:
                return candidate
        return fallback

    def _resolve_angle_map(matrix: Dict[float, Dict[str, float]], lookup: Optional[float]) -> Dict[str, float]:
        if lookup is not None:
            if lookup in matrix:
                return matrix[lookup]
            for key in matrix:
                if math.isfinite(key) and math.isfinite(lookup) and abs(key - lookup) < 1e-6:
                    return matrix[key]
        if matrix:
            first_key = next(iter(matrix))
            return matrix[first_key]
        return {}

    def _resolve_entry_context(scenario: ProjectScenario, entry: dict[str, object]) -> dict[str, object]:
        lookup = _attenuation_value(entry)
        att_display = _att_display_text(entry)
        item_rx = _item_label_rx(entry, scenario)
        item_tx = _item_label_tx(entry, item_rx)
        rx_angles = _resolve_angle_map(scenario.angle_rx_matrix, lookup)
        tx_angles = _resolve_angle_map(scenario.angle_tx_matrix, lookup)
        rssi_rx_angles = _resolve_angle_map(scenario.angle_rssi_rx_matrix, lookup)
        rssi_tx_angles = _resolve_angle_map(scenario.angle_rssi_tx_matrix, lookup)
        standard_text = _rvo_standard_text(lookup) if isinstance(lookup, float) else None
        return {
            "lookup": lookup,
            "att": att_display,
            "item_rx": item_rx,
            "item_tx": item_tx,
            "rx_angles": rx_angles,
            "tx_angles": tx_angles,
            "rssi_rx": rssi_rx_angles,
            "rssi_tx": rssi_tx_angles,
            "standard": standard_text,
        }

    def _write_rx_row(row_index: int, scenario: ProjectScenario, ctx: dict[str, object]) -> None:
        att_display = ctx["att"]  # type: ignore[assignment]
        item_label = ctx["item_rx"]  # type: ignore[assignment]
        rx_angle_map = ctx["rx_angles"]  # type: ignore[assignment]
        rssi_rx_map = ctx["rssi_rx"]  # type: ignore[assignment]
        standard_text = ctx["standard"]  # type: ignore[assignment]

        _set_cell(
            ws,
            row_index,
            channel_col,
            scenario.channel,
            font=FONT_BODY,
            alignment=ALIGN_CENTER,
            border=True,
        )
        _set_cell(
            ws,
            row_index,
            left_att_col,
            att_display,
            font=FONT_BODY,
            alignment=ALIGN_CENTER,
            border=True,
        )
        for idx, angle in enumerate(angle_headers):
            col = throughput_start_col + idx
            _set_cell(
                ws,
                row_index,
                col,
                rx_angle_map.get(angle),
                font=FONT_BODY,
                alignment=ALIGN_CENTER,
                fill=COLOR_RATE_PRIMARY,
                border=True,
            )
        if angle_count:
            throughput_range_rx = f"{throughput_start_letter}{row_index}:{throughput_end_letter}{row_index}"
            rx_average_formula = f"=AVERAGE({throughput_range_rx})"
            rx_ovality_formula = f"=MIN({throughput_range_rx})/{average_letter}{row_index}"
        else:
            rx_average_formula = None
            rx_ovality_formula = None
        _set_cell(
            ws,
            row_index,
            average_col,
            rx_average_formula,
            font=FONT_BODY,
            alignment=ALIGN_CENTER,
            border=True,
            number_format="0.00",
        )
        _set_cell(
            ws,
            row_index,
            ovality_col,
            rx_ovality_formula,
            font=FONT_BODY,
            alignment=ALIGN_CENTER,
            border=True,
            number_format="0.00%",
        )
        _set_cell(
            ws,
            row_index,
            aml_standard_col,
            None,
            font=FONT_STANDARD,
            alignment=ALIGN_LEFT_WRAP,
            border=True,
        )
        _set_cell(
            ws,
            row_index,
            aml_result_col,
            None,
            font=FONT_BODY,
            alignment=ALIGN_CENTER,
            border=True,
        )
        _set_cell(
            ws,
            row_index,
            right_att_col,
            att_display,
            font=FONT_BODY,
            alignment=ALIGN_CENTER,
            border=True,
        )
        for idx, angle in enumerate(angle_headers):
            col = rssi_start_col + idx
            _set_cell(
                ws,
                row_index,
                col,
                rssi_rx_map.get(angle),
                font=FONT_BODY,
                alignment=ALIGN_CENTER,
                fill=COLOR_RSSI_RX,
                border=True,
            )

    def _write_tx_row(row_index: int, scenario: ProjectScenario, ctx: dict[str, object]) -> None:
        att_display = ctx["att"]  # type: ignore[assignment]
        item_label = ctx["item_tx"]  # type: ignore[assignment]
        tx_angle_map = ctx["tx_angles"]  # type: ignore[assignment]
        rssi_tx_map = ctx["rssi_tx"]  # type: ignore[assignment]
        standard_text = ctx["standard"]  # type: ignore[assignment]

        _set_cell(
            ws,
            row_index,
            channel_col,
            scenario.channel,
            font=FONT_BODY,
            alignment=ALIGN_CENTER,
            border=True,
        )
        _set_cell(
            ws,
            row_index,
            left_att_col,
            att_display,
            font=FONT_BODY,
            alignment=ALIGN_CENTER,
            border=True,
        )
        for idx, angle in enumerate(angle_headers):
            col = throughput_start_col + idx
            _set_cell(
                ws,
                row_index,
                col,
                tx_angle_map.get(angle),
                font=FONT_BODY,
                alignment=ALIGN_CENTER,
                fill=COLOR_RATE_PRIMARY,
                border=True,
            )
        if angle_count:
            throughput_range_tx = f"{throughput_start_letter}{row_index}:{throughput_end_letter}{row_index}"
            tx_average_formula = f"=AVERAGE({throughput_range_tx})"
            tx_ovality_formula = f"=MIN({throughput_range_tx})/{average_letter}{row_index}"
        else:
            tx_average_formula = None
            tx_ovality_formula = None
        _set_cell(
            ws,
            row_index,
            average_col,
            tx_average_formula,
            font=FONT_BODY,
            alignment=ALIGN_CENTER,
            border=True,
            number_format="0.00",
        )
        _set_cell(
            ws,
            row_index,
            ovality_col,
            tx_ovality_formula,
            font=FONT_BODY,
            alignment=ALIGN_CENTER,
            border=True,
            number_format="0.00%",
        )
        _set_cell(
            ws,
            row_index,
            aml_standard_col,
            None,
            font=FONT_STANDARD,
            alignment=ALIGN_LEFT_WRAP,
            border=True,
        )
        _set_cell(
            ws,
            row_index,
            aml_result_col,
            None,
            font=FONT_BODY,
            alignment=ALIGN_CENTER,
            border=True,
        )
        _set_cell(
            ws,
            row_index,
            right_att_col,
            att_display,
            font=FONT_BODY,
            alignment=ALIGN_CENTER,
            border=True,
        )
        for idx, angle in enumerate(angle_headers):
            col = rssi_start_col + idx
            _set_cell(
                ws,
                row_index,
                col,
                rssi_tx_map.get(angle),
                font=FONT_BODY,
                alignment=ALIGN_CENTER,
                fill=COLOR_RSSI_TX,
                border=True,
            )

    data_start_row = sub_header_row_rx + 1
    current_row = data_start_row
    first_data_row: Optional[int] = None
    last_data_row: Optional[int] = None
    scenario_contexts: Dict[str, List[dict[str, object]]] = {}
    for scenario in group.channels:
        scenario_key = getattr(scenario, "key", "")
        scenario_rows = rows_by_scenario.get(scenario_key, []) if scenario_key else []
        if not scenario_rows:
            continue
        context_list: List[dict[str, object]] = []
        scenario_contexts[scenario_key] = context_list
        scenario_start = current_row
        for row_entry in scenario_rows:
            ctx = _resolve_entry_context(scenario, row_entry)
            context_list.append(ctx)
            if first_data_row is None:
                first_data_row = current_row
            last_data_row = current_row
            _write_rx_row(current_row, scenario, ctx)
            current_row += 1
        scenario_end = current_row - 1
        if scenario_start <= scenario_end:
            if scenario_end > scenario_start:
                _merge(ws, f"B{scenario_start}:B{scenario_end}")
            _set_cell(
                ws,
                scenario_start,
                channel_col,
                scenario.channel,
                font=FONT_BODY,
                alignment=ALIGN_CENTER,
                border=True,
            )
            for row_index in range(scenario_start + 1, scenario_end + 1):
                cell = ws.cell(row=row_index, column=channel_col)
                cell.border = BORDER_THIN
                cell.alignment = ALIGN_CENTER

    has_rx_rows = current_row > sub_header_row_rx + 1
    if has_rx_rows:
        current_row += 1
    header_row_tx = max(current_row, sub_header_row_rx + 2)
    if current_row < header_row_tx:
        current_row = header_row_tx
    sub_header_row_tx = header_row_tx + 1
    _write_header_block(
        header_row_tx,
        throughput_title="TX (Unit:Mb)",
        rssi_title="TX_RSSI (Unit:dBm)",
        include_item=False,
    )

    current_row = sub_header_row_tx + 1
    for scenario in group.channels:
        scenario_key = getattr(scenario, "key", "")
        scenario_rows = rows_by_scenario.get(scenario_key, []) if scenario_key else []
        if not scenario_rows:
            continue
        context_list = scenario_contexts.get(scenario_key, [])
        scenario_start = current_row
        for index, row_entry in enumerate(scenario_rows):
            if index < len(context_list):
                ctx = context_list[index]
            else:
                ctx = _resolve_entry_context(scenario, row_entry)
                context_list.append(ctx)
            if first_data_row is None:
                first_data_row = current_row
            last_data_row = current_row
            _write_tx_row(current_row, scenario, ctx)
            current_row += 1
        scenario_end = current_row - 1
        if scenario_start <= scenario_end:
            if scenario_end > scenario_start:
                _merge(ws, f"B{scenario_start}:B{scenario_end}")
            _set_cell(
                ws,
                scenario_start,
                channel_col,
                scenario.channel,
                font=FONT_BODY,
                alignment=ALIGN_CENTER,
                border=True,
            )
            for row_index in range(scenario_start + 1, scenario_end + 1):
                cell = ws.cell(row=row_index, column=channel_col)
                cell.border = BORDER_THIN
                cell.alignment = ALIGN_CENTER

    data_end_row = max(current_row - 1, sub_header_row_tx)
    if first_data_row is not None and last_data_row is not None and first_data_row <= last_data_row:
        group_label_start = first_data_row
        group_label_end = last_data_row
        for merged_range in list(ws.merged_cells.ranges):
            if merged_range.min_col == item_col and merged_range.max_col == item_col:
                if merged_range.min_row <= group_label_end and merged_range.max_row >= group_label_start:
                    ws.unmerge_cells(str(merged_range))
        _merge(ws, f"A{group_label_start}:A{group_label_end}")
        _set_cell(
            ws,
            group_label_start,
            item_col,
            group.summary_label or group.title,
            font=FONT_BODY,
            alignment=ALIGN_CENTER_WRAP,
            border=True,
        )
    used_last_col = rssi_end_col if angle_count else ovality_col
    total_rows = max(data_end_row - header_row_rx, 0)
    LOGGER.info(
        'RVO matrix written | group=%s rows=%d angles=%d',
        group.key,
        total_rows,
        raw_angle_count,
    )
    return data_end_row, used_last_col

def _nice_number(value: float, *, round_up: bool = True) -> float:
    if value <= 0:
        return 0.0
    exponent = math.floor(math.log10(value))
    fraction = value / (10 ** exponent)
    candidates = (1, 2, 2.5, 5, 10)
    if round_up:
        for candidate in candidates:
            if fraction <= candidate:
                fraction = candidate
                break
        else:
            fraction = candidates[-1]
    else:
        for candidate in reversed(candidates):
            if fraction >= candidate:
                fraction = candidate
                break
        else:
            fraction = candidates[0]
    return fraction * (10 ** exponent)


def _resolve_throughput_axis(values: Sequence[Optional[float]]) -> Tuple[float, float]:
    numeric_values: list[float] = []
    for value in values:
        if value is None:
            continue
        try:
            numeric = float(value)
        except (TypeError, ValueError):
            continue
        if not math.isfinite(numeric):
            continue
        numeric_values.append(numeric)
    if not numeric_values:
        return 10.0, 2.0
    max_value = max(numeric_values)
    if max_value <= 0:
        return 10.0, 2.0
    padded = max_value * 1.1
    upper = _nice_number(padded, round_up=True)
    if upper <= 0:
        upper = 10.0
    major = _nice_number(upper / 5.0, round_up=True)
    if major <= 0:
        major = upper
    LOGGER.info(
        "Resolved throughput axis | max=%.2f padded=%.2f upper=%.2f major=%.2f",
        max_value,
        padded,
        upper,
        major,
    )
    return upper, major


def _style_chart(
    chart: ScatterChart,
    *,
    data_points: Sequence[Optional[float]],
    show_markers: bool,
    step_values: Sequence[float],
) -> None:
    chart.width = 12.0
    chart.height = 7.5
    chart.varyColors = False
    LOGGER.info("Configured scatter chart | vary_colors=%s", chart.varyColors)

    if chart.legend is None:
        chart.legend = Legend()
    chart.legend.position = 'b'
    chart.legend.overlay = False
    LOGGER.info(
        'Chart legend configured | position=%s overlay=%s series=%d',
        chart.legend.position,
        chart.legend.overlay,
        len(chart.series),
    )
    chart.x_axis.majorGridlines = None
    chart.y_axis.majorGridlines = ChartLines()
    chart.y_axis.majorGridlines.spPr = GraphicalProperties(
        ln=LineProperties(solidFill=COLOR_GRIDLINE, w=12000)
    )
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.title = None
    chart.y_axis.title = None
    chart.x_axis.majorTickMark = "cross"
    chart.y_axis.majorTickMark = "out"
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "nextTo"
    chart.x_axis.crosses = "min"
    chart.y_axis.crosses = "min"
    chart.x_axis.number_format = "0"
    chart.y_axis.number_format = "0"
    chart.x_axis.tickLblSkip = 1
    chart.x_axis.tickMarkSkip = 1
    if step_values:
        x_min = float(step_values[0])
        x_max = float(step_values[-1])
        if len(step_values) > 1:
            raw_step = float(step_values[1] - step_values[0])
            major_step = raw_step if raw_step > 0 else 1.0
        else:
            major_step = 1.0
    else:
        x_min = 0.0
        x_max = float(max(len(data_points) - 1, 0))
        major_step = 1.0 if x_max == 0 else x_max / max(len(data_points) - 1, 1)
        if major_step <= 0:
            major_step = 1.0

    chart.x_axis.scaling.min = x_min
    chart.x_axis.scaling.max = x_max
    chart.x_axis.majorUnit = max(major_step, 1.0)

    upper, major = _resolve_throughput_axis(data_points)
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = upper
    chart.y_axis.majorUnit = major
    chart.y_axis.minorTickMark = "none"
    chart.x_axis.minorTickMark = "none"

    chart.plot_area.layout = Layout(
        manualLayout=ManualLayout(x=0.06, y=0.12, w=0.8, h=0.7)
    )
    LOGGER.info(
        "Chart layout + axis | width=%.1f height=%.1f y_max=%.2f major=%.2f x_min=%.2f x_max=%.2f x_step=%.2f layout=(x=%.2f y=%.2f w=%.2f h=%.2f)",
        chart.width,
        chart.height,
        upper,
        major,
        x_min,
        x_max,
        chart.x_axis.majorUnit,
        chart.plot_area.layout.manualLayout.x,
        chart.plot_area.layout.manualLayout.y,
        chart.plot_area.layout.manualLayout.w,
        chart.plot_area.layout.manualLayout.h,
    )

    for idx, series in enumerate(chart.series):
        color = SERIES_COLORS[idx % len(SERIES_COLORS)]
        if hasattr(series, "graphicalProperties") and hasattr(series.graphicalProperties, "line"):
            series.graphicalProperties.line.width = 20000  # 2pt
            series.graphicalProperties.line.solidFill = color
        series.smooth = False
        if show_markers:
            marker = Marker(symbol="circle")
            marker.graphicalProperties = GraphicalProperties(solidFill=color)
            marker.size = 6
        else:
            marker = Marker(symbol="none")
        series.marker = marker
        LOGGER.info(
            "Styled chart series | title=%s line_width=%s marker_symbol=%s marker_size=%s smooth=%s",
            getattr(series, "title", None),
            getattr(series.graphicalProperties.line, "width", None) if hasattr(series, "graphicalProperties") else None,
            marker.symbol,
            getattr(marker, "size", None),
            series.smooth,
        )


def _build_throughput_chart(
    *,
    title: str,
    series_definitions: Sequence[tuple[str, Reference, Reference]],
    data_points: Sequence[Optional[float]],
    show_markers: bool,
    step_values: Sequence[float],
) -> ScatterChart:
    chart = ScatterChart()
    chart.scatterStyle = 'line'
    chart.varyColors = False
    chart.title = title
    for series_title, categories, values in series_definitions:
        series = Series(values, xvalues=categories, title=series_title)
        series.smooth = False
        chart.series.append(series)
    _style_chart(
        chart,
        data_points=data_points,
        show_markers=show_markers,
        step_values=step_values,
    )
    LOGGER.info(
        'Built throughput chart | title=%s series_count=%d point_count=%d markers=%s',
        title,
        len(series_definitions),
        len(data_points),
        show_markers,
    )
    return chart

def _add_group_charts(
    ws: Worksheet,
    group: ScenarioGroup,
    layout: GroupLayout,
    sections: Sequence[dict[str, object]],
    *,
    anchor_row: int,
) -> int:
    if not sections:
        return anchor_row

    attenuations = sorted(group.attenuation_steps) or DEFAULT_ATTENUATIONS
    rx_series_defs: list[tuple[str, Reference, Reference]] = []
    tx_series_defs: list[tuple[str, Reference, Reference]] = []
    rx_points: list[float] = []
    tx_points: list[float] = []
    for idx, section in enumerate(sections):
        scenario = section['scenario']  # type: ignore[index]
        data_start = section['data_start']  # type: ignore[index]
        data_end = section['data_end']  # type: ignore[index]
        categories = Reference(ws, min_col=2, min_row=data_start, max_row=data_end)
        rx_col = layout.rx_cols[idx] if idx < len(layout.rx_cols) else 4
        tx_col = layout.tx_cols[idx] if idx < len(layout.tx_cols) else 5
        rx_values = Reference(ws, min_col=rx_col, min_row=data_start, max_row=data_end)
        tx_values = Reference(ws, min_col=tx_col, min_row=data_start, max_row=data_end)
        rx_series_defs.append((scenario.channel, categories, rx_values))
        tx_series_defs.append((scenario.channel, categories, tx_values))
        rx_points.extend(float(val) for val in scenario.rx_values.values() if val is not None)
        tx_points.extend(float(val) for val in scenario.tx_values.values() if val is not None)

    rx_chart = _build_throughput_chart(
        title=f"{group.title} RVR Throughput_RX",
        series_definitions=rx_series_defs,
        data_points=rx_points,
        show_markers=False,
        step_values=attenuations,
    )
    tx_chart = _build_throughput_chart(
        title=f"{group.title} RVR Throughput_TX",
        series_definitions=tx_series_defs,
        data_points=tx_points,
        show_markers=False,
        step_values=attenuations,
    )

    data_last_row = max(section['data_end'] for section in sections)  # type: ignore[index]
    first_anchor_row = max(data_last_row + 2, CHART_MIN_TOP_ROW)
    base_col_index = layout.rx_cols[0] if layout.rx_cols else 3
    rx_col_anchor = get_column_letter(base_col_index)

    gap_columns = 3
    tx_col_index = (layout.tx_cols[-1] if layout.tx_cols else base_col_index) + gap_columns
    if layout.right_rx_cols:
        tx_col_index = min(tx_col_index, layout.right_rx_cols[0] - 1)
        if tx_col_index <= base_col_index:
            tx_col_index = layout.right_rx_cols[0]
    tx_col_index = max(tx_col_index, base_col_index + 1)
    tx_col_anchor = get_column_letter(tx_col_index)

    rx_chart.anchor = f"{rx_col_anchor}{first_anchor_row}"
    ws.add_chart(rx_chart)

    tx_chart.anchor = f"{tx_col_anchor}{first_anchor_row}"
    ws.add_chart(tx_chart)

    chart_bottom = first_anchor_row + CHART_VERTICAL_HEIGHT_ROWS
    bottom_row = max(data_last_row, chart_bottom)

    LOGGER.info(
        'Group charts placed | group=%s anchor=%s rx_series=%d tx_series=%d bottom_row=%d',
        group.key,
        f"{rx_col_anchor}{first_anchor_row}",
        len(rx_series_defs),
        len(tx_series_defs),
        bottom_row,
    )
    return bottom_row


def _prepare_rvo_chart_context(result_file: Path | str) -> dict[str, Any]:
    chart_dir = Path(result_file).resolve().parent / "rvo_charts"
    chart_dir.mkdir(parents=True, exist_ok=True)
    generator = PerformanceRvrChartGenerator()
    try:
        dataframe = generator._load_rvr_dataframe(Path(result_file))  # type: ignore[attr-defined]
    except Exception:
        LOGGER.exception('Failed to load RVO dataframe for charts: %s', result_file)
        dataframe = None
    return {"generator": generator, "dataframe": dataframe, "directory": chart_dir}


def _add_rvo_polar_chart(
    ws: Worksheet,
    group: ScenarioGroup,
    layout: Optional[GroupLayout],
    sections: Sequence[dict[str, object]],
    *,
    anchor_row: int,
    context: dict[str, Any],
    data_end_row: Optional[int] = None,
    data_end_col: Optional[int] = None,
) -> int:
    data_last_row = data_end_row if data_end_row is not None else max(
        (section['data_end'] for section in sections),
        default=anchor_row,
    )
    generator: Optional[PerformanceRvrChartGenerator] = context.get("generator")
    dataframe = context.get("dataframe")
    chart_dir: Optional[Path] = context.get("directory")

    if generator is None or chart_dir is None:
        LOGGER.warning('Invalid RVO chart context; skip chart for %s', group.key)
        return data_last_row

    if dataframe is None or getattr(dataframe, "empty", True):
        LOGGER.warning('RVO dataframe empty; polar chart unavailable for %s', group.key)
        base_col = data_end_col if data_end_col is not None else None
        if base_col is None and layout is not None:
            base_col = layout.right_tx_cols[-1] if layout.right_tx_cols else layout.tx_cols[-1]
        placeholder_col = (base_col or 2) + CHART_COLUMN_GAP + 1
        _set_cell(
            ws,
            anchor_row,
            placeholder_col,
            "Polar chart unavailable",
            font=FONT_BODY,
            alignment=ALIGN_CENTER,
            border=True,
        )
        return max(data_last_row, anchor_row)

    try:
        df = dataframe
        if "Scenario_Group_Key" in df.columns:
            df = df[df["Scenario_Group_Key"].isin(group.raw_keys)]
        if "__test_type_display__" in df.columns:
            df = df[df["__test_type_display__"].astype(str).str.upper() == "RVO"]
        if df.empty:
            raise ValueError("No RVO rows matched group keys")

        direction_column = "__direction_display__" if "__direction_display__" in df.columns else "Direction"
        if direction_column in df.columns:
            directions = [
                str(direction).strip()
                for direction in df[direction_column].unique()
                if str(direction).strip()
            ]
        else:
            directions = ["RX"]
        def _is_rx_direction(value: str) -> bool:
            upper = value.upper()
            return any(token in upper for token in ("RX", "DL", "DOWN"))

        def _is_tx_direction(value: str) -> bool:
            upper = value.upper()
            return any(token in upper for token in ("TX", "UL", "UP"))

        ordered_dirs: list[tuple[str, str]] = []
        rx_direction = next((d for d in directions if _is_rx_direction(d)), directions[0] if directions else None)
        if rx_direction:
            ordered_dirs.append(("RX", rx_direction))
        remaining = [d for d in directions if d != rx_direction]
        tx_direction = next((d for d in remaining if _is_tx_direction(d)), None)
        if tx_direction:
            ordered_dirs.append(("TX", tx_direction))
            remaining = [d for d in remaining if d != tx_direction]
        if not ordered_dirs and remaining:
            ordered_dirs.append(("RX", remaining[0]))
            remaining = remaining[1:]
        if tx_direction is None and remaining:
            ordered_dirs.append(("TX", remaining[0]))

        channel_column = "__channel_display__" if "__channel_display__" in df.columns else "Channel"
        channel_labels = [
            str(value).strip() or "Unknown"
            for value in df[channel_column].dropna().unique()
        ] if channel_column in df.columns else []

        generated_images: list[tuple[int, int, str, str, Path]] = []
        max_row_idx = -1
        for row_idx, channel_label in enumerate(channel_labels or [""]):
            channel_df = df[df[channel_column] == channel_label] if channel_column in df.columns else df
            if channel_df.empty:
                continue
            for col_idx, (label, direction) in enumerate(ordered_dirs or [("RX", "RX")]):
                subset = channel_df[channel_df[direction_column] == direction] if direction_column in channel_df.columns else channel_df
                if subset.empty:
                    continue
                base_title = generator._format_chart_title(
                    group.standard,
                    group.bandwidth,
                    group.freq,
                    group.test_type,
                    direction,
                )
                title = f"{channel_label} {base_title}" if channel_label else base_title
                image_path = generator._save_rvo_chart(subset, title, chart_dir)
                if image_path is not None:
                    generated_images.append((row_idx, col_idx, channel_label, direction, image_path))
                    max_row_idx = max(max_row_idx, row_idx)

        if not generated_images:
            raise ValueError("Failed to generate polar chart")

        if layout is not None and getattr(layout, "rx_cols", None):
            base_col_index = layout.rx_cols[0] if layout.rx_cols else 2
        else:
            base_col_index = 2
        column_step = 16
        row_step = POLAR_CHART_HEIGHT_ROWS + POLAR_ROW_SPACING
        for row_idx, col_idx, channel_label, direction, image_path in generated_images:
            anchor_col_index = base_col_index + col_idx * column_step
            anchor_letter = get_column_letter(anchor_col_index)
            chart_anchor_row = anchor_row + row_idx * row_step
            image = Image(str(image_path))
            image.width = POLAR_IMAGE_WIDTH
            image.height = POLAR_IMAGE_HEIGHT
            image.anchor = f"{anchor_letter}{chart_anchor_row}"
            ws.add_image(image)
            LOGGER.info(
                'Inserted RVO polar chart | group=%s channel=%s direction=%s path=%s anchor=%s',
                group.key,
                channel_label,
                direction,
                image_path,
                f"{anchor_letter}{chart_anchor_row}",
            )
        chart_bottom = anchor_row + POLAR_CHART_HEIGHT_ROWS if max_row_idx < 0 else anchor_row + max_row_idx * row_step + POLAR_CHART_HEIGHT_ROWS
    except Exception:
        LOGGER.exception('Failed to insert RVO polar chart for group %s', group.key)
        if layout is not None and getattr(layout, "rx_cols", None):
            fallback_col = layout.rx_cols[0] if layout.rx_cols else 2
        else:
            fallback_col = 2
        _set_cell(
            ws,
            anchor_row,
            fallback_col,
            "Polar chart unavailable",
            font=FONT_BODY,
            alignment=ALIGN_CENTER,
            border=True,
        )
        chart_bottom = anchor_row + POLAR_CHART_HEIGHT_ROWS
    bottom_row = max(data_last_row, chart_bottom)
    return bottom_row

# ---------------------------------------------------------------------------
# Data extraction
# ---------------------------------------------------------------------------


def _load_scenario_groups(result_file: Path | str, *, test_type: str | None = None) -> List[ScenarioGroup]:
    path = Path(result_file)
    if not path.exists():
        LOGGER.warning('Result CSV not found for project report: %s', path)
        return []

    normalized_filter = _normalize_test_type(test_type) if test_type else None
    buckets: dict[str, dict[str, object]] = {}
    total_rows = 0
    filtered_rows = 0
    matched_rows = 0
    type_counts: Counter[str] = Counter()
    try:
        with path.open('r', encoding='utf-8-sig', newline='') as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                if not row:
                    continue
                total_rows += 1
                row_type = _detect_row_test_type(row)
                type_counts[row_type] += 1
                if normalized_filter and row_type != normalized_filter:
                    filtered_rows += 1
                    continue
                matched_rows += 1
                raw_key = row.get('Scenario_Group_Key')
                base_key = _group_base_key(raw_key)
                bucket = buckets.setdefault(
                    base_key,
                    {
                        'raw_keys': set(),
                        'freq': Counter(),
                        'standard': Counter(),
                        'bandwidth': Counter(),
                        'angle': Counter(),
                        'angle_order': [],
                        'attenuations': set(),
                        'channels': {},
                        'channel_order': [],
                        'test_type': Counter(),
                    },
                )

                bucket['raw_keys'].add(_normalize_scenario_key(raw_key))
                bucket['test_type'][row_type] += 1

                freq = str(row.get('Freq_Band') or '').strip()
                if freq:
                    bucket['freq'][freq] += 1
                standard = str(row.get('Standard') or '').strip()
                if standard:
                    bucket['standard'][standard] += 1
                bandwidth_raw = row.get('BW') or row.get('Bandwidth')
                bandwidth = _format_bandwidth(bandwidth_raw)
                if bandwidth:
                    bucket['bandwidth'][bandwidth] += 1
                angle = _format_angle(row.get('Angel') or row.get('Angle'))
                if angle:
                    bucket['angle'][angle] += 1

                channel_label = _format_channel(row.get('CH_Freq_MHz') or row.get('Channel'))
                channels = bucket['channels']
                channel_bucket = channels.setdefault(
                    channel_label,
                    {
                        'scenario_key': _normalize_scenario_key(raw_key),
                        'rx': {},
                        'tx': {},
                        'rssi_rx': {},
                        'rssi_tx': {},
                        'attenuations': set(),
                        'angle_order': [],
                        'rvo_rx': {},
                        'rvo_tx': {},
                        'rvo_rssi_rx': {},
                        'rvo_rssi_tx': {},
                    },
                )
                if channel_label not in bucket['channel_order']:
                    bucket['channel_order'].append(channel_label)

                attenuation = _sanitize_number(row.get('DB') or row.get('Total_Path_Loss'))
                if attenuation is None:
                    continue
                bucket['attenuations'].add(attenuation)
                channel_bucket['attenuations'].add(attenuation)

                direction = str(row.get('Direction') or '').upper()
                throughput = _sanitize_number(row.get('Throughput'))
                if throughput is not None:
                    if row_type == 'RVO':
                        angle_label = angle or '0deg'
                        angle_order = bucket['angle_order']
                        if angle_label not in angle_order:
                            angle_order.append(angle_label)
                        channel_angle_order = channel_bucket['angle_order']
                        if angle_label not in channel_angle_order:
                            channel_angle_order.append(angle_label)
                        if direction in {'DL', 'RX'}:
                            angle_map = channel_bucket['rvo_rx'].setdefault(float(attenuation), {})
                            angle_map[angle_label] = throughput
                        elif direction in {'UL', 'TX'}:
                            angle_map = channel_bucket['rvo_tx'].setdefault(float(attenuation), {})
                            angle_map[angle_label] = throughput

                    if direction in {'DL', 'RX'}:
                        channel_bucket['rx'][attenuation] = throughput
                    elif direction in {'UL', 'TX'}:
                        channel_bucket['tx'][attenuation] = throughput

                rssi = _sanitize_number(row.get('RSSI'))
                if rssi is not None:
                    if direction in {'DL', 'RX'}:
                        channel_bucket['rssi_rx'][attenuation] = rssi
                        if row_type == 'RVO':
                            angle_label = angle or '0deg'
                            angle_map = channel_bucket['rvo_rssi_rx'].setdefault(float(attenuation), {})
                            angle_map[angle_label] = rssi
                    elif direction in {'UL', 'TX'}:
                        channel_bucket['rssi_tx'][attenuation] = rssi
                        if row_type == 'RVO':
                            angle_label = angle or '0deg'
                            angle_map = channel_bucket['rvo_rssi_tx'].setdefault(float(attenuation), {})
                            angle_map[angle_label] = rssi
    except Exception:
        LOGGER.exception('Failed to parse project CSV: %s', path)
        return []

    groups: List[ScenarioGroup] = []
    for base_key, bucket in buckets.items():
        attenuation_steps = sorted(bucket['attenuations'])
        group_attenuations = attenuation_steps if attenuation_steps else DEFAULT_ATTENUATIONS.copy()
        group = ScenarioGroup(
            key=base_key,
            freq=_most_common(bucket['freq'], default='5G'),
            standard=_most_common(bucket['standard'], default='Auto'),
            bandwidth=_most_common(bucket['bandwidth'], default='20/40/80 MHz'),
            angle_label=_most_common(bucket['angle'], default='0deg'),
            attenuation_steps=group_attenuations,
        )
        group.step_summary = _summarize_attenuation_steps(group_attenuations)
        group.raw_keys = set(bucket['raw_keys'])
        group.test_type = _most_common(
            bucket['test_type'],
            default=_normalize_test_type(normalized_filter or 'RVR'),
        )
        group.angle_order = _sorted_unique_angles(bucket.get('angle_order', []))

        channel_scenarios: List[ProjectScenario] = []
        for channel_label in bucket['channel_order']:
            channel_bucket = bucket['channels'].get(channel_label) or {}
            channel_atts = sorted(channel_bucket.get('attenuations', set()))
            scenario_angle_order = _sorted_unique_angles(
                channel_bucket.get('angle_order', []) or group.angle_order
            )

            def _ordered_angle_map(raw_map: Dict[float, Dict[str, float]]) -> Dict[float, Dict[str, float]]:
                ordered: Dict[float, Dict[str, float]] = {}
                for att_key, angle_map in raw_map.items():
                    if not isinstance(att_key, (int, float)):
                        continue
                    att_float = float(att_key)
                    ordered_angles: Dict[str, float] = {}
                    for angle_name in scenario_angle_order:
                        if angle_name in angle_map:
                            ordered_angles[angle_name] = angle_map[angle_name]
                    if not ordered_angles and angle_map:
                        ordered_angles = dict(sorted(angle_map.items()))
                    ordered[att_float] = ordered_angles
                return ordered

            scenario = ProjectScenario(
                key=channel_bucket.get('scenario_key') or base_key,
                freq=group.freq,
                standard=group.standard,
                bandwidth=group.bandwidth,
                channel_label=channel_label,
                angle_label=group.angle_label,
                attenuation_steps=channel_atts if channel_atts else group_attenuations.copy(),
                step_summary=group.step_summary,
                rx_values=dict(sorted(channel_bucket.get('rx', {}).items())),
                tx_values=dict(sorted(channel_bucket.get('tx', {}).items())),
                rssi_rx=dict(sorted(channel_bucket.get('rssi_rx', {}).items())),
                rssi_tx=dict(sorted(channel_bucket.get('rssi_tx', {}).items())),
                angle_order=scenario_angle_order,
                angle_rx_matrix=_ordered_angle_map(channel_bucket.get('rvo_rx', {})),
                angle_tx_matrix=_ordered_angle_map(channel_bucket.get('rvo_tx', {})),
                angle_rssi_rx_matrix=_ordered_angle_map(channel_bucket.get('rvo_rssi_rx', {})),
                angle_rssi_tx_matrix=_ordered_angle_map(channel_bucket.get('rvo_rssi_tx', {})),
            )
            LOGGER.info(
                'Channel aggregated | base_key=%s channel=%s attenuations=%d rx_points=%d tx_points=%d',
                base_key,
                channel_label,
                len(scenario.attenuation_steps),
                len(scenario.rx_values),
                len(scenario.tx_values),
            )
            channel_scenarios.append(scenario)

        if not channel_scenarios:
            continue
        group.channels = channel_scenarios
        LOGGER.info(
            'Scenario group aggregated | base_key=%s freq=%s standard=%s channels=%d step_summary=%s',
            base_key,
            group.freq,
            group.standard,
            len(group.channels),
            group.step_summary or 'N/A',
        )
        groups.append(group)

    groups.sort(
        key=lambda grp: (
            grp.freq.upper(),
            grp.standard.upper(),
            grp.bandwidth.upper(),
            grp.key,
        ),
    )
    distribution = ', '.join(
        f"{name}={count}" for name, count in sorted(type_counts.items())
    ) or 'none'
    LOGGER.info(
        'Scenario test type summary | total_rows=%d matched=%d filtered=%d filter=%s distribution=%s',
        total_rows,
        matched_rows,
        filtered_rows,
        normalized_filter or 'ALL',
        distribution,
    )
    LOGGER.info(
        'Loaded project scenario groups | count=%d total_rows=%d source=%s filter=%s',
        len(groups),
        total_rows,
        path,
        normalized_filter or 'ALL',
    )
    return groups

# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def generate_project_report(
    result_file: Path | str,
    output_path: Path | str,
    forced_test_type: str | None = None,
) -> Path:
    normalized_type = _normalize_test_type(forced_test_type) if forced_test_type else None
    groups = _load_scenario_groups(result_file, test_type=normalized_type)
    actual_type = normalized_type or (groups[0].test_type if groups else "RVR")
    LOGGER.info(
        "Preparing project %s report | groups=%d source=%s",
        actual_type,
        len(groups),
        result_file,
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet_name = actual_type.lower()
    sheet.title = sheet_name[:31]
    _configure_sheet(sheet)

    rvo_att_entries: List[Tuple[Optional[float], str, str]] = []
    if actual_type == "RVO":
        rvo_att_entries = _resolve_rvo_att_steps()

    chart_context: Optional[dict[str, Any]] = None
    if actual_type == "RVO":
        chart_context = _prepare_rvo_chart_context(result_file)

    if groups:
        report_last_col_index = _determine_report_last_column(
            groups,
            actual_type,
            rvo_att_entries=rvo_att_entries,
        )
    else:
        report_last_col_index = column_index_from_string("M")
    report_last_col_index = max(1, report_last_col_index)
    report_last_column_letter = get_column_letter(report_last_col_index)

    if not groups:
        title_end_row = _write_report_title(
            sheet,
            groups=[],
            test_type=actual_type,
            start_row=1,
            last_column=report_last_column_letter,
        )
        placeholder_row = title_end_row + 2
        _merge(sheet, f"A{placeholder_row}:{report_last_column_letter}{placeholder_row}")
        _set_cell(
            sheet,
            placeholder_row,
            1,
            "No Wi-Fi data available for project report",
            font=FONT_SECTION,
            alignment=ALIGN_CENTER,
            border=True,
        )
        LOGGER.warning("No scenarios found; generated placeholder sheet named %s.", sheet.title)
    else:
        title_end_row = _write_report_title(
            sheet,
            groups=groups,
            test_type=actual_type,
            start_row=1,
            last_column=report_last_column_letter,
        )
        current_row = title_end_row + 2
        for group_index, group in enumerate(groups):
            if group_index > 0:
                current_row += 1
            group_header_row = _write_group_header(
                sheet,
                group,
                start_row=current_row,
                last_column=report_last_column_letter,
            )
            header_row = group_header_row + 1

            if actual_type == "RVO":
                override_steps = [entry[0] for entry in rvo_att_entries]
                att_labels = [entry[2] for entry in rvo_att_entries]
                item_labels = [entry[1] for entry in rvo_att_entries]
                angles, entries = _prepare_rvo_table_entries(
                    group,
                    override_steps,
                    att_labels,
                    item_labels,
                )
                matrix_angles = angles if angles else group.angle_order
                if not matrix_angles:
                    LOGGER.warning('Group %s missing RVO angle definitions', group.key)
                data_end, last_col = _write_rvo_table(
                    sheet,
                    group,
                    start_row=header_row,
                    angles=matrix_angles,
                    entries=entries,
                )
                current_row = data_end + 1
                if chart_context is not None:
                    chart_anchor_row = data_end + 2
                    last_row = _add_rvo_polar_chart(
                        sheet,
                        group,
                        None,
                        [],
                        anchor_row=chart_anchor_row,
                        context=chart_context,
                        data_end_row=data_end,
                        data_end_col=last_col,
                    )
                    current_row = max(current_row, last_row + 6)
                continue

            group_layout = _build_group_layout(len(group.channels))
            sub_header_row = _write_headers(
                sheet,
                group,
                group.channels,
                group_layout,
                header_row=header_row,
            )
            data_start = sub_header_row + 1
            data_end = _write_data(
                sheet,
                group,
                group.channels,
                group_layout,
                start_row=data_start,
            )
            current_row = data_end + 1

            used_channels = list(group.channels)
            sections = [
                {"scenario": scenario, "header_row": header_row, "data_start": data_start, "data_end": data_end}
                for scenario in used_channels
            ]

            if sections:
                anchor_row = header_row
                if actual_type == "RVO" and chart_context is not None:
                    last_row = _add_rvo_polar_chart(
                        sheet,
                        group,
                        group_layout,
                        sections,
                        anchor_row=anchor_row,
                        context=chart_context,
                    )
                else:
                    last_row = _add_group_charts(sheet, group, group_layout, sections, anchor_row=anchor_row)
                current_row = max(current_row, last_row + 6)
            else:
                LOGGER.warning("Group %s has no channel sections to chart.", group.key)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    LOGGER.info("Project report saved to %s (sheet=%s)", output, sheet.title)
    return output.resolve()


__all__ = ["generate_project_report", "ProjectScenario"]

