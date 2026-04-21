from __future__ import annotations

import json
import logging
from dataclasses import dataclass
from datetime import datetime
import hashlib
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

from openpyxl import load_workbook
from PyQt5.QtCore import QObject, QThread, pyqtSignal
from PyQt5.QtWidgets import QFileDialog, QDialog
from qfluentwidgets import MessageBox

from src.tools.mysql_tool import MySqlClient
from src.tools.mysql_tool.operations import ensure_project, ensure_test_report
from src.tools.mysql_tool.operations import PerformanceTableManager
from src.tools.mysql_tool.schema import ensure_report_tables
from src.tools.mysql_tool.sql_writer import SqlWriter
from src.ui.view.titlebar.import_dialog import ImportDialog
from src.ui.view.titlebar.import_sheets_dialog import ImportSheetsDialog
from src.util.constants import (
    AP_MODEL_CHOICES,
    AP_REGION_CHOICES,
    BT_DEVICE_CHOICES,
    BT_REMOTE_CHOICES,
    BT_TYPE_CHOICES,
    DUT_OS_CHOICES,
    HW_PHASE_CHOICES,
    IDENTIFIER_SANITIZE_PATTERN,
    LAB_ENV_COEX_MODE_CHOICES,
    LAB_ENV_CONNECT_TYPE_CHOICES,
    PROJECT_ID_CHOICES,
    PROJECT_TYPES,
    RUN_TYPE_WIFI_SMARTTEST,
    TEST_REPORT_CHOICES,
    TEST_REPORT_COMPATIBILITY,
    TEST_REPORT_PEAK_THROUGHPUT,
    TEST_REPORT_RVO,
    TEST_REPORT_RVR,
    WIFI_MODULE_CHOICES,
    WIFI_PRODUCT_PROJECT_MAP,
)


def _store_excel_artifact(
    client: MySqlClient,
    *,
    test_report_id: int,
    excel_path: str,
) -> int:
    content = Path(excel_path).read_bytes()
    digest = hashlib.sha256(content).hexdigest()
    file_name = Path(excel_path).name
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    size_bytes = len(content)

    insert_sql = (
        "INSERT INTO `artifact` "
        "(`test_report_id`, `file_name`, `content_type`, `sha256`, `size_bytes`, `content`) "
        "VALUES (%s, %s, %s, %s, %s, %s) "
        "ON DUPLICATE KEY UPDATE `id`=LAST_INSERT_ID(`id`)"
    )
    artifact_id = client.insert(
        insert_sql,
        (
            int(test_report_id),
            file_name,
            content_type,
            digest,
            int(size_bytes),
            content,
        ),
    )
    return int(artifact_id)


def _insert_performance_rows(
    manager: PerformanceTableManager,
    *,
    test_report_id: int,
    execution_id: int,
    csv_name: str,
    data_type: str,
    rows: List[Mapping[str, Any]],
) -> int:
    if not rows:
        return 0

    manager.ensure_schema_initialized()
    # insert_columns = [name for name, _ in manager._BASE_COLUMNS]
    # insert_columns.extend(column.name for column in manager._STATIC_COLUMNS)
    BASE_COLUMN_NAMES = ["test_report_id", "execution_id", "csv_name"]
    insert_columns = BASE_COLUMN_NAMES[:]
    insert_columns.extend(column.name for column in manager._STATIC_COLUMNS)

    writer = SqlWriter(manager.TABLE_NAME)
    insert_sql = writer.insert_statement(insert_columns)

    headers = list(rows[0].keys())
    throughput_aliases = manager._collect_throughput_headers(headers)

    values: List[List[Any]] = []
    for row in rows:
        base_values: List[Any] = [
            test_report_id,
            execution_id,
            csv_name,
        ]
        row_values: List[Any] = base_values[:]

        for column in manager._STATIC_COLUMNS:
            if column.original == "Throughput":
                samples: List[Any] = []
                if throughput_aliases:
                    for alias in throughput_aliases:
                        value = row.get(alias)
                        if value is not None:
                            samples.append(value)
                else:
                    value = row.get(column.original)
                    if value is not None:
                        samples.append(value)
                raw_value = manager._compute_throughput_average(samples)
            else:
                raw_value = row.get(column.original)
            if column.normalizer is not None:
                raw_value = column.normalizer(raw_value)
            row_values.append(manager._normalize_cell(raw_value, column.sql_type))
        values.append(row_values)

    return manager._client.executemany(insert_sql, values)


# @dataclass(frozen=True)
# class ScenarioSpec:
#     band_label: str
#     ssid: str
#     wireless_mode: str
#     channel: int
#     bandwidth_label: str
#     security_mode: str
#     password: str
#     tx: int
#     rx: int
#
#     @property
#     def band_token(self) -> str:
#         if self.band_label.startswith("2.4"):
#             return "2.4"
#         if self.band_label.startswith("5"):
#             return "5"
#         return self.band_label
#
#     @property
#     def bandwidth_mhz(self) -> int:
#         text = self.bandwidth_label.strip().lower()
#         digits = "".join(ch for ch in text if ch.isdigit())
#         return int(digits or 0)


@dataclass(frozen=True)
class SheetImportBatch:
    report_type: str
    sheet_name: str
    rows: List[Dict[str, Any]]


_PROJECT_PAYLOAD_KEYS: tuple[str, ...] = (
    "customer",
    "project_type",
    "nickname",
    "project_name",
    "project_id",
    "soc",
    "wifi_module",
    "odm",
    "interface",
    "ecosystem",
)


# DEFAULT_SCENARIOS: Tuple[ScenarioSpec, ...] = (
#     #2.4G 20M
#     ScenarioSpec("2.4G", "ax3600_2g", "11ax", 1, "20 MHz", "Open System", "", 1, 1),
#     ScenarioSpec("2.4G", "ax3600_2g", "11ax", 6, "20 MHz", "Open System", "", 1, 1),
#     ScenarioSpec("2.4G", "ax3600_2g", "11ax", 11, "20 MHz", "Open System", "", 1, 1),
#     #2.4G 40M
#     ScenarioSpec("2.4G", "ax3600_2g", "11ax", 1, "40 MHz", "Open System", "", 1, 1),
#     ScenarioSpec("2.4G", "ax3600_2g", "11ax", 6, "40 MHz", "Open System", "", 1, 1),
#     ScenarioSpec("2.4G", "ax3600_2g", "11ax", 11, "40 MHz", "Open System", "", 1, 1),
#     ScenarioSpec("5G", "ax3600_5g", "11ax", 36, "80 MHz", "Open System", "", 1, 1),
#     ScenarioSpec("5G", "ax3600_5g", "11ax", 52, "80 MHz", "Open System", "", 1, 1),
#     ScenarioSpec("5G", "ax3600_5g", "11ax", 64, "80 MHz", "Open System", "", 1, 1),
#     ScenarioSpec("5G", "ax3600_5g", "11ax", 149, "80 MHz", "Open System", "", 1, 1),
#     ScenarioSpec("5G", "ax3600_5g", "11ax", 161, "80 MHz", "Open System", "", 1, 1),
# )

from collections import namedtuple

ScenarioSpec = namedtuple(
    "ScenarioSpec",
    [
        "band_token",  # e.g., "2.4G", "5G"
        "ssid_token",  # e.g., "ax3600_2g"
        "mode",  # e.g., "11n", "11ac", "11ax"
        "channel",  # e.g., 1, 6, 11, 36, 149
        "bandwidth",  # e.g., "20 MHz", "40 MHz"
        "security",  # e.g., "Open System"
        "extra_params",  # currently unused
        "tx_antennas",  # e.g., 1
        "rx_antennas",  # e.g., 1
    ],
)

# === 定义所有维度的配置 ===
# 1. 频段与对应的 SSID 映射
BAND_CONFIGS = {
    "2.4G": "ax3600_2g",
    "5G": "ax3600_5g"
}

# 2. 协议标准与支持的带宽
# 注意：这里我们直接使用数据库中期望的带宽字符串格式
MODE_BANDWIDTH_MAP = {
    "11n": ["20 MHz", "40 MHz"],
    "11ac": ["80 MHz", "160 MHz"],  # 11ac 通常在5G，但这里按逻辑配置
    "11ax": ["20 MHz", "40 MHz", "80 MHz", "160 MHz"]
}

# 3. 频段与合法信道的映射
BAND_CHANNEL_MAP = {
    "2.4G": [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11],  # 可以根据需要添加 13 等
    "5G": [36, 40, 44, 48, 52, 56, 60, 64, 100, 104, 108, 112, 116, 120, 124, 128, 132, 136, 140, 144, 149, 153, 157, 161, 165]
}

# 4. 其他固定参数
SECURITY = "Open System"
EXTRA_PARAMS = ""
TX_ANTENNAS = 1
RX_ANTENNAS = 1


# === 动态生成 DEFAULT_SCENARIOS ===
def _generate_default_scenarios() -> Tuple[ScenarioSpec, ...]:
    scenarios = []

    for band, ssid in BAND_CONFIGS.items():
        channels = BAND_CHANNEL_MAP[band]
        for mode, bandwidths in MODE_BANDWIDTH_MAP.items():
            # 过滤不合理的组合：11ac 不应在 2.4G 频段
            if band == "2.4G" and mode == "11ac":
                continue

            for channel in channels:
                for bandwidth in bandwidths:
                    # 对于 2.4G, 跳过 80/160MHz (实践中很少见)
                    if band == "2.4G" and bandwidth in ("80 MHz", "160 MHz"):
                        continue

                    spec = ScenarioSpec(
                        band_token=band,
                        ssid_token=ssid,
                        mode=mode,
                        channel=channel,
                        bandwidth=bandwidth,
                        security=SECURITY,
                        extra_params=EXTRA_PARAMS,
                        tx_antennas=TX_ANTENNAS,
                        rx_antennas=RX_ANTENNAS,
                    )
                    scenarios.append(spec)

    return tuple(scenarios)


# 最终的配置
DEFAULT_SCENARIOS: Tuple[ScenarioSpec, ...] = _generate_default_scenarios()



def build_scenario_group_key(spec: ScenarioSpec) -> str:
    def normalize(value: Any) -> str:
        text = "" if value is None else str(value).strip()
        if not text:
            return ""
        sanitized = IDENTIFIER_SANITIZE_PATTERN.sub("_", text).strip("_")
        return sanitized.upper()

    parts: list[str] = [
        f"BAND={normalize(spec.band_token)}",
        f"SSID={normalize(spec.ssid_token)}",
        f"MODE={normalize(spec.mode)}",
        f"CHANNEL={normalize(spec.channel)}",
        f"BANDWIDTH={normalize(spec.bandwidth)}",
        f"SECURITY={normalize(spec.security)}",
        f"TX={normalize(spec.tx_antennas)}",
        f"RX={normalize(spec.rx_antennas)}",
    ]
    return "SCENARIO|" + "|".join(parts)


def build_peak_scenario_group_key(
    spec: ScenarioSpec,
    *,
    standard: str,
    bandwidth_mhz: int,
) -> str:
    def normalize(value: Any) -> str:
        text = "" if value is None else str(value).strip()
        if not text:
            return ""
        sanitized = IDENTIFIER_SANITIZE_PATTERN.sub("_", text).strip("_")
        return sanitized.upper()

    #band_label = "2.4G" if spec.band_token == "2.4" else "5G" if spec.band_token == "5" else spec.band_label
    band_label = spec.band_token
    bw_label = f"{int(bandwidth_mhz)} MHz"
    mode_label = standard or spec.mode
    #mode_label = standard or spec.wireless_mode

    parts: list[str] = [
        f"BAND={normalize(band_label)}",
        f"SSID={normalize(spec.ssid_token)}",
        f"MODE={normalize(mode_label)}",
        f"CHANNEL={normalize(spec.channel)}",
        f"BANDWIDTH={normalize(bw_label)}",
        f"SECURITY={normalize(spec.security)}",
        f"TX={normalize(spec.tx_antennas)}",
        f"RX={normalize(spec.rx_antennas)}",
    ]
    return "SCENARIO|" + "|".join(parts)


def _parse_angle(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value)
    digits = []
    dot_seen = False
    for ch in text:
        if ch.isdigit():
            digits.append(ch)
            continue
        if ch == "." and not dot_seen:
            digits.append(ch)
            dot_seen = True
            continue
        if digits:
            break
    number = "".join(digits).strip(".")
    return float(number) if number else None


def _parse_standard(value: Any) -> str:
    text = "" if value is None else str(value).strip().lower()
    for token in ("11be", "11ax", "11ac", "11n", "11g", "11b", "11a"):
        if token in text:
            return token
    if "802.11ax" in text:
        return "11ax"
    if "802.11ac" in text:
        return "11ac"
    if "802.11n" in text:
        return "11n"
    return ""


def _parse_band_token(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    if "2.4" in text or "2.4G" in text:
        return "2.4"
    if "5" in text or "5G" in text:
        return "5"
    if "6" in text or "6G" in text:
        return "6"
    return ""


def _parse_bandwidth_mhz(value: Any) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, int) and not isinstance(value, bool):
        return int(value)
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = str(value).strip().upper()
    digits = "".join(ch for ch in text if ch.isdigit())
    return int(digits) if digits else None


def _performance_row(
    *,
    test_category: str,
    standard: str,
    band: str,
    bandwidth_mhz: int,
    channel: int,
    protocol: str,
    mode: Optional[str],
    direction: str,
    path_loss_db: float,
    rssi: Optional[float],
    angle_deg: Optional[float],
    throughput_mbps: Optional[float],
    throughput_peak_mbps: Optional[float],
    scenario_group_key: str,
) -> Dict[str, Any]:
    return {
        "Test_Category": test_category,
        "Standard": standard,
        "Freq_Band": band,
        "BW": bandwidth_mhz,
        "CH_Freq_MHz": channel,
        "Protocol": protocol,
        "Mode": mode,
        "Direction": direction,
        "DB": path_loss_db,
        "RSSI": rssi,
        "Angel": angle_deg,
        "Throughput": throughput_mbps,
        "Max_Rate": throughput_peak_mbps,
        "Profile_Mode": "",
        "Profile_Value": "",
        "Scenario_Group_Key": scenario_group_key,
    }


class PerformanceExcelImporter:
    def __init__(self, *, scenarios: Sequence[ScenarioSpec] = DEFAULT_SCENARIOS) -> None:
        self._scenarios = tuple(scenarios)
        self._scenario_by_channel: Dict[Tuple[str, int], ScenarioSpec] = {
            (spec.band_token, spec.channel): spec for spec in self._scenarios
        }

    def _load_workbook(self, path: str | Path):
        return load_workbook(Path(path), data_only=True)

    def build_peak_throughput_rows(self, workbook, *, sheet_name: str) -> Tuple[List[Dict[str, Any]], List[str]]:
        ws = workbook[sheet_name]

        # 初始化上下文变量用于处理合并单元格
        context_frequency = None
        context_wifi_mode = None
        context_bandwidth = None
        context_channel = None
        context_rssi = None
        context_scenario = None

        seen_mode: set[Tuple[str, int, int, str, str]] = set()
        rows: List[Dict[str, Any]] = []

        print(f"[PEAK_DEBUG] Starting to parse Peak Throughput sheet: {sheet_name}")
        print(f"[PEAK_DEBUG] Sheet has {ws.max_row} rows and {ws.max_column} columns")

        # 从第2行开始（跳过标题行）
        for r in range(2, ws.max_row + 1):
            # 获取实际列数据（基于您提供的格式：12列）
            try:
                no_col = ws.cell(r, 1).value  # No. 列 (A)
                frequency = ws.cell(r, 2).value  # Frequency (B)
                wifi_mode = ws.cell(r, 3).value  # WiFi Mode (C)
                bandwidth = ws.cell(r, 4).value  # Bandwidth (D)
                channel = ws.cell(r, 5).value  # Channel (E)
                rssi = ws.cell(r, 6).value  # RSSI(dBm) (F)
                scenario = ws.cell(r, 7).value  # Scenario (G)
                tx_mbps = ws.cell(r, 8).value  # TX (Mbps) (H)
                rx_mbps = ws.cell(r, 9).value  # RX (Mbps) (I)
                link_rate = ws.cell(r, 10).value  # Link Rate(Mbps) (J)
                standard = ws.cell(r, 11).value  # Standard (K)
                result = ws.cell(r, 12).value  # Result (L)

                print(
                    f"[PEAK_DEBUG] Row {r} raw data: freq={frequency}, mode={wifi_mode}, bw={bandwidth}, ch={channel}, rssi={rssi}, tx={tx_mbps}, rx={rx_mbps}")

            except Exception as e:
                print(f"[PEAK_DEBUG] Error reading row {r}: {e}")
                continue

            # 跳过完全空的行
            if all(val is None for val in [frequency, wifi_mode, bandwidth, channel, rssi, tx_mbps, rx_mbps]):
                print(f"[PEAK_DEBUG] Row {r} is empty, skipping")
                continue

            # 更新上下文变量（处理合并单元格）
            if frequency is not None:
                context_frequency = frequency
            if wifi_mode is not None:
                context_wifi_mode = wifi_mode
            if bandwidth is not None:
                context_bandwidth = bandwidth
            if channel is not None:
                context_channel = channel
            if rssi is not None:
                context_rssi = rssi
            if scenario is not None:
                context_scenario = scenario

            print(
                f"[PEAK_DEBUG] Row {r} context: freq={context_frequency}, mode={context_wifi_mode}, bw={context_bandwidth}, ch={context_channel}, rssi={context_rssi}")

            # 解析频段和带宽
            band = _parse_band_token(context_frequency)
            bw_mhz = _parse_bandwidth_mhz(context_bandwidth)

            print(f"[PEAK_DEBUG] Row {r} parsed: band={band}, bw_mhz={bw_mhz}, channel={context_channel}")

            # 验证必要字段（现在只需要 band, bw_mhz, channel 存在即可）
            if not band or bw_mhz is None or context_channel is None:
                print(f"[PEAK_DEBUG] Row {r} missing required fields, skipping")
                continue

            # 处理 TX 数据 - uplink
            if tx_mbps is not None and isinstance(tx_mbps, (int, float)) and not isinstance(tx_mbps, bool):
                direction = "uplink"
                throughput_mbps = float(tx_mbps)

                # 使用 scenario 作为 mode，如果没有则用默认值
                mode_token = self._normalize_peak_mode(context_scenario) if context_scenario else "default"
                unique = (band, int(bw_mhz), int(context_channel), direction, mode_token)

                if unique not in seen_mode:
                    seen_mode.add(unique)
                    print(f"[PEAK_DEBUG] Row {r} - Processing TX: {throughput_mbps} Mbps")

                    # 创建简单的 scenario_group_key，不需要复杂的场景匹配
                    scenario_group_key = f"{band}_{int(bw_mhz)}_{int(context_channel)}"

                    rows.append(
                        _performance_row(
                            test_category=TEST_REPORT_PEAK_THROUGHPUT,
                            standard=_parse_standard(context_wifi_mode) or context_wifi_mode or "unknown",
                            band=band,
                            bandwidth_mhz=int(bw_mhz),
                            channel=int(context_channel),
                            protocol="TCP",  # Peak Throughput 默认使用 TCP
                            mode=mode_token,
                            direction=direction,
                            path_loss_db=0.0,
                            rssi=float(context_rssi) if isinstance(context_rssi, (int, float)) and not isinstance(
                                context_rssi, bool) else None,
                            angle_deg=None,
                            throughput_mbps=throughput_mbps,
                            throughput_peak_mbps=throughput_mbps,
                            scenario_group_key=scenario_group_key,
                        )
                    )

            # 处理 RX 数据 - downlink
            if rx_mbps is not None and isinstance(rx_mbps, (int, float)) and not isinstance(rx_mbps, bool):
                direction = "downlink"
                throughput_mbps = float(rx_mbps)

                mode_token = self._normalize_peak_mode(context_scenario) if context_scenario else "default"
                unique = (band, int(bw_mhz), int(context_channel), direction, mode_token)

                if unique not in seen_mode:
                    seen_mode.add(unique)
                    print(f"[PEAK_DEBUG] Row {r} - Processing RX: {throughput_mbps} Mbps")

                    scenario_group_key = f"{band}_{int(bw_mhz)}_{int(context_channel)}"

                    rows.append(
                        _performance_row(
                            test_category=TEST_REPORT_PEAK_THROUGHPUT,
                            standard=_parse_standard(context_wifi_mode) or context_wifi_mode or "unknown",
                            band=band,
                            bandwidth_mhz=int(bw_mhz),
                            channel=int(context_channel),
                            protocol="TCP",
                            mode=mode_token,
                            direction=direction,
                            path_loss_db=0.0,
                            rssi=float(context_rssi) if isinstance(context_rssi, (int, float)) and not isinstance(
                                context_rssi, bool) else None,
                            angle_deg=None,
                            throughput_mbps=throughput_mbps,
                            throughput_peak_mbps=throughput_mbps,
                            scenario_group_key=scenario_group_key,
                        )
                    )

        print(f"[PEAK_DEBUG] Finished parsing. Total rows generated: {len(rows)}")
        return rows, []

    @staticmethod
    def _normalize_peak_mode(value: Any) -> str:
        text = "" if value is None else str(value).strip().upper()
        if not text:
            return ""
        if "BLE" in text and "CLASSIC" in text:
            return "Wi-Fi BLE CLASSIC"
        if "BLE" in text:
            return "Wi-Fi BLE"
        if "WIFI" in text:
            return "Wi-Fi Only"
        return ""

    def build_rvr_rows(self, workbook, *, sheet_name: str = "RVR") -> Tuple[List[Dict[str, Any]], List[str]]:
        ws = workbook[sheet_name]
        print(f"[RVR_DEBUG] Parsing sheet: {sheet_name}, max_row={ws.max_row}, max_col={ws.max_column}")
        print(f"[RVR_DEBUG] Total scenarios loaded: {len(self._scenarios)}")
        rows: List[Dict[str, Any]] = []
        allowed_channels_by_band: Dict[str, set[int]] = {}
        for spec in self._scenarios:
            allowed_channels_by_band.setdefault(spec.band_token, set()).add(int(spec.channel))

        # === 新增的辅助函数：用于添加单向数据行 ===
        def _append_simplex_row(
                *,
                standard: str,
                band: str,
                bandwidth_mhz: int,
                channel: int,
                path_loss_db: float,
                angle_deg: float,
                direction: str,  # "downlink" for RX, "uplink" for TX
                throughput_mbps: Optional[float],
                rssi: Optional[float],
                scenario_group_key: str,
        ) -> None:
            rows.append(
                _performance_row(
                    test_category=TEST_REPORT_RVR,
                    standard=standard,
                    band=band,
                    bandwidth_mhz=bandwidth_mhz,
                    channel=channel,
                    protocol="TCP",
                    mode=None,
                    direction=direction,
                    path_loss_db=path_loss_db,
                    angle_deg=angle_deg,
                    rssi=rssi,
                    throughput_mbps=throughput_mbps,
                    throughput_peak_mbps=None,
                    scenario_group_key=scenario_group_key,
                )
            )

        # =========================================

        def safe_get_cell(ws, row, col):
            """安全获取单元格值，避免 openpyxl 抛出 'Row or column values must be at least 1' 错误"""
            if row < 1 or col < 1:
                return None
            try:
                return ws.cell(row, col).value
            except Exception:
                return None

        def find_title_row(start_row: int) -> Optional[str]:
            print(f"[RVR_TITLE_DEBUG] Looking for title above header row {start_row} (checking up to 7 rows)")
            for offset in range(1, 8):
                r = start_row - offset
                if r <= 1:
                    break
                v = safe_get_cell(ws, r, 1)  # 只检查第一列 (A列)
                print(f"[RVR_TITLE_DEBUG] Checking row {r}, col A: {repr(v)}")  # 调试日志
                if not isinstance(v, str):
                    continue
                text = v.strip()
                if not text:
                    continue
                upper = text.upper()
                if "2.4G" in upper or "5G" in upper or "6G" in upper:
                    print(f"[RVR_TITLE_DEBUG] Title FOUND at row {r}: '{text}'")
                    return text
            print(f"[RVR_TITLE_DEBUG] No valid title found above header row {start_row}")
            return None

        def parse_title(text: str) -> Tuple[str, str, int]:
            upper = (text or "").strip().upper()
            if "2.4G" in upper:
                band = "2.4G"
            elif "5G" in upper:
                band = "5G"
            elif "6G" in upper:
                band = "6G"
            else:
                band = ""

            if "11AX" in upper or "HE" in upper:
                standard = "11ax"
            elif "11AC" in upper or "VHT" in upper:
                standard = "11ac"
            elif "11N" in upper or "HT" in upper:
                standard = "11n"
            else:
                standard = ""

            bw = 0
            for token, value in (("HE20", 20), ("HT20", 20), ("HE40", 40), ("HT40", 40), ("VHT80", 80), ("HE80", 80),
                                 ("HE160", 160)):
                if token in upper:
                    bw = value
                    break
            if bw == 0:
                if "20M" in upper:
                    bw = 20
                elif "40M" in upper:
                    bw = 40
                elif "80M" in upper:
                    bw = 80
                elif "160M" in upper:
                    bw = 160
            return band, standard, bw

        def parse_channel_label(value: Any) -> Optional[int]:
            if not isinstance(value, str):
                return None
            text = value.strip().upper()
            if not text.startswith("CH"):
                return None
            digits = "".join(ch for ch in text if ch.isdigit())
            return int(digits) if digits else None

        def find_rssi_item_col(header_row: int) -> Optional[int]:
            for c in range(2, ws.max_column + 1):
                v = ws.cell(header_row, c).value
                if v != "Item":
                    continue
                for probe in range(c, min(ws.max_column, c + 15) + 1):
                    cell = ws.cell(header_row, probe).value
                    if isinstance(cell, str) and "RSSI" in cell.upper():
                        return c
            return None

        for header_row in range(1, ws.max_row + 1):
            try:
                cell_a_value = ws.cell(header_row, 1).value
                print(f"[RVR_FULL_SCAN] Row {header_row}, Col A: {repr(cell_a_value)}")

                if ws.cell(header_row, 1).value != "Item":
                    continue
                cell_b_value = ws.cell(header_row, 2).value
                cell_c_value = ws.cell(header_row, 3).value
                print(
                    f"[RVR_FULL_SCAN] Found 'Item' at row {header_row}. Checking B/C cols: B={repr(cell_b_value)}, C={repr(cell_c_value)}")
                if not isinstance(ws.cell(header_row, 2).value, str):
                    continue
                if not isinstance(ws.cell(header_row, 3).value, str):
                    continue

                rx_header_col = None
                tx_header_col = None
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(header_row, c).value
                    if not isinstance(v, str):
                        continue
                    upper = v.upper()
                    if rx_header_col is None and "RX" in upper and "MBPS" in upper:
                        rx_header_col = c
                    if tx_header_col is None and "TX" in upper and "MBPS" in upper:
                        tx_header_col = c
                print(f"[RVR_FULL_SCAN] RX/TX columns found: RX={rx_header_col}, TX={tx_header_col}")
                # === 修改点: 不再强制要求 RX 和 TX 同时存在 ===
                if rx_header_col is None and tx_header_col is None:
                    continue  # 如果两者都没有，才跳过
                # ===========================================

                title = find_title_row(header_row)
                if not title:
                    continue
                band, standard, bw_mhz = parse_title(title)
                if not band or bw_mhz <= 0:
                    continue
                # if band == "2.4" and bw_mhz == 20:
                #     continue

                allowed_channels = allowed_channels_by_band.get(band, set())
                if not allowed_channels:
                    continue

                channel_row = header_row + 1

                # === 修改点: 分别构建 RX 和 TX 的通道映射 ===
                rx_cols: Dict[int, int] = {}
                tx_cols: Dict[int, int] = {}

                # 确定 RX 列的结束位置
                rx_end_col = tx_header_col if tx_header_col is not None else (ws.max_column + 1)
                for c in range(rx_header_col or 1, rx_end_col):
                    ch = parse_channel_label(ws.cell(channel_row, c).value)
                    if ch is not None and ch in allowed_channels and ch not in rx_cols:
                        rx_cols[ch] = c

                # 确定 TX 列的起始位置
                tx_start_col = tx_header_col or (rx_header_col + len(rx_cols) if rx_header_col else 1)
                for c in range(tx_start_col, ws.max_column + 1):
                    ch = parse_channel_label(ws.cell(channel_row, c).value)
                    if ch is not None and ch in allowed_channels and ch not in tx_cols:
                        tx_cols[ch] = c
                    if ch is None and tx_cols:
                        break

                # 移除旧的、过于严格的检查
                # if not rx_cols or not tx_cols:
                #     logging.warning(f"[RVR_DEBUG] Missing RX/TX columns at header row {header_row}")
                #     continue
                # ===========================================

                rssi_item_col = find_rssi_item_col(header_row)
                rssi_rx_header_col = None
                rssi_tx_header_col = None
                rssi_rx_cols: Dict[int, int] = {}
                rssi_tx_cols: Dict[int, int] = {}
                if rssi_item_col is not None:
                    for c in range(rssi_item_col, ws.max_column + 1):
                        v = ws.cell(header_row, c).value
                        if not isinstance(v, str):
                            continue
                        upper = v.upper().replace(" ", "_")
                        if rssi_rx_header_col is None and "RX_RSSI" in upper:
                            rssi_rx_header_col = c
                        if rssi_tx_header_col is None and "TX_RSSI" in upper:
                            rssi_tx_header_col = c
                    if rssi_rx_header_col is not None and rssi_tx_header_col is not None:
                        for c in range(rssi_rx_header_col, rssi_tx_header_col):
                            ch = parse_channel_label(ws.cell(channel_row, c).value)
                            if ch is not None and ch in allowed_channels and ch not in rssi_rx_cols:
                                rssi_rx_cols[ch] = c
                        for c in range(rssi_tx_header_col, ws.max_column + 1):
                            ch = parse_channel_label(ws.cell(channel_row, c).value)
                            if ch is not None and ch in allowed_channels and ch not in rssi_tx_cols:
                                rssi_tx_cols[ch] = c
                            if ch is None and rssi_tx_cols:
                                break

                data_row = header_row + 2
                current_angle: Optional[float] = None
                while data_row <= ws.max_row:
                    att = ws.cell(data_row, 2).value
                    if att is None:
                        break
                    att_value = int(float(att))
                    parsed_angle = _parse_angle(ws.cell(data_row, 3).value)
                    if parsed_angle is not None:
                        current_angle = float(parsed_angle)
                    angle_value = float(current_angle) if current_angle is not None else 180.0

                    # === 修改点: 分别处理 RX 和 TX 数据 ===
                    # 处理 RX 数据
                    if rx_header_col is not None:
                        for ch, rx_col in rx_cols.items():
                            rx_value = ws.cell(data_row, rx_col).value
                            rx_throughput = (
                                float(rx_value) if isinstance(rx_value, (int, float)) and not isinstance(rx_value,
                                                                                                         bool) else None
                            )
                            rx_rssi = None
                            if rssi_rx_cols:
                                col = rssi_rx_cols.get(ch)
                                if col is not None and col >= 1:
                                    v = safe_get_cell(ws, data_row, col)
                                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                                        rx_rssi = float(v)

                            spec = self._scenario_by_channel.get((band, ch))
                            if spec is None:
                                logging.warning(
                                    f"[RVR_DEBUG] No predefined scenario found for band='{band}' and channel={ch}. "
                                    f"This channel will be skipped. "
                                    f"Check if BAND_CHANNEL_MAP or _generate_default_scenarios needs to be updated.")
                                continue
                            scenario_key = build_peak_scenario_group_key(
                                spec,
                                standard=standard,
                                bandwidth_mhz=bw_mhz,
                            )
                            _append_simplex_row(
                                standard=standard,
                                band=band,
                                bandwidth_mhz=bw_mhz,
                                channel=ch,
                                path_loss_db=float(att_value),
                                angle_deg=angle_value,
                                direction="downlink",
                                throughput_mbps=rx_throughput,
                                rssi=rx_rssi,
                                scenario_group_key=scenario_key,
                            )
                            print(
                                f"[RVR_DEBUG] Channel {ch} - RX: {rx_value} (col {rx_col})")

                    # 处理 TX 数据
                    if tx_header_col is not None:
                        for ch, tx_col in tx_cols.items():
                            tx_value = ws.cell(data_row, tx_col).value
                            tx_throughput = (
                                float(tx_value) if isinstance(tx_value, (int, float)) and not isinstance(tx_value,
                                                                                                         bool) else None
                            )
                            tx_rssi = None
                            if rssi_tx_cols:
                                col = rssi_tx_cols.get(ch)
                                if col is not None and col >= 1:
                                    v = safe_get_cell(ws, data_row, col)
                                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                                        tx_rssi = float(v)

                            spec = self._scenario_by_channel.get((band, ch))
                            if spec is None:
                                logging.warning(
                                    f"[RVR_DEBUG] No predefined scenario found for band='{band}' and channel={ch}. "
                                    f"This channel will be skipped. "
                                    f"Check if BAND_CHANNEL_MAP or _generate_default_scenarios needs to be updated.")
                                continue
                            scenario_key = build_peak_scenario_group_key(
                                spec,
                                standard=standard,
                                bandwidth_mhz=bw_mhz,
                            )
                            _append_simplex_row(
                                standard=standard,
                                band=band,
                                bandwidth_mhz=bw_mhz,
                                channel=ch,
                                path_loss_db=float(att_value),
                                angle_deg=angle_value,
                                direction="uplink",
                                throughput_mbps=tx_throughput,
                                rssi=tx_rssi,
                                scenario_group_key=scenario_key,
                            )
                            print(
                                f"[RVR_DEBUG] Channel {ch} - TX: {tx_value} (col {tx_col})")
                    # ===================================

                    data_row += 1
            except Exception as e:
                logging.error(f"[RVR_DEBUG] Error processing block starting at header row {header_row}: {e}",
                              exc_info=True)

        return rows, []

    def build_rvo_rows(self, workbook, *, sheet_name: str = "RVO") -> Tuple[List[Dict[str, Any]], List[str]]:
        ws = workbook[sheet_name]
        rows: List[Dict[str, Any]] = []
        print(f"[RVO_DEBUG] Parsing sheet: {sheet_name}, max_row={ws.max_row}, max_col={ws.max_column}")
        print(f"[RVO_DEBUG] Total scenarios loaded: {len(self._scenarios)}")

        def parse_bandwidth_mhz(bandwidth_str: str) -> int:
            if not bandwidth_str:
                return 0
            clean_str = bandwidth_str.replace(" ", "").replace("MHz", "").replace("MHZ", "")
            digits = "".join(ch for ch in clean_str if ch.isdigit())
            return int(digits) if digits else 0

        required_by_band_bw: Dict[Tuple[str, int], List[int]] = {}
        for spec in self._scenarios:
            bw_mhz = parse_bandwidth_mhz(spec.bandwidth)  # 正确解析 bandwidth 字符串
            if bw_mhz > 0:
                required_by_band_bw.setdefault((spec.band_token, bw_mhz), []).append(spec.channel)

        def normalize_text(value: Any) -> str:
            return "" if value is None else str(value).strip().upper().replace(" ", "")

        def parse_title(value: Any) -> Optional[Tuple[str, int, str]]:
            if not isinstance(value, str):
                return None
            upper = value.strip().upper()
            if not upper:
                return None

            band = "2.4G" if "2.4G" in upper else "5G" if "5G" in upper else "6G" if "6G" in upper else ""
            if not band:
                return None

            bw = 0
            for token, value in (("HE20", 20), ("HT20", 20), ("HE40", 40), ("HT40", 40), ("VHT80", 80), ("HE80", 80),
                                 ("HE160", 160)):
                if token in upper:
                    bw = value
                    break
            if bw == 0:
                if "20M" in upper:
                    bw = 20
                elif "40M" in upper:
                    bw = 40
                elif "80M" in upper:
                    bw = 80
                elif "160M" in upper:
                    bw = 160

            if not bw:
                return None

            standard = "11ax" if ("11AX" in upper or "HE" in upper) else "11ac" if ("11AC" in upper or "VHT" in upper) else "11n" if ("11N" in upper or "HT" in upper) else ""
            return band, bw, standard

        def find_title_above(row: int) -> Optional[Tuple[str, int, str]]:
            for r in range(row - 1, max(0, row - 40), -1):
                parsed = parse_title(ws.cell(r, 1).value)
                if parsed is not None:
                    print(f"[RVO_DEBUG] Found title at row {r}: {ws.cell(r, 1).value} -> {parsed}")
                    return parsed
            print(f"[RVO_DEBUG] No title found above row {row}")
            return None

        def parse_att_db(value: Any) -> Optional[float]:
            if value is None:
                return None
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                return float(value)
            text = str(value).strip().upper()
            if not text:
                return None
            digits: list[str] = []
            dot_seen = False
            sign_seen = False
            for ch in text:
                if ch in "+-" and not digits and not sign_seen:
                    digits.append(ch)
                    sign_seen = True
                    continue
                if ch.isdigit():
                    digits.append(ch)
                    continue
                if ch == "." and not dot_seen:
                    digits.append(ch)
                    dot_seen = True
                    continue
                if digits:
                    break
            number = "".join(digits).strip(".")
            return float(number) if number else None

        metric_tokens = (
            ("RX(UNIT:MBPS)", "downlink"),
            ("RXUNIT:MBPS", "downlink"),
            ("TX(UNIT:MBPS)", "uplink"),
            ("TXUNIT:MBPS", "uplink"),
            ("TX(UNIT:MB)", "uplink"),
            ("TXUNIT:MB", "uplink"),
        )
        print(f"[RVO_DEBUG] Starting to scan for headers in {ws.max_row} rows")
        for header_row in range(1, ws.max_row):
            title = find_title_above(header_row)
            if title is None:
                continue
            band, bw_mhz, standard = title
            print(
                f"[RVO_DEBUG] Processing header row {header_row} with title: band={band}, bw={bw_mhz}, std={standard}")
            required_channels = required_by_band_bw.get((band, bw_mhz), [])
            print(f"[RVO_DEBUG] Required channels for ({band}, {bw_mhz}): {required_channels}")
            if not required_channels:
                continue

            for metric_col in range(1, ws.max_column + 1):
                metric_cell = normalize_text(ws.cell(header_row, metric_col).value)
                direction = ""
                for token, resolved_direction in metric_tokens:
                    direction = resolved_direction if token in metric_cell else direction
                if not direction:
                    continue

                ch_col = 0
                att_col = 0
                for c in range(max(1, metric_col - 10), metric_col):
                    if normalize_text(ws.cell(header_row, c).value) == "CH":
                        ch_col = c
                    if normalize_text(ws.cell(header_row, c).value).endswith("ATT"):
                        att_col = c
                if not ch_col:
                    continue
                if not att_col:
                    att_col = ch_col + 1

                angle_row = header_row + 1
                angles: List[Tuple[int, float]] = []
                for c in range(metric_col, ws.max_column + 1):
                    angle_value = _parse_angle(ws.cell(angle_row, c).value)
                    if angle_value is None:
                        if angles:
                            break
                        continue
                    angles.append((c, float(angle_value)))
                if not angles:
                    continue

                current_channel: Optional[int] = None
                for data_row in range(angle_row + 1, ws.max_row + 1):
                    if normalize_text(ws.cell(data_row, ch_col).value) == "CH" and ws.cell(data_row, metric_col).value is not None:
                        break

                    ch_value = ws.cell(data_row, ch_col).value
                    if isinstance(ch_value, str):
                        ch_text = ch_value.strip().upper()
                        digits = "".join(ch for ch in ch_text if ch.isdigit()) if ch_text.startswith("CH") else ""
                        current_channel = int(digits) if digits else None
                    channel = current_channel
                    if channel is None or channel not in required_channels:
                        continue

                    att_db = parse_att_db(ws.cell(data_row, att_col).value)
                    if att_db is None:
                        continue

                    spec = self._scenario_by_channel[(band, channel)]
                    scenario_key = build_peak_scenario_group_key(spec, standard=standard, bandwidth_mhz=bw_mhz)
                    for col_idx, angle in angles:
                        val = ws.cell(data_row, col_idx).value
                        print(
                            f"[RVO_DEBUG] Row {data_row}, Col {col_idx}, Channel {channel}, Angle {angle}: raw_value={val}")
                        throughput = float(val) if isinstance(val, (int, float)) and not isinstance(val, bool) else None
                        if throughput is None:
                            print(f"[RVO_DEBUG] Skipping invalid throughput value: {val}")
                            continue
                        rows.append(
                            _performance_row(
                                test_category=TEST_REPORT_RVO,
                                standard=standard,
                                band=band,
                                bandwidth_mhz=bw_mhz,
                                channel=channel,
                                protocol="TCP",
                                mode=None,
                                direction=direction,
                                path_loss_db=float(att_db),
                                rssi=None,
                                angle_deg=float(angle),
                                throughput_mbps=throughput,
                                throughput_peak_mbps=None,
                                scenario_group_key=scenario_key,
                            )
                        )

        return rows, []

    def build_rows(
        self,
        path: str | Path,
        *,
        types: Iterable[str],
        throughput_sheet_name: Optional[str] = None,
        rvr_sheet_name: str = "RVR",
        rvo_sheet_name: str = "RVO",
    ) -> Tuple[Dict[str, List[Dict[str, Any]]], List[str]]:
        workbook = self._load_workbook(path)
        sheetnames = set(workbook.sheetnames)

        selected: list[str] = []
        seen: set[str] = set()
        for raw in types:
            text = str(raw).strip()
            if not text:
                continue
            key = text
            if key not in seen:
                selected.append(key)
                seen.add(key)

        def _resolve_peak_sheet() -> Optional[str]:
            if throughput_sheet_name:
                return throughput_sheet_name
            if "Peak Throughput" in sheetnames:
                return "Peak Throughput"
            excluded = {"Summary", "Test Setup", "RVR", "RVO", "MI_HW_cases-65", "BT-distance"}
            for name in workbook.sheetnames:
                if name not in excluded:
                    return name
            return None

        out: Dict[str, List[Dict[str, Any]]] = {}
        issues: List[str] = []
        peak_sheet: Optional[str] = None
        if TEST_REPORT_PEAK_THROUGHPUT in selected:
            peak_sheet = _resolve_peak_sheet()
            if not peak_sheet:
                issues.append(f"{TEST_REPORT_PEAK_THROUGHPUT}: missing throughput sheet")
            elif peak_sheet not in sheetnames:
                issues.append(f"{TEST_REPORT_PEAK_THROUGHPUT}: worksheet {peak_sheet!r} not found")
                peak_sheet = None

        for key in selected:
            if key == TEST_REPORT_PEAK_THROUGHPUT:
                if not peak_sheet:
                    continue
                try:
                    rows, row_issues = self.build_peak_throughput_rows(workbook, sheet_name=peak_sheet)
                except Exception as exc:
                    issues.append(f"{TEST_REPORT_PEAK_THROUGHPUT}: parse failed ({exc})")
                    continue
            elif key == TEST_REPORT_RVR:
                if rvr_sheet_name not in sheetnames:
                    issues.append(f"RVR: worksheet {rvr_sheet_name!r} not found")
                    continue
                try:
                    rows, row_issues = self.build_rvr_rows(workbook, sheet_name=rvr_sheet_name)
                except Exception as exc:
                    issues.append(f"RVR: parse failed ({exc})")
                    continue
            elif key == TEST_REPORT_RVO:
                if rvo_sheet_name not in sheetnames:
                    issues.append(f"RVO: worksheet {rvo_sheet_name!r} not found")
                    continue
                try:
                    rows, row_issues = self.build_rvo_rows(workbook, sheet_name=rvo_sheet_name)
                except Exception as exc:
                    issues.append(f"RVO: parse failed ({exc})")
                    continue
            else:
                issues.append(f"Unknown import type: {key}")
                continue

            out[key] = rows
            if not rows:
                issues.append(f"{key}: no rows parsed")
            issues.extend(row_issues)

        return out, issues

    def build_sheet_batches(
        self,
        path: str | Path,
        *,
        sheet_entries: Sequence[tuple[str, str]],
    ) -> Tuple[List[SheetImportBatch], List[str]]:
        workbook = self._load_workbook(path)
        sheetnames = set(workbook.sheetnames)
        batches: List[SheetImportBatch] = []
        issues: List[str] = []
        print(f"[DEBUG] Actual sheet names in the file: {list(workbook.sheetnames)}")

        for report_type, raw_sheet_name in sheet_entries:
            normalized_type = str(report_type).strip()
            sheet_name = str(raw_sheet_name).strip()
            if not normalized_type or not sheet_name:
                continue
            if sheet_name not in sheetnames:
                issues.append(f"{normalized_type}: worksheet {sheet_name!r} not found")
                continue
            try:
                if normalized_type == TEST_REPORT_PEAK_THROUGHPUT:
                    rows, row_issues = self.build_peak_throughput_rows(workbook, sheet_name=sheet_name)
                elif normalized_type == TEST_REPORT_RVR:
                    rows, row_issues = self.build_rvr_rows(workbook, sheet_name=sheet_name)
                elif normalized_type == TEST_REPORT_RVO:
                    rows, row_issues = self.build_rvo_rows(workbook, sheet_name=sheet_name)
                else:
                    issues.append(f"Unknown import type: {normalized_type}")
                    continue
            except Exception as exc:
                if 'rows' in locals() and rows:
                    batches.append(SheetImportBatch(
                        report_type=normalized_type,
                        sheet_name=sheet_name,
                        rows=rows,
                    ))
                    issues.append(f"{normalized_type}: partial parse with error ({exc})")
                else:
                    issues.append(f"{normalized_type}: parse failed ({exc})")
                continue

            batches.append(
                SheetImportBatch(
                    report_type=normalized_type,
                    sheet_name=sheet_name,
                    rows=rows,
                )
            )
            issues.extend(f"{sheet_name}: {issue}" for issue in row_issues)

        return batches, issues


class ImportController:
    def __init__(self, main_window) -> None:
        self._main = main_window

    @staticmethod
    def _format_tester_name(value: str) -> str:
        text = str(value).strip()
        if not text:
            return ""
        lowered = text.lower()
        if "." in lowered:
            return lowered
        parts = [p for p in text.replace("\t", " ").split(" ") if p.strip()]
        if len(parts) >= 2:
            return f"{parts[0]}.{parts[-1]}".lower()
        return lowered

    def run_import(self) -> None:
        dialog = ImportDialog(self._main)
        if dialog.exec_() != dialog.Accepted:
            return
        import_as_golden = dialog.import_as_golden()
        print("[IMPORT_DEBUG] import_as_golden=", int(import_as_golden))

        file_path, _ = QFileDialog.getOpenFileName(
            self._main,
            "Select Excel file to import",
            str(Path.cwd()),
            "Excel Files (*.xlsx)",
        )
        if not file_path:
            return

        importer = PerformanceExcelImporter()
        workbook = importer._load_workbook(file_path)
        summary_name = self._find_summary_sheet_name(workbook.sheetnames)
        summary_payload: dict[str, str] = {}
        if summary_name is not None:
            summary_payload = self._parse_summary_payload(workbook[summary_name])
        tester = str(summary_payload.get("tester") or "").strip()
        summary_project_id = str(summary_payload.get("project_id") or "").strip()
        if tester:
            print("[IMPORT_DEBUG] tester=", tester)
        if summary_project_id:
            print("[IMPORT_DEBUG] summary_project_id=", summary_project_id)
        sheet_candidates = [name for name in workbook.sheetnames if "SUMMARY" not in str(name).strip().upper()]

        sheets_box = ImportSheetsDialog(
            self._main,
            summary_lines=[f"Excel: {Path(file_path).name}", f"Golden: {1 if import_as_golden else 0}"],
            sheet_names=sheet_candidates,
            on_import=None,
        )
        inserted_holder: dict[str, int] = {"value": 0}
        issues_holder: dict[str, list[str]] = {"value": []}
        skipped_holder: dict[str, list[str]] = {"value": []}
        failed_holder: dict[str, str] = {"value": ""}

        def start_import(selected_sheets: list[str]) -> None:
            print("[IMPORT_DEBUG] selected_sheets=", selected_sheets)

            sheet_entries: list[tuple[str, str]] = []
            for sheet_name in selected_sheets:
                token = str(sheet_name).strip().upper()
                if "RVR" in token:
                    sheet_entries.append((TEST_REPORT_RVR, str(sheet_name)))
                    continue
                if "RVO" in token:
                    sheet_entries.append((TEST_REPORT_RVO, str(sheet_name)))
                    continue
                sheet_entries.append((TEST_REPORT_PEAK_THROUGHPUT, str(sheet_name)))

            print(
                "[IMPORT_DEBUG] resolved_sheet_entries=",
                sheet_entries,
            )

            try:
                sheet_batches, issues = importer.build_sheet_batches(
                    file_path,
                    sheet_entries=sheet_entries,
                )
            except Exception as exc:
                failed_holder["value"] = str(exc)
                sheets_box.reject()
                return

            payload = self._build_import_payload(
                {},
                summary_payload or {},
                summary_project_id=summary_project_id or "",
            )
            if tester:
                payload = dict(payload)
                payload["tester"] = tester

            summary_issues = self._validate_summary_payload(payload)
            summary_issues.extend(self._validate_import_project_payload(payload))
            if summary_issues:
                MessageBox(
                    "Import blocked",
                    "SUMMARY sheet contains unsupported values:\n\n"
                    + "\n".join(f"- {issue}" for issue in summary_issues),
                    self._main,
                ).exec()
                sheets_box.set_loading(False)
                sheets_box.reject()
                return

            resolved_project_id = None
            project_key = str(payload.get("project_id") or "").strip()
            if not project_key:
                MessageBox("Import blocked", "Missing Project ID in SUMMARY sheet.", self._main).exec()
                sheets_box.set_loading(False)
                sheets_box.reject()
                return
            with MySqlClient() as client:
                ensure_report_tables(client)
                if str(project_key or "").strip().upper() == "NA":
                    row = None
                else:
                    row = client.query_one(
                        "SELECT `id` FROM `project` WHERE `project_id`=%s ORDER BY `id` DESC LIMIT 1",
                        (project_key,),
                    )
                if row and row.get("id"):
                    resolved_project_id = int(row["id"])
                else:
                    resolved_project_id = int(ensure_project(client, payload))
            print("[IMPORT_DEBUG] resolved_project_id=", int(resolved_project_id), flush=True)
            payload = dict(payload)
            payload["resolved_project_id"] = str(int(resolved_project_id))

            def do_import() -> tuple[int, list[str]]:
                if import_as_golden:
                    return self._sync_golden_to_db(payload, file_path, sheet_batches, project_id=int(resolved_project_id)), []
                return self._sync_report_to_db(
                    payload,
                    file_path,
                    sheet_batches,
                    project_id=int(resolved_project_id),
                )

            class _ImportWorker(QObject):
                finished = pyqtSignal(object)
                failed = pyqtSignal(str)

                def __init__(self, fn):
                    super().__init__()
                    self._fn = fn

                def run(self) -> None:
                    try:
                        result = self._fn()
                    except Exception as exc:
                        self.failed.emit(str(exc))
                        return
                    self.finished.emit(result)

            thread = QThread(self._main)
            worker = _ImportWorker(do_import)
            worker.moveToThread(thread)

            def _cleanup() -> None:
                thread.quit()
                thread.wait()
                worker.deleteLater()
                thread.deleteLater()

            def _on_failed(message: str) -> None:
                failed_holder["value"] = message
                _cleanup()
                sheets_box.set_loading(False)
                sheets_box.reject()

            def _on_finished(result: object) -> None:
                inserted, skipped = result
                inserted_holder["value"] = int(inserted)
                issues_holder["value"] = issues
                skipped_holder["value"] = list(skipped)
                _cleanup()
                sheets_box.set_loading(False)
                sheets_box.accept()

            thread.started.connect(worker.run)
            worker.failed.connect(_on_failed)
            worker.finished.connect(_on_finished)
            thread.start()

        sheets_box._on_import = start_import
        if sheets_box.exec_() != sheets_box.Accepted:
            if failed_holder["value"]:
                MessageBox("Import failed", failed_holder["value"], self._main).exec()
            return

        inserted = inserted_holder["value"]
        issues = issues_holder["value"]
        skipped = skipped_holder["value"]

        if skipped and inserted == 0:
            MessageBox(
                "Import skipped",
                "All selected worksheets were already imported.\n\nAlready imported:\n"
                + "\n".join(f"- {t}" for t in skipped),
                self._main,
            ).exec()
        elif skipped:
            MessageBox(
                "Import notice",
                "Some selected worksheets were already imported and were skipped:\n"
                + "\n".join(f"- {t}" for t in skipped),
                self._main,
            ).exec()

        summary_lines = [f"Inserted {inserted} performance row(s)."]
        if issues:
            print(f"[SKIP] issues {issues}, Skip:{skipped}")
            summary_lines.append("")
            summary_lines.append("Some selected types were skipped/failed. See validation details above.")
        MessageBox("Import completed", "\n".join(summary_lines), self._main).exec()
        return

    @staticmethod
    def _find_summary_sheet_name(sheetnames: Sequence[object]) -> Optional[str]:
        for name in sheetnames:
            if "SUMMARY" in str(name).strip().upper():
                return str(name)
        return None

    @staticmethod
    def _validate_summary_payload(payload: Mapping[str, Any]) -> list[str]:
        validations: tuple[tuple[str, str, Sequence[str]], ...] = (
            ("Project ID", "project_id", PROJECT_ID_CHOICES),
            ("Product Type", "project_type", PROJECT_TYPES),
            ("WiFi Module", "wifi_module", WIFI_MODULE_CHOICES),
            ("HW Phase", "hw_phase", HW_PHASE_CHOICES),
            ("OS", "os", DUT_OS_CHOICES),
            ("BT Remote", "bt_remote", BT_REMOTE_CHOICES),
            ("BT Device", "bt_device", BT_DEVICE_CHOICES),
            ("BT Type", "bt_type", BT_TYPE_CHOICES),
            ("Connect Type", "lab_enviroment.connect_type", LAB_ENV_CONNECT_TYPE_CHOICES),
            ("Coex Mode", "lab_enviroment.coex_mode", LAB_ENV_COEX_MODE_CHOICES),
            ("AP Name", "ap_name", AP_MODEL_CHOICES),
            ("AP Region", "ap_region", AP_REGION_CHOICES),
        )
        issues: list[str] = []
        for label, key, allowed in validations:
            value = str(payload.get(key) or "").strip()
            if not value:
                continue
            if key == "lab_enviroment.connect_type":
                value = value.replace('\u2011', '-').replace('\u2013', '-').replace('\u2014', '-')
            if value not in allowed:
                issues.append(f"{label}: {value!r} not in {list(allowed)!r}")

        return issues

    def _parse_summary_payload(self, ws) -> dict[str, str]:
        label_aliases: dict[str, str] = {
            "TESTED BY": "tester",
            "TESTER": "tester",
            "PROJECT ID": "project_id",
            "PROJECTID": "project_id",
            "PROJECT NAME": "project_name",
            "PRODUCT TYPE": "project_type",
            "PROJECT TYPE": "project_type",
            "CUSTOMER": "customer",
            "ODM": "odm",
            "SOC": "soc",
            "WIFI MODULE": "wifi_module",
            "WIFI MODULE SN": "wifi_module_sn",
            "USB CABLE": "usb_cable",
            "HDMI CABLE": "hdmi_cable",
            "BT DEVICE": "bt_device",
            "BLUETOOTH DEVICE": "bt_device",
            "BT REMOTE": "bt_remote",
            "BT TYPE": "bt_type",
            "BLUETOOTH TYPE": "bt_type",
            "TV DEVICE": "tv_device",
            "SOFTWARE VERSION": "software_version",
            "SW VERSION": "software_version",
            "SW_VER": "software_version",
            "DRIVER VERSION": "driver_version",
            "WIFI DRIVER": "driver_version",
            "ANDROID VERSION": "android_version",
            "OS VERSION": "android_version",
            "KERNEL VERSION": "kernel_version",
            "KERNEL": "kernel_version",
            "OS": "os",
            "HW PHASE": "hw_phase",
            "MAC ADDRESS": "mac_address",
            "MAC": "mac_address",
            "CONNECT TYPE": "lab_enviroment.connect_type",
            "DUT TYPE": "lab_enviroment.connect_type",
            "COEX MODE": "lab_enviroment.coex_mode",
            "AP NAME": "ap_name",
            "ROUTER": "ap_name",
            "AP REGION": "ap_region",
            "ADB DEVICE": "adb_device",
            "DUT IP": "telnet_ip",
            "TELNET IP": "telnet_ip",
            "SERIAL NUMBER": "serial_number",
            "SERIAL": "serial_number",
            "SN": "serial_number",
            "NICKNAME": "nickname",
        }

        out: dict[str, str] = {}
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            for idx, cell in enumerate(row):
                label = str(cell or "").strip().upper()
                key = label_aliases.get(label)
                if not key:
                    continue
                value = ""
                for j in range(idx + 1, len(row)):
                    candidate = str(row[j] or "").strip()
                    if candidate:
                        value = candidate
                        break
                if not value:
                    continue
                if key == "tester":
                    out[key] = self._format_tester_name(value)
                else:
                    out[key] = value
        return out

    @staticmethod
    def _build_import_payload(
        default_payload: Mapping[str, Any],
        summary_payload: Mapping[str, Any],
        *,
        summary_project_id: str,
    ) -> dict[str, Any]:
        payload = dict(default_payload)
        summary_values = {k: v for k, v in summary_payload.items() if v not in (None, "")}
        normalized_project_id = str(summary_project_id or "").strip()
        logging.info(
            "[IMPORT_DEBUG] build_payload start project_id=%s default_project=%s summary_project=%s",
            normalized_project_id,
            {k: default_payload.get(k) for k in _PROJECT_PAYLOAD_KEYS if k in default_payload},
            {k: summary_values.get(k) for k in _PROJECT_PAYLOAD_KEYS if k in summary_values},
        )

        if normalized_project_id.upper() == "NA":
            for key in _PROJECT_PAYLOAD_KEYS:
                payload.pop(key, None)
            payload.update({k: v for k, v in summary_values.items() if k in _PROJECT_PAYLOAD_KEYS})
            payload["project_id"] = normalized_project_id
            #Handel NA nickname
            if "nickname" not in payload:
                payload["nickname"] = "NA"
            logging.info(
                "[IMPORT_DEBUG] build_payload NA project=%s",
                {k: payload.get(k) for k in _PROJECT_PAYLOAD_KEYS if k in payload},
            )
            return payload

        payload.update(summary_values)
        if normalized_project_id:
            payload["project_id"] = normalized_project_id
            for project_type, brands in WIFI_PRODUCT_PROJECT_MAP.items():
                for brand, projects in brands.items():
                    for _, info in projects.items():
                        if str(info.get("ProjectID") or "").strip() != normalized_project_id:
                            continue
                        payload["customer"] = str(brand)
                        payload["project_type"] = str(project_type)
                        payload["project_name"] = str(info.get("ProjectName") or payload.get("project_name") or "")
                        payload["soc"] = str(info.get("main_chip") or payload.get("soc") or "")
                        payload["wifi_module"] = str(info.get("wifi_module") or payload.get("wifi_module") or "")
                        payload["interface"] = str(info.get("interface") or payload.get("interface") or "")
                        payload["ecosystem"] = str(info.get("ecosystem") or payload.get("ecosystem") or "")
                        logging.info(
                            "[IMPORT_DEBUG] build_payload mapped project=%s",
                            {k: payload.get(k) for k in _PROJECT_PAYLOAD_KEYS if k in payload},
                        )
                        return payload
        logging.info(
            "[IMPORT_DEBUG] build_payload final project=%s",
            {k: payload.get(k) for k in _PROJECT_PAYLOAD_KEYS if k in payload},
        )
        return payload

    @staticmethod
    def _validate_import_project_payload(payload: Mapping[str, Any]) -> list[str]:
        issues: list[str] = []
        project_id = str(payload.get("project_id") or "").strip()
        if project_id.upper() != "NA":
            return issues
        for key, label in (
            ("customer", "Customer"),
            ("project_type", "Product Type"),
        ):
            value = str(payload.get(key) or "").strip()
            if not value:
                issues.append(f"{label}: required when Project ID is 'NA'")
        return issues

    def _sync_report_to_db(
            self,
            payload: Mapping[str, Any],
            excel_path: str,
            sheet_batches: Sequence[SheetImportBatch],
            *,
            project_id: int,
    ) -> tuple[int, list[str]]:
        """
        Modified to create a separate test_report for each unique report_type.
        This ensures that RVR and RVO data from the same file are stored in
        distinct reports with the correct report_type, making them queryable.
        """
        original_report_name = Path(excel_path).name
        notes = "\n".join([excel_path, "source=excel"])

        with MySqlClient() as client:
            ensure_report_tables(client)
            manager = PerformanceTableManager(client)

            skipped: list[str] = []
            inserted_total = 0

            # Group batches by their report_type
            batches_by_type: Dict[str, List[SheetImportBatch]] = {}
            for batch in sheet_batches:
                data_type = str(batch.report_type).strip()
                if data_type not in batches_by_type:
                    batches_by_type[data_type] = []
                batches_by_type[data_type].append(batch)

            # Process each report_type group independently
            for report_type, batches in batches_by_type.items():
                # Create a unique report name for this type, e.g., "original_file.xlsx_RVR"
                type_specific_report_name = f"{Path(original_report_name).stem}_{report_type}{Path(original_report_name).suffix}"

                # Check if a report of this specific type and name already exists for the project
                existing = client.query_one(
                    "SELECT `id` FROM `test_report` WHERE `project_id`=%s AND `report_name`=%s ORDER BY `id` DESC LIMIT 1",
                    (int(project_id), type_specific_report_name),
                )
                report_id = int(existing["id"]) if existing and existing.get("id") else None

                # Get the set of sheets already imported for this specific report
                imported_sheets: set[str] = set()
                if report_id is not None:
                    rows = client.query_all(
                        "SELECT DISTINCT `sheet_name` "
                        "FROM `execution` "
                        "WHERE `test_report_id`=%s AND `sheet_name` IS NOT NULL AND TRIM(`sheet_name`) <> ''",
                        (int(report_id),),
                    )
                    imported_sheets = {str(r.get("sheet_name") or "").strip() for r in rows if r.get("sheet_name")}

                print(
                    f"[IMPORT_DEBUG] Handling report type '{report_type}': "
                    f"report_name='{type_specific_report_name}', "
                    f"report_id={report_id}, "
                    f"imported_sheets={sorted(imported_sheets)}",
                )

                # If no report exists for this type, create one
                if report_id is None:
                    report_id = ensure_test_report(
                        client,
                        project_id=int(project_id),
                        report_name=type_specific_report_name,
                        case_path=None,
                        is_golden=False,
                        report_type=report_type,  # Correctly set the report_type
                        golden_group=None,
                        notes=notes,
                        tester=payload.get("tester"),
                    )
                    _store_excel_artifact(
                        client,
                        test_report_id=int(report_id),
                        excel_path=excel_path,
                    )

                # Process all batches of this specific report_type
                for batch in batches:
                    sheet_name = str(batch.sheet_name).strip()
                    rows = list(batch.rows)
                    print(
                        f"[IMPORT_DEBUG] Processing batch: {report_type} from sheet '{sheet_name}', rows count: {len(rows)}")
                    if not rows:
                        continue

                    # Skip if this exact sheet has already been imported into this specific report
                    if sheet_name in imported_sheets:
                        print(
                            f"[IMPORT_DEBUG] Skipping {sheet_name} ({report_type}) - already exists in report {report_id}")
                        skipped.append(f"{sheet_name} ({report_type})")
                        continue

                    execution_id = self._insert_execution(
                        client,
                        test_report_id=int(report_id),
                        project_id=int(project_id),
                        execution_type=report_type,
                        sheet_name=sheet_name,
                        csv_name=f"{Path(excel_path).name}:{report_type}",
                        csv_path=excel_path,
                        run_source="import",
                        payload={
                            "source": "excel",
                            "excel_path": excel_path,
                            "data_type": report_type,
                            "ui_payload": dict(payload),
                        },
                    )
                    inserted_total += _insert_performance_rows(
                        manager,
                        test_report_id=int(report_id),
                        execution_id=execution_id,
                        csv_name=original_report_name,
                        data_type=report_type,
                        rows=list(rows),
                    )
                    imported_sheets.add(sheet_name)

            if skipped:
                print("[IMPORT_DEBUG] skipped=", skipped)
            print("[IMPORT_DEBUG] inserted_total=", inserted_total)

            return inserted_total, skipped

    def _sync_golden_to_db(
        self,
        payload: Mapping[str, Any],
        excel_path: str,
        sheet_batches: Sequence[SheetImportBatch],
        *,
        project_id: int,
    ) -> int:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_digest = hashlib.sha256(Path(excel_path).read_bytes()).hexdigest()
        #excel_digest = "DEBUG_FORCE_DIFFERENT_" + excel_digest
        with MySqlClient() as client:
            ensure_report_tables(client)
            manager = PerformanceTableManager(client)
            inserted_total = 0

            for batch in sheet_batches:
                data_type = str(batch.report_type).strip()
                sheet_name = str(batch.sheet_name).strip()
                rows = list(batch.rows)
                if not rows:
                    continue

                existing = self._find_existing_golden_report(
                    client,
                    project_id=int(project_id),
                    report_type=str(data_type),
                    excel_digest=excel_digest,
                )
                if existing:
                    artifact = client.query_one(
                        "SELECT `sha256` FROM `artifact` WHERE `test_report_id`=%s ORDER BY `id` DESC LIMIT 1",
                        (int(existing["test_report_id"]),),
                    )
                    existing_sha = str((artifact or {}).get("sha256") or "")
                    if existing_sha and existing_sha == excel_digest:
                        # MessageBox(
                        #     "Import skipped",
                        #     f"Golden data already imported for type: {data_type}\n\nExisting report: {existing.get('report_name')}\nExcel: {Path(excel_path).name}",
                        #     self._main,
                        # ).exec()
                        print(f"[DEBUG] Would skip import for type: {data_type}. File is identical.")
                        continue
                if existing and not self._confirm_overwrite_existing(existing, report_type=str(data_type)):
                    continue
                if existing:
                    client.execute(
                        "DELETE FROM `test_report` WHERE `id`=%s",
                        (int(existing["test_report_id"]),),
                    )

                report_name = f"GOLDEN_{data_type}_{stamp}"
                notes = "\n".join([excel_path, f"type={data_type}"])
                report_id = ensure_test_report(
                    client,
                    project_id=int(project_id),
                    report_name=report_name,
                    case_path=None,
                    is_golden=True,
                    report_type=str(data_type),
                    golden_group="GOLDEN",
                    notes=notes,
                    tester=payload.get("tester"),
                )
                _store_excel_artifact(
                    client,
                    test_report_id=int(report_id),
                    excel_path=excel_path,
                )
                execution_id = self._insert_execution(
                    client,
                    test_report_id=int(report_id),
                    project_id=int(project_id),
                    execution_type=data_type,
                    sheet_name=sheet_name,
                    csv_name=f"{Path(excel_path).name}:{data_type}",
                    csv_path=excel_path,
                    run_source="import",
                    payload={
                        "source": "excel",
                        "excel_path": excel_path,
                        "data_type": data_type,
                        "ui_payload": dict(payload),
                    },
                )
                inserted_total += _insert_performance_rows(
                    manager,
                    test_report_id=int(report_id),
                    execution_id=execution_id,
                    csv_name=Path(excel_path).name,
                    data_type=data_type,
                    rows=list(rows),
                )
            print("[GOLDEN_IMPORT_DEBUG] inserted_total=", inserted_total)
            return inserted_total

    def _find_existing_golden_report(
        self,
        client: MySqlClient,
        *,
        project_id: int,
        report_type: str,
        excel_digest: Optional[str] = None,
    ) -> Optional[Dict[str, Any]]:
        if excel_digest:
            sql = (
                "SELECT "
                "tr.id AS test_report_id, tr.report_name, tr.created_at, tr.updated_at, "
                "a.file_name, a.created_at AS artifact_created_at, a.sha256 "
                "FROM test_report AS tr "
                "LEFT JOIN artifact AS a ON a.test_report_id = tr.id "
                "WHERE tr.project_id = %s AND tr.is_golden = 1 "
                "AND tr.report_type = %s AND tr.golden_group = %s "
                "AND a.sha256 = %s "  # 新增哈希匹配条件
                "ORDER BY tr.updated_at DESC, tr.id DESC "
                "LIMIT 1"
            )
            rows = client.query_all(sql, (int(project_id), report_type, "GOLDEN", excel_digest))
            if rows:
                return rows[0]
        else:
            sql = (
                "SELECT "
                "tr.id AS test_report_id, tr.report_name, tr.created_at, tr.updated_at, "
                "a.file_name, a.created_at AS artifact_created_at "
                "FROM test_report AS tr "
                "LEFT JOIN artifact AS a ON a.test_report_id = tr.id "
                "WHERE tr.project_id = %s AND tr.is_golden = 1 "
                "AND tr.report_type = %s AND tr.golden_group = %s "
                "ORDER BY tr.updated_at DESC, tr.id DESC "
                "LIMIT 1"
            )
            rows = client.query_all(sql, (int(project_id), report_type, "GOLDEN"))
            if rows:
                return rows[0]

        legacy_sql = (
            "SELECT "
            "tr.id AS test_report_id, tr.report_name, tr.created_at, tr.updated_at, "
            "a.file_name, a.created_at AS artifact_created_at, "
            "GROUP_CONCAT(DISTINCT ex.run_type ORDER BY ex.run_type) AS execution_types "
            "FROM test_report AS tr "
            "JOIN execution AS ex ON ex.test_report_id = tr.id "
            "LEFT JOIN artifact AS a ON a.test_report_id = tr.id "
            "WHERE tr.project_id = %s AND tr.is_golden = 1 "
            "AND (tr.report_type IS NULL OR tr.report_type = '') "
            "AND ex.run_type = %s "
            "GROUP BY tr.id "
            "ORDER BY tr.updated_at DESC, tr.id DESC "
            "LIMIT 1"
        )
        rows = client.query_all(legacy_sql, (int(project_id), report_type))
        if rows:
            row = dict(rows[0])
            row["legacy"] = True
            return row
        return None

    def _confirm_overwrite_existing(self, existing: Mapping[str, Any], *, report_type: str) -> bool:
        file_name = existing.get("file_name") or "(unknown)"
        uploaded_at = existing.get("artifact_created_at") or existing.get("updated_at") or existing.get("created_at")
        report_name = existing.get("report_name") or "(unknown)"
        legacy = bool(existing.get("legacy"))
        legacy_types = existing.get("execution_types") or ""

        details_lines = [
            f"Project already has golden data for type: {report_type}",
            f"Existing report: {report_name}",
            f"Existing excel: {file_name}",
            f"Uploaded at: {uploaded_at}",
        ]
        if legacy:
            details_lines.append("")
            details_lines.append("Legacy golden report detected.")
            if legacy_types:
                details_lines.append(f"This legacy report contains types: {legacy_types}")
            details_lines.append("Overwriting will delete the entire legacy report.")

        # box = MessageBox("Overwrite golden data?", "\n".join(details_lines), self._main)
        # box.yesButton.setText("Overwrite")
        # box.cancelButton.setText("Cancel")
        # box.exec()
        # return box.result() == QDialog.Accepted
        print(
            f"[DEBUG] Would ask to overwrite existing golden data for type: {report_type}. Auto-confirming 'Yes' for debug.")
        return True

    def _insert_execution(
        self,
        client: MySqlClient,
        *,
        test_report_id: int,
        project_id: int,
        execution_type: str,
        sheet_name: str,
        csv_name: str,
        csv_path: str,
        run_source: str,
        payload: Mapping[str, Any],
        duration_seconds: Optional[float] = None,
    ) -> int:
        ui_payload = payload.get("ui_payload", {}) or {}
        lab_id: int | None = None
        lab_name = str(ui_payload.get("lab_name") or "").strip()
        if lab_name:
            lab_id = client.insert(
                "INSERT INTO `lab` (`lab_name`) "
                "VALUES (%s) "
                "ON DUPLICATE KEY UPDATE "
                "`id`=LAST_INSERT_ID(`id`), "
                "`lab_name`=VALUES(`lab_name`)",
                (lab_name,),
            )
            client.execute(
                "UPDATE `test_report` SET `lab_id`=%s WHERE `id`=%s",
                (int(lab_id), int(test_report_id)),
            )

            env_payload = {
                "lab_id": int(lab_id),
                "ap_name": ui_payload.get("ap_name") or ui_payload.get("router_name"),
                "ap_address": ui_payload.get("ap_address") or ui_payload.get("router_address"),
                "distance": ui_payload.get("distance"),
                "ap_region": ui_payload.get("ap_region"),
                "connect_type": ui_payload.get("lab_enviroment.connect_type"),
                "coex_mode": ui_payload.get("lab_enviroment.coex_mode"),
                "bt_remote": ui_payload.get("bt_remote"),
                "usb_cable": ui_payload.get("usb_cable"),
                "hdmi_cable": ui_payload.get("hdmi_cable"),
                "bt_device": ui_payload.get("bt_device"),
                "bt_type": ui_payload.get("bt_type"),
                "tv_device": ui_payload.get("tv_device"),
            }
            ap_name = str(env_payload.get("ap_name") or "").strip()
            if ap_name and ap_name not in AP_MODEL_CHOICES:
                raise ValueError(
                    f"Unsupported lab_enviroment.ap_name={ap_name!r}; "
                    f"allowed={list(AP_MODEL_CHOICES)!r}"
                )
            ap_region = str(env_payload.get("ap_region") or "").strip()
            if ap_region and ap_region not in AP_REGION_CHOICES:
                raise ValueError(
                    f"Unsupported lab_enviroment.ap_region={ap_region!r}; "
                    f"allowed={list(AP_REGION_CHOICES)!r}"
                )
            client.insert(
                "INSERT INTO `lab_environment` "
                "(`lab_id`, `ap_name`, `ap_address`, `distance`, `ap_region`, `connect_type`, `coex_mode`, `bt_remote`, `usb_cable`, `hdmi_cable`, `bt_device`, `bt_type`, `tv_device`) "
                "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) "
                "ON DUPLICATE KEY UPDATE "
                "`lab_id`=VALUES(`lab_id`), "
                "`ap_name`=VALUES(`ap_name`), "
                "`ap_address`=VALUES(`ap_address`), "
                "`distance`=VALUES(`distance`), "
                "`ap_region`=VALUES(`ap_region`), "
                "`connect_type`=VALUES(`connect_type`), "
                "`coex_mode`=VALUES(`coex_mode`), "
                "`bt_remote`=VALUES(`bt_remote`), "
                "`usb_cable`=VALUES(`usb_cable`), "
                "`hdmi_cable`=VALUES(`hdmi_cable`), "
                "`bt_device`=VALUES(`bt_device`), "
                "`bt_type`=VALUES(`bt_type`), "
                "`tv_device`=VALUES(`tv_device`)",
                (
                    int(lab_id),
                    env_payload.get("ap_name"),
                    env_payload.get("ap_address"),
                    env_payload.get("distance"),
                    env_payload.get("ap_region"),
                    env_payload.get("connect_type"),
                    env_payload.get("coex_mode"),
                    env_payload.get("bt_remote"),
                    env_payload.get("usb_cable"),
                    env_payload.get("hdmi_cable"),
                    env_payload.get("bt_device"),
                    env_payload.get("bt_type"),
                    env_payload.get("tv_device"),
                ),
            )
        raw_sn = ui_payload.get("sn") or ui_payload.get("serial_number")
        mac_addr = ui_payload.get("mac_address")
        if raw_sn and str(raw_sn).strip().upper() == "NA":
            final_sn = mac_addr
        else:
            final_sn = raw_sn

        dut_payload = {
            "test_report_id": int(test_report_id),
            "sn": final_sn,
            "mac_address": mac_addr,
            "adb_device": ui_payload.get("adb_device"),
            "ip": ui_payload.get("ip") or ui_payload.get("telnet_ip"),
            "software_version": ui_payload.get("software_version"),
            "driver_version": ui_payload.get("driver_version"),
            "android_version": ui_payload.get("android_version"),
            "kernel_version": ui_payload.get("kernel_version"),
            "hw_phase": ui_payload.get("hw_phase"),
            "wifi_module_sn": ui_payload.get("wifi_module_sn"),
            "antenna": ui_payload.get("antenna"),
        }
        dut_id = client.insert(
            "INSERT INTO `dut` "
            "(`test_report_id`, `sn`, `mac_address`, `adb_device`, `ip`, "
            "`software_version`, `driver_version`, `android_version`, `kernel_version`, `hw_phase`, `wifi_module_sn`, `antenna`) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) "
            "ON DUPLICATE KEY UPDATE "
            "`id`=LAST_INSERT_ID(`id`), "
            "`test_report_id`=VALUES(`test_report_id`), "
            "`sn`=VALUES(`sn`), "
            "`mac_address`=VALUES(`mac_address`), "
            "`adb_device`=VALUES(`adb_device`), "
            "`ip`=VALUES(`ip`), "
            "`software_version`=VALUES(`software_version`), "
            "`driver_version`=VALUES(`driver_version`), "
            "`android_version`=VALUES(`android_version`), "
            "`kernel_version`=VALUES(`kernel_version`), "
            "`hw_phase`=VALUES(`hw_phase`), "
            "`wifi_module_sn`=VALUES(`wifi_module_sn`), "
            "`antenna`=VALUES(`antenna`)",
            (
                dut_payload.get("test_report_id"),
                dut_payload.get("sn"),
                dut_payload.get("mac_address"),
                dut_payload.get("adb_device"),
                dut_payload.get("ip"),
                dut_payload.get("software_version"),
                dut_payload.get("driver_version"),
                dut_payload.get("android_version"),
                dut_payload.get("kernel_version"),
                dut_payload.get("hw_phase"),
                dut_payload.get("wifi_module_sn"),
                dut_payload.get("antenna"),
            ),
        )
        insert_sql = (
            "INSERT INTO `execution` "
            "(`test_report_id`, `dut_id`, `sheet_name`, `run_type`, `run_source`, `duration_seconds`, `csv_name`, `execution_type`) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"
        )
        return client.insert(
            insert_sql,
            (
                test_report_id,
                int(dut_id),
                sheet_name,
                RUN_TYPE_WIFI_SMARTTEST,
                (run_source or "import")[:32],
                int(duration_seconds) if duration_seconds is not None else None,
                csv_name,
                execution_type,
            ),
        )
