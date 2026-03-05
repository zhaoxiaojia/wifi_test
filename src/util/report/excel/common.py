"""Common Excel helpers (openpyxl).

These helpers focus on *in-place* updates so we don't destroy existing
styles/formulas/validations in formatted templates.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class HeaderLocator:
    header_row: int
    columns: Dict[str, int]  # name -> 1-based column index


def _normalize_header(value: Any) -> str:
    text = "" if value is None else str(value)
    return text.strip()


def find_header(
    ws: Worksheet,
    *,
    required: Iterable[str],
    search_rows: int = 50,
    header_row: int | None = None,
    preferred: Iterable[str] = (),
) -> Optional[HeaderLocator]:
    required_set = {name.strip() for name in required if str(name).strip()}
    if not required_set:
        return None
    preferred_set = {name.strip() for name in preferred if str(name).strip()}

    max_ws_row = int(ws.max_row or 0)
    if max_ws_row <= 0:
        return None

    if header_row is not None:
        candidate_rows = [int(header_row)]
    else:
        max_row = min(max_ws_row, max(1, int(search_rows)))
        candidate_rows = list(range(1, max_row + 1))

    best: Optional[HeaderLocator] = None
    best_required_matches = -1
    best_preferred_matches = -1

    for row in candidate_rows:
        if row < 1 or row > max_ws_row:
            continue
        columns: Dict[str, int] = {}
        max_col = int(ws.max_column or 0)
        for col in range(1, max_col + 1):
            header = _normalize_header(ws.cell(row=row, column=col).value)
            if header:
                columns[header] = col

        required_matches = sum(1 for name in required_set if name in columns)
        if required_matches <= 0:
            continue
        preferred_matches = sum(1 for name in preferred_set if name in columns) if preferred_set else 0

        if required_set.issubset(columns.keys()):
            return HeaderLocator(header_row=row, columns=columns)

        if (required_matches, preferred_matches) > (best_required_matches, best_preferred_matches):
            best = HeaderLocator(header_row=row, columns=columns)
            best_required_matches = required_matches
            best_preferred_matches = preferred_matches

    if header_row is not None:
        return None
    return best
    return None


def open_workbook(path: str | Path):
    return load_workbook(Path(path))


def resolve_sheet(workbook, sheet_name: str | None = None) -> Worksheet:
    if sheet_name:
        return workbook[sheet_name]
    return workbook.active


def locate_plan_header(ws: Worksheet, required: Iterable[str]) -> HeaderLocator:
    locator = find_header(ws, required=required)
    if locator is None:
        required_text = ", ".join(str(n) for n in required)
        raise ValueError(f"Failed to locate header row; required columns: {required_text}")
    required_set = {str(name).strip() for name in required if str(name).strip()}
    if required_set and not required_set.issubset(locator.columns.keys()):
        missing = sorted(required_set.difference(locator.columns.keys()))
        raise ValueError(f"Failed to locate header row; missing columns: {missing}")
    return locator


def as_str(value: Any) -> str:
    return "" if value is None else str(value)


def match_cell_text(value: Any, expected: str) -> bool:
    return as_str(value).strip() == (expected or "").strip()


class ExcelTable:
    """Worksheet helper that provides header-based access by column name."""

    def __init__(
        self,
        workbook,
        ws: Worksheet,
        *,
        required_columns: Iterable[str],
        search_rows: int = 50,
        header_row: int | None = None,
        preferred_columns: Iterable[str] = (),
    ) -> None:
        self.workbook = workbook
        self.ws = ws
        self.locator = find_header(
            ws,
            required=required_columns,
            search_rows=search_rows,
            header_row=header_row,
            preferred=preferred_columns,
        )
        if self.locator is None:
            required_text = ", ".join(str(n) for n in required_columns)
            raise ValueError(f"Failed to locate header row; required columns: {required_text}")
        required_set = {str(name).strip() for name in required_columns if str(name).strip()}
        if required_set and not required_set.issubset(self.locator.columns.keys()):
            missing = sorted(required_set.difference(self.locator.columns.keys()))
            raise ValueError(f"Failed to locate header row; missing columns: {missing}")

    @classmethod
    def open(
        cls,
        path: str | Path,
        *,
        sheet_name: str | None = None,
        required_columns: Iterable[str],
        search_rows: int = 50,
        header_row: int | None = None,
        preferred_columns: Iterable[str] = (),
    ) -> "ExcelTable":
        workbook = open_workbook(path)
        ws = resolve_sheet(workbook, sheet_name)
        table = cls(
            workbook,
            ws,
            required_columns=required_columns,
            search_rows=search_rows,
            header_row=header_row,
            preferred_columns=preferred_columns,
        )
        return table

    def data_first_row(self) -> int:
        return self.locator.header_row + 1

    def get_col(self, name: str) -> int:
        try:
            return self.locator.columns[name]
        except KeyError as exc:
            raise KeyError(f"Unknown column {name!r}; available: {sorted(self.locator.columns)}") from exc

    def set_cell(self, *, row: int, column_name: str, value: Any) -> None:
        col = self.get_col(column_name)
        self.ws.cell(row=row, column=col, value=value)

    def find_first_row(self, *, column_name: str, equals_text: str) -> Optional[int]:
        col = self.get_col(column_name)
        expected = (equals_text or "").strip()
        if not expected:
            return None
        start = self.data_first_row()
        end = (self.ws.max_row or self.locator.header_row)
        for row in range(start, end + 1):
            if match_cell_text(self.ws.cell(row=row, column=col).value, expected):
                return row
        return None

    def save(self, path: str | Path) -> None:
        self.workbook.save(Path(path))
