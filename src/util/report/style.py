"""Default style/layout constants for Excel reports.

Keep everything related to look-and-feel here so future formatting changes are localized.
"""

from __future__ import annotations

from typing import Dict

from typing import Any, Optional, Union

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Style constants derived from Demo.xlsx
# ---------------------------------------------------------------------------

COLUMN_WIDTHS: Dict[str, float] = {
    "A": 13.75,
    "B": 12.625,
    "C": 11.5,
    "D": 15.0,
    "E": 14.875,
    "F": 15.125,
    "G": 18.375,
    "H": 16.125,
    "I": 14.375,
    "J": 15.0,
    "K": 12.25,
    "L": 13.5,
    "M": 12.5,
    "N": 14.0,
    "O": 12.25,
    "P": 12.25,
    "Q": 13.5,
    "R": 11.5,
    "S": 12.0,
    "T": 12.0,
    "U": 12.0,
    "V": 12.0,
    "W": 12.0,
    "X": 12.0,
    "Y": 12.0,
    "Z": 12.0,
    "AA": 12.0,
    "AB": 12.0,
    "AC": 12.0,
    "AD": 12.0,
    "AE": 12.0,
    "AF": 12.0,
    "AG": 12.0,
    "AH": 12.0,
    "AI": 12.0,
    "AJ": 12.0,
    "AK": 12.0,
    "AL": 12.0,
    "AM": 12.0,
    "AN": 12.0,
}

REPORT_LAST_COLUMN = "AN"

ROW_HEIGHT_TITLE = 42.95

CHART_COLUMN_GAP = 1
CHART_TITLE_ROW = 4
CHART_MIN_TOP_ROW = 5
CHART_VERTICAL_HEIGHT_ROWS = 18
POLAR_CHART_HEIGHT_ROWS = 20
POLAR_IMAGE_WIDTH = 525
POLAR_IMAGE_HEIGHT = 400
POLAR_ROW_SPACING = 6

COLOR_BRAND_BLUE = "2D529F"
COLOR_SUBHEADER = "B4C6E7"
COLOR_RATE_PRIMARY = "FFF2CC"
COLOR_RATE_SECONDARY = "D9E1F2"
COLOR_HEADER_GENERAL = "E7E6E6"
COLOR_RSSI_RX = "CFE2F3"
COLOR_RSSI_TX = "E2F0D9"
COLOR_GRIDLINE = "BFBFBF"
SERIES_COLORS = [
    "2D529F",  # brand blue
    "FF8C00",  # orange
    "2CA02C",  # green
    "9467BD",  # purple
    "D62728",  # red
    "17BECF",  # teal
    "8C564B",  # brown
]

FONT_TITLE = Font(name="Arial", color="FFFFFF", bold=True, size=20)
FONT_SECTION = Font(name="Arial", color="1F4E78", bold=True, size=16)
FONT_HEADER = Font(name="Arial", color="FFFFFF", bold=True, size=11)
FONT_SUBHEADER = Font(name="Arial", color="1F4E78", bold=True, size=11)
FONT_HEADER_DARK = Font(name="Arial", color="333333", bold=True, size=11)
FONT_BODY = Font(name="Arial", color="333333", size=11)
FONT_STANDARD = Font(name="Arial", color="333333", size=11)

ALIGN_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

BORDER_THIN = Border(
    left=Side(style="thin", color="8A8A8A"),
    right=Side(style="thin", color="8A8A8A"),
    top=Side(style="thin", color="8A8A8A"),
    bottom=Side(style="thin", color="8A8A8A"),
)

BORDER_THIN_DIAGONAL_DOWN = Border(
    left=Side(style="thin", color="8A8A8A"),
    right=Side(style="thin", color="8A8A8A"),
    top=Side(style="thin", color="8A8A8A"),
    bottom=Side(style="thin", color="8A8A8A"),
    diagonal=Side(style="thin", color="8A8A8A"),
    diagonalDown=True,
)

BORDER_THIN_DIAGONAL_UP = Border(
    left=Side(style="thin", color="8A8A8A"),
    right=Side(style="thin", color="8A8A8A"),
    top=Side(style="thin", color="8A8A8A"),
    bottom=Side(style="thin", color="8A8A8A"),
    diagonal=Side(style="thin", color="8A8A8A"),
    diagonalUp=True,
)


def set_cell(
    ws: Worksheet,
    row: int,
    column: int,
    value,
    *,
    font: Optional[Font] = None,
    alignment: Optional[Alignment] = None,
    fill: Optional[str] = None,
    border: Union[bool, Border] = False,
    number_format: Optional[str] = None,
) -> None:
    cell = ws.cell(row=row, column=column, value=value)
    if font is not None:
        cell.font = font
    if alignment is not None:
        cell.alignment = alignment
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    if border:
        cell.border = BORDER_THIN if border is True else border
    if number_format:
        cell.number_format = number_format


def apply_cells(
    ws: Worksheet,
    *,
    column: int,
    start_row: int,
    end_row: int,
    border: Union[bool, Border] = False,
    alignment: Optional[Alignment] = None,
    font: Optional[Font] = None,
    fill: Optional[str] = None,
    number_format: Optional[str] = None,
) -> None:
    """Apply a common style to a contiguous column segment."""

    if start_row > end_row:
        return
    resolved_border: Union[None, Border] = None
    if border:
        resolved_border = BORDER_THIN if border is True else border
    resolved_fill = PatternFill("solid", fgColor=fill) if fill else None
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=column)
        if font is not None:
            cell.font = font
        if alignment is not None:
            cell.alignment = alignment
        if resolved_fill is not None:
            cell.fill = resolved_fill
        if resolved_border is not None:
            cell.border = resolved_border
        if number_format:
            cell.number_format = number_format


def set_title(ws: Worksheet, row: int, column: int, value) -> None:
    set_cell(
        ws,
        row,
        column,
        value,
        font=FONT_TITLE,
        alignment=ALIGN_CENTER,
        fill=COLOR_BRAND_BLUE,
        border=True,
    )


def set_brand_header(ws: Worksheet, row: int, column: int, value, *, wrap: bool = True) -> None:
    set_cell(
        ws,
        row,
        column,
        value,
        font=FONT_HEADER,
        alignment=ALIGN_CENTER_WRAP if wrap else ALIGN_CENTER,
        fill=COLOR_BRAND_BLUE,
        border=True,
    )


def set_brand_header_border(
    ws: Worksheet,
    row: int,
    column: int,
    value,
    *,
    border: Union[bool, Border] = True,
    wrap: bool = True,
) -> None:
    set_cell(
        ws,
        row,
        column,
        value,
        font=FONT_HEADER,
        alignment=ALIGN_CENTER_WRAP if wrap else ALIGN_CENTER,
        fill=COLOR_BRAND_BLUE,
        border=border,
    )


def set_subheader(ws: Worksheet, row: int, column: int, value) -> None:
    set_cell(
        ws,
        row,
        column,
        value,
        font=FONT_SUBHEADER,
        alignment=ALIGN_CENTER,
        fill=COLOR_SUBHEADER,
        border=True,
    )


def set_subheader_fill(ws: Worksheet, row: int, column: int, value, *, fill: str) -> None:
    set_cell(
        ws,
        row,
        column,
        value,
        font=FONT_SUBHEADER,
        alignment=ALIGN_CENTER,
        fill=fill,
        border=True,
    )


def set_body_center(ws: Worksheet, row: int, column: int, value, *, fill: Optional[str] = None) -> None:
    set_cell(
        ws,
        row,
        column,
        value,
        font=FONT_BODY,
        alignment=ALIGN_CENTER,
        fill=fill,
        border=True,
    )


def set_body_number(
    ws: Worksheet,
    row: int,
    column: int,
    value,
    *,
    number_format: str,
    fill: Optional[str] = None,
) -> None:
    set_cell(
        ws,
        row,
        column,
        value,
        font=FONT_BODY,
        alignment=ALIGN_CENTER,
        fill=fill,
        border=True,
        number_format=number_format,
    )


def set_body_center_wrap(ws: Worksheet, row: int, column: int, value, *, fill: Optional[str] = None) -> None:
    set_cell(
        ws,
        row,
        column,
        value,
        font=FONT_BODY,
        alignment=ALIGN_CENTER_WRAP,
        fill=fill,
        border=True,
    )


def set_standard_left_wrap(ws: Worksheet, row: int, column: int, value) -> None:
    set_cell(
        ws,
        row,
        column,
        value,
        font=FONT_STANDARD,
        alignment=ALIGN_LEFT_WRAP,
        border=True,
    )


def merge_row(ws: Worksheet, *, row: int, start_col: int, end_col: int) -> None:
    if end_col <= start_col:
        return
    ws.merge_cells(f"{get_column_letter(start_col)}{row}:{get_column_letter(end_col)}{row}")


def merge_row_letters(ws: Worksheet, *, row: int, start_col_letter: str, end_col_letter: str) -> None:
    if not start_col_letter or not end_col_letter:
        return
    if start_col_letter == end_col_letter:
        return
    ws.merge_cells(f"{start_col_letter}{row}:{end_col_letter}{row}")


def set_merged_row(
    ws: Worksheet,
    *,
    row: int,
    start_col_letter: str,
    end_col_letter: str,
    value: Any,
    font: Optional[Font] = None,
    alignment: Optional[Alignment] = None,
    fill: Optional[str] = None,
    border: Union[bool, Border] = False,
) -> None:
    merge_row_letters(ws, row=row, start_col_letter=start_col_letter, end_col_letter=end_col_letter)
    start_col = column_index_from_string(start_col_letter)
    set_cell(
        ws,
        row,
        start_col,
        value,
        font=font,
        alignment=alignment,
        fill=fill,
        border=border,
    )


def apply_report_remark(ws: Worksheet, *, row: int, end_col_letter: str, text: str) -> None:
    set_merged_row(
        ws,
        row=row,
        start_col_letter="A",
        end_col_letter=end_col_letter,
        value=text,
        font=FONT_SUBHEADER,
        alignment=ALIGN_CENTER_WRAP,
        border=True,
    )


def apply_report_summary(ws: Worksheet, *, row: int, end_col_letter: str, text: str) -> None:
    set_merged_row(
        ws,
        row=row,
        start_col_letter="A",
        end_col_letter=end_col_letter,
        value=text,
        font=FONT_BODY,
        alignment=ALIGN_CENTER_WRAP,
        border=True,
    )


def apply_section_header(ws: Worksheet, *, row: int, end_col_letter: str, text: str) -> None:
    set_merged_row(
        ws,
        row=row,
        start_col_letter="A",
        end_col_letter=end_col_letter,
        value=text,
        font=FONT_SECTION,
        alignment=ALIGN_CENTER,
        border=True,
    )


def apply_empty_report_placeholder(ws: Worksheet, *, row: int, end_col_letter: str, text: str) -> None:
    apply_section_header(ws, row=row, end_col_letter=end_col_letter, text=text)


def apply_header_block(
    ws: Worksheet,
    *,
    row: int,
    start_col: int,
    end_col: int,
    label: str,
    wrap: bool = True,
) -> None:
    """Merge a horizontal header range and place a brand header label in the left-most cell."""

    if end_col > start_col:
        merge_row_letters(
            ws,
            row=row,
            start_col_letter=get_column_letter(start_col),
            end_col_letter=get_column_letter(end_col),
        )
    set_brand_header(ws, row, start_col, label, wrap=wrap)
