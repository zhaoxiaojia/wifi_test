# !/usr/bin/env python
# -*-coding:utf-8 -*-

"""
# File       : test_rvr.py
# Time       ：2023/9/15 14:03
# Author     ：chao.li
# version    ：python 3.9
# Description：
"""

import bisect
import csv
import itertools
import logging
import os
import re
import signal
import subprocess
import threading
import time

import pytest

from tools.router_tool.Router import Router
from tools.connect_tool.TelnetInterface import TelnetInterface
from tools.yamlTool import yamlTool
from tools.router_tool.AsusRouter.Asusax88uControl import Asusax88uControl
import openpyxl
from copy import copy

filename = 'XiaoMi-Rvr.xlsx'
rvr_xlsx = openpyxl.load_workbook(filename)
sheet = rvr_xlsx['Sheet1']
new_sheet = rvr_xlsx.create_sheet(title=f'{pytest.timestamp}')

for row in sheet.iter_rows(values_only=False):
    for cell in row:
        new_sheet[cell.coordinate].value = copy(cell.value)
        new_sheet[cell.coordinate].font = copy(cell.font)
        new_sheet[cell.coordinate].border = copy(cell.border)
        new_sheet[cell.coordinate].fill = copy(cell.fill)
        new_sheet[cell.coordinate].number_format = copy(cell.number_format)
        new_sheet[cell.coordinate].protection = copy(cell.protection)
        new_sheet[cell.coordinate].alignment = copy(cell.alignment)

for merged_range in sheet.merged_cells.ranges:
    new_sheet.merge_cells(str(merged_range))

rvr_xlsx.save(filename)


def writeInExcelArea(value, row_num, col_num):
    for i in range(0, len(value)):
        logging.info(f'execl write {row_num} {i + col_num}')
        new_sheet.cell(row=row_num, column=i + col_num, value=value[i])


# 读取 测试配置
with open(os.getcwd() + '/config/asusax88u.csv', 'r') as f:
    reader = csv.reader(f)
    test_data = [Router(*[i.strip() for i in row]) for row in reader][1:]

logging.info(test_data)

# 设置为True 时 开启 衰减测试流程
rf_needed = False
# 设置为True 时 开启 状态测试流程
corner_needed = False
# 设置为True 时 开启 路由相关配置
router_needed = True

# 设置是否需要push iperf
iperf_tool = False

# 无法使用 命令行 连接wifi 是 设置为true
third_dut = False
if pytest.connect_type == 'telnet':
    third_dut = True

sum_list_lock = threading.Lock()

# loading config.yaml 文件 获取数据  dict 数据类型
wifi_yaml = yamlTool(os.getcwd() + '/config/config.yaml')
command_data = wifi_yaml.get_note('env_command')
router_name = wifi_yaml.get_note('router')['name']
router = ''

rvr_tool = wifi_yaml.get_note('rvr')['tool']
if rvr_tool == 'iperf':
    test_tool = wifi_yaml.get_note('rvr')[rvr_tool]['version']
    tool_path = wifi_yaml.get_note('rvr')[rvr_tool]['path'] or ''
    logging.info(f'test_tool {test_tool}')
if rvr_tool == 'ixchariot':
    # Todo
    ...
    # logging.info(f'test_tool {test_tool}')

# 实例路由器对象
if router_needed:
    exec(f'router = {router_name.capitalize()}Control()')

# env_control = wifi_yaml.get_note('env_control')


# 初始化 衰减 & 转台 对象
if rf_needed:
    # 读取衰减 配置
    rf_step_list = []
    model = wifi_yaml.get_note('rf_solution')['model']
    if model != 'RADIORACK-4-220' and model != 'RC4DAT-8G-95':
        raise EnvironmentError("Doesn't support this model")

    if model == 'RADIORACK-4-220':
        rf_ip = wifi_yaml.get_note('rf_solution')[model]['ip_address']
    if model == 'RC4DAT-8G-95':
        rf_ip = '192.168.50.19'

    logging.info('test rf')
    rf_tool = TelnetInterface(rf_ip)
    logging.info(f'rf_ip {rf_ip}')
    rf_step_list = wifi_yaml.get_note('rf_solution')['step']
    rf_step_list = [i for i in range(*rf_step_list)][::2]
    logging.info(f'rf_step_list {rf_step_list}')

if corner_needed:
    corner_step_list = []
    # 配置衰减
    corner_ip = wifi_yaml.get_note('corner_angle')['ip_address']

    logging.info('test corner')
    corner_tool = TelnetInterface(corner_ip)
    logging.info(f'corner_ip {corner_ip}')
    corner_step_list = wifi_yaml.get_note('corner_angle')['step']
    corner_step_list = [i for i in range(*corner_step_list)][::45]
    logging.info(f'corner step_list {corner_step_list}')

step_list = [1]
if rf_needed and rf_step_list:
    step_list = rf_step_list
if corner_needed and corner_step_list:
    step_list = corner_step_list
if rf_needed and corner_needed and rf_step_list and corner_step_list:
    step_list = itertools.product(corner_step_list, rf_step_list)

logging.info(f'finally step_list {step_list}')

# 配置 测试报告
pytest.testResult.x_path = [] if (rf_needed and corner_needed) == 'both' else step_list
pytest.testResult.init_rvr_result()
tx_result_list, rx_result_list = [], []
rx_result, tx_result = '', ''


def iperf_on(command, adb, direction='tx'):
    def server_on():
        logging.info(f'server {command} ->{adb}<-')
        with open('temp.txt', 'w') as f:
            if adb == 'executer':
                iperf_log = pytest.dut.checkoutput(command)
                f.write(iperf_log)
            popen = subprocess.Popen(command.split(), stdout=f, encoding='gbk')
        # logging.info(subprocess.run('tasklist | findstr "iperf"'.replace('iperf',pc_ipef),shell=True,encoding='gbk'))
        # logging.info(pytest.dut.checkoutput('ps -A|grep "iperf"'.replace('iperf',dut_iperf)))
        logging.info('write done')
        return popen

    if adb:
        if test_tool == 'iperf3':
            command = 'iperf3 -s -1'
        if adb != 'executer':
            command = f'adb -s {adb} shell ' + command
    else:
        if test_tool == 'iperf3':
            command = f'iperf3 -c {pytest.dut.dut_ip} -i1 -t30 -P5'
            if direction == 'tx':
                command = f'iperf3 -c {pytest.dut.dut_ip} -i1 -t30 -P5 -R'

    logging.info(f'command {command} ->{adb}<-')

    if re.findall(r'iperf[3]?.*?-s', command):
        popen = server_on()
    else:
        if adb == 'executer':
            pytest.dut.checkoutput(command)
        popen = subprocess.Popen(command.split(), encoding='gbk')
    return popen


def server_off(popen):
    if pytest.connect_type == 'adb':
        if not isinstance(popen, subprocess.Popen):
            logging.warning('pls pass in the popen object')
            return 'pls pass in the popen object'
        try:
            os.kill(popen.pid, signal.SIGTERM)
        except Exception as e:
            ...
        popen.terminate()
    elif pytest.connect_type == 'telnet':
        pytest.dut.tn.close()


def get_logcat(pair):
    # pytest.dut.kill_iperf()
    # 分析 iperf 测试结果
    result_list = []
    with open('temp.txt', 'r') as f:
        for line in f.readlines():
            logging.info(f'line : {line.strip()}')
            if pair != 1:
                if '[SUM]' not in line:
                    continue
            if re.findall(r'.*?\d+\.\d*-\s*\d+\.\d*.*?(\d+\.*\d*)\s+Mbits/sec.*?', line.strip(), re.S):
                result_list.append(
                    int(float(
                        re.findall(r'.*?\d+\.\d*-\s*\d+\.\d*.*?(\d+\.*\d*)\s+Mbits/sec.*?', line.strip(), re.S)[0])))

    if result_list:
        logging.info(f'{sum(result_list) / len(result_list)}')
        logging.info(f'{result_list}')
        result = sum(result_list) / len(result_list)
    else:
        result = 0
    return result


def push_iperf():
    if iperf_tool and pytest.connect_type == 'adb' and (
            pytest.dut.checkoutput('[ -e /system/bin/iperf ] && echo yes || echo no').strip() != 'yes'):
        path = os.path.join(os.getcwd(), 'res/iperf')
        pytest.dut.push(path, '/system/bin')
        pytest.dut.checkoutput('chmod a+x /system/bin/iperf')


@pytest.fixture(scope='session', autouse=True, params=test_data)
def wifi_setup_teardown(request):
    global x_result, tx_resul, tx_result_list, rx_result_list
    tx_result_list.clear()
    rx_result_list.clear()
    logging.info('==== wifi env setup start')
    # push_iperf()
    router_info = request.param
    if router_needed:
        # 修改路由器配置
        assert router.change_setting(router_info), "Can't set ap , pls check first"

    logging.info('wifi env set done')
    with open(pytest.testResult.detail_file, 'a', encoding='gbk') as f:
        f.write(f'Testing {router_info} \n')

    # 重置衰减&转台

    # 衰减器置0
    if rf_needed:
        logging.info('Reset rf value')
        rf_tool.execute_rf_cmd(0)
        logging.info(rf_tool.get_rf_current_value())
        time.sleep(10)

    # 转台置0
    if corner_needed:
        logging.info('Reset corner')
        corner_tool.set_turntable_zero()
        logging.info(corner_tool.get_turntanle_current_angle())
        time.sleep(3)

    if third_dut:
        connect_status = True
        if not router_needed:
            # router 有修改时 等待 30 秒 让板子回连
            time.sleep(30)
    else:
        # 连接 网络 最多三次重试
        for _ in range(3):
            if not router_needed:
                break
            try:
                type = 'wpa3' if 'WPA3' in router_info.authentication_method else 'wpa2'
                if router_info.authentication_method.lower() in \
                        ['open', '不加密', '无', 'open system', '无加密(允许所有人连接)', 'none']:
                    logging.info('no passwd')
                    cmd = pytest.dut.CMD_WIFI_CONNECT.format(router_info.ssid, "open", "")
                else:
                    cmd = pytest.dut.CMD_WIFI_CONNECT.format(router_info.ssid, type,
                                                             router_info.wpa_passwd)
                if router_info.hide_ssid == '是':
                    if int(pytest.dut.getprop('ro.build.version.sdk')) >= 31:
                        cmd += pytest.dut.CMD_WIFI_HIDE
                    else:
                        cmd = (pytest.dut.WIFI_CONNECT_COMMAND_REGU.format(router_info.ssid) +
                               pytest.dut.WIFI_CONNECT_PASSWD_REGU.format(router_info.wpa_passwd) +
                               pytest.dut.WIFI_CONNECT_HIDE_SSID_REGU.format(router_info.hide_type))
                pytest.dut.checkoutput(cmd)
                if pytest.dut.wait_for_wifi_address():
                    connect_status = True
                    break
            except Exception as e:
                logging.info(e)
                connect_status = False

    if pytest.connect_type == 'telnet':
        pytest.dut.dut_ip = pytest.dut.ip
    else:
        dut_info = pytest.dut.checkoutput('ifconfig wlan0')
        pytest.dut.dut_ip = re.findall(r'inet addr:(\d+\.\d+\.\d+\.\d+)', dut_info, re.S)[0]
    logging.info(f'dut_ip:{pytest.dut.dut_ip}')
    connect_status = True
    ipfoncig_info = pytest.dut.checkoutput_term('ipconfig').strip()
    pytest.dut.pc_ip = re.findall(r'IPv4 地址.*?(\d+\.\d+\.\d+\.\d+)', ipfoncig_info, re.S)[0]
    logging.info(f'pc_ip:{pytest.dut.pc_ip}')
    logging.info('==== wifi env setup done')
    yield connect_status, router_info
    # 后置动作
    kill_iperf()
    # 重置结果
    logging.info(f'tx_result_list {tx_result_list}')
    logging.info(f'rx_result_list {rx_result_list}')
    logging.info(f'len  {len(tx_result_list)}')
    if len(tx_result_list) == 3:
        writeInExcelArea(rx_result_list, row_num=int(router_info.data_row), col_num=10)
        writeInExcelArea(tx_result_list, row_num=int(router_info.data_row) + 1, col_num=10)

    rvr_xlsx.save(filename)
    # if not router_debug:
    #     router.router_control.driver.quit()


# @pytest.fixture(scope='session', autouse=True, params=command_data)
# def session_setup_teardown(request):
#     logging.info('==== debug command setup start')
# iperf 配置


# 获取板子的ip 地址

# command_info = request.param
# logging.info(f"command_info : {command_info}")
# pytest.dut.subprocess_run(command_info)
# logging.info('==== debug command setup done')
# yield
# 后置动作

# 生成 pdf
# if step_list != [0]:
#     pytest.testResult.write_to_excel()
#     if test_type == 'rf':
#         # 重置衰减
#         if not rf_debug:
#             rf_tool.execute_rf_cmd(0)
#         # 生成折线图
#         pytest.testResult.write_attenuation_data_to_pdf()
#     elif test_type == 'corner':
#         # 转台重置
#         if not rf_debug:
#             corner_tool.set_turntable_zero()
#         # 生成雷达图
#         pytest.testResult.write_corner_data_to_pdf()
#     else:
#         ...


pair_count = {
    'n': {
        '2': wifi_yaml.get_note('pair_num')['n']['2'],
        '5': wifi_yaml.get_note('pair_num')['n']['5']
    },
    'ac': {
        '5': wifi_yaml.get_note('pair_num')['ac']['5']
    },
    'ax': {
        '2': wifi_yaml.get_note('pair_num')['ax']['2'],
        '5': wifi_yaml.get_note('pair_num')['ax']['5'],
    },
    'auto': {
        '2': wifi_yaml.get_note('pair_num')['auto']['2'],
        '5': wifi_yaml.get_note('pair_num')['auto']['5'],
    }
}


def set_pair_count(router_info, rssi_num, type, dire):
    '''
    匹配 打流通道数
    '''
    if 'AX' in router_info.wireless_mode:
        type = 'ax'
    elif 'AC' in router_info.wireless_mode:
        type = 'ac'
    elif '自动' in router_info.wireless_mode:
        type = 'auto'
    else:
        type = 'n'

    if '2' in router_info.band:
        band = '2'
    else:
        band = '5'

    _2g = [10, 20, 30, 40]
    _5g = [10, 20, 30, 40, 50]

    if band == '2':
        target_list = _2g
    else:
        target_list = _5g
    logging.info(f'rf current db {rssi_num}')
    pair = pair_count[type][band][dire][bisect.bisect(target_list, rssi_num)]
    logging.info(f'pair {pair}')
    return int(pair)


def kill_iperf():
    # kill iperf
    if rvr_tool == 'iperf':
        pytest.dut.subprocess_run(pytest.dut.IPERF_KILL.replace('iperf', test_tool))
        pytest.dut.popen_term(pytest.dut.IPERF_WIN_KILL.replace('iperf', test_tool))


def get_tx_rate(router_info, pair, freq_num, rssi_num, type, corner_set='', db_set=''):
    global tx_result_list

    # 最多三次 重试机会
    for _ in range(3):
        logging.info('run tx ')
        tx_result = 0
        mcs_tx = 0
        # pytest.dut.checkoutput(pytest.dut.CLEAR_DMESG_COMMAND)
        # pytest.dut.checkoutput(pytest.dut.MCS_TX_KEEP_GET_COMMAND)
        # kill iperf
        if rvr_tool == 'iperf':
            kill_iperf()
            time.sleep(1)
            if test_tool == 'iperf3':
                adb_popen = iperf_on(tool_path + pytest.dut.IPERF_CLIENT_REGU[type]['tx'].format(
                    pytest.dut.pc_ip,
                    pytest.dut.IPERF_TEST_TIME,
                    pair if type == 'TCP' else 1), pytest.dut.serialnumber)
                pc_popen = iperf_on(pytest.dut.IPERF_SERVER[type], '')
            else:
                pc_popen = iperf_on(pytest.dut.IPERF_SERVER[type], '')
                adb_popen = iperf_on(tool_path + pytest.dut.IPERF_CLIENT_REGU[type]['tx'].format(
                    pytest.dut.pc_ip,
                    pytest.dut.IPERF_TEST_TIME,
                    pair if type == 'TCP' else 1), pytest.dut.serialnumber)

            time.sleep(pytest.dut.IPERF_WAIT_TIME)
            if pytest.connect_type == 'telnet':
                time.sleep(15)
            time.sleep(3)
            server_off(adb_popen)
            server_off(pc_popen)
            tx_result = get_logcat(pair if type == 'TCP' else 1)

        if rvr_tool == 'ixchariot':
            # Todo
            ...

        if tx_result == False:
            logging.info("Connect failed")
            continue

        mcs_tx = pytest.dut.get_mcs_tx()
        if tx_result and mcs_tx:
            logging.info(f'{tx_result}, {mcs_tx}')
            tx_result_list.append(tx_result)
            break

    corner = 'None'
    corner = corner_tool.get_turntanle_current_angle() if corner_needed else corner_set

    tx_result_info = (
        f'P0 RvR Standalone NULL Null {router_info.wireless_mode.split()[0]} {router_info.band.split()[0]} '
        f'{router_info.bandwidth.split()[0]} Rate_Adaptation {router_info.channel} {type} UL NULL NULL {db_set} {rssi_num} {corner} NULL {tx_result} {mcs_tx if mcs_tx else "NULL"}')
    logging.info(tx_result_info)
    pytest.testResult.save_result(tx_result_info.replace(' ', ','))
    with open(pytest.testResult.detail_file, 'a') as f:
        f.write(f'Tx {type} result : {tx_result}\n')
        f.write('-' * 40 + '\n\n')


def get_rx_rate(router_info, pair, freq_num, rssi_num, type, corner_set='', db_set=''):
    global rx_result_list
    for _ in range(3):
        logging.info('run rx ')
        rx_result = 0
        mcs_rx = 0
        # clear mcs data
        # pytest.dut.checkoutput(pytest.dut.CLEAR_DMESG_COMMAND)
        # pytest.dut.checkoutput(pytest.dut.MCS_RX_CLEAR_COMMAND)
        # kill iperf
        if rvr_tool == 'iperf':
            kill_iperf()
            time.sleep(1)
            adb_popen = iperf_on(tool_path + pytest.dut.IPERF_SERVER[type], pytest.dut.serialnumber)
            pc_popen = iperf_on(
                pytest.dut.IPERF_CLIENT_REGU[type]['rx'].format(
                    pytest.dut.dut_ip, pytest.dut.IPERF_TEST_TIME,
                    pair if type == 'TCP' else 4), '', direction='rx')
            time.sleep(pytest.dut.IPERF_WAIT_TIME)
            if pytest.connect_type == 'telnet':
                time.sleep(15)
            server_off(adb_popen)
            server_off(pc_popen)
            rx_result = get_logcat(pair if type == 'TCP' else 4)

        if rvr_tool == 'ixchariot':
            # Todo
            ...
        if rx_result == False:
            logging.info("Connect failed")
            continue
        time.sleep(3)

        # get mcs data
        mcs_rx = pytest.dut.get_mcs_rx()
        if rx_result and mcs_rx:
            logging.info(f'{rx_result}, {mcs_rx}')
            rx_result_list.append(rx_result)
            break
    corner = 'None'
    corner = corner_tool.get_turntanle_current_angle() if corner_needed else corner_set

    rx_result_info = (
        f'P0 RvR Standalone NULL Null {router_info.wireless_mode.split()[0]} {router_info.band.split()[0]} '
        f'{router_info.bandwidth.split()[0]} Rate_Adaptation {router_info.channel} {type} DL NULL NULL {db_set} {rssi_num} {corner} NULL {rx_result} {mcs_rx if mcs_rx else "NULL"}')
    pytest.testResult.save_result(rx_result_info.replace(' ', ','))
    with open(pytest.testResult.detail_file, 'a', encoding='gbk') as f:
        logging.info('writing')
        f.write(f'Rx {type} result : {rx_result}\n')
        f.write('-' * 40 + '\n\n')


# 测试 iperf
@pytest.mark.repeat(3)
@pytest.mark.parametrize("rf_value", step_list)
def test_wifi_rvr(wifi_setup_teardown, rf_value):
    global rx_result, tx_result
    # 判断板子是否存在  ip
    if not wifi_setup_teardown[0]:
        logging.info("Can't connect wifi ,input 0")
        # rx_result_list.append('0')
        # tx_result_list.append('0')
        with open(pytest.testResult.detail_file, 'a') as f:
            f.write("\n Can't connect wifi , skip this loop\n\n")
        return
    router_info = wifi_setup_teardown[1]

    # 执行 修改 步长
    # 修改衰减
    if rf_needed:
        logging.info('set rf value')
        value = rf_value[1] if type(rf_value) == tuple else rf_value
        logging.info(value)
        rf_tool.execute_rf_cmd(value)
        # 获取当前衰减值
        logging.info(rf_tool.get_rf_current_value())
    if corner_needed:
        logging.info('set corner value')
        value = rf_value[0] if type(rf_value) == tuple else rf_value
        logging.info(value)
        corner_tool.execute_turntable_cmd('rt', angle=value * 10)
        # 获取转台角度
        logging.info(corner_tool.get_turntanle_current_angle())

    with open(pytest.testResult.detail_file, 'a') as f:
        f.write('-' * 40 + '\n')
        info, corner_set = '', ''
        db_set = 0
        if rf_needed:
            db_set = rf_value[1] if type(rf_value) == tuple else rf_value
            info += 'db_set : ' + str(db_set) + '\n'

        if corner_needed:
            corner_set = rf_value[0] if type(rf_value) == tuple else rf_value
            info += 'corner_set : ' + str(corner_set) + '\n'

        f.write(info)
    # time.sleep(1)

    # 获取rssi
    for i in range(3):
        rssi_info = pytest.dut.checkoutput(pytest.dut.IW_LINNK_COMMAND)
        time.sleep(1)
        if 'signal' in rssi_info:
            break
    logging.info(rssi_info)
    if 'Not connected' in rssi_info:
        logging.info('The signal strength is not enough ,input 0')
        rx_result, tx_result, rssi_num = 0, 0, 0
        with open(pytest.testResult.detail_file, 'a') as f:
            f.write('signal strength is not enough no rssi \n')
        assert False, "Wifi is not connected"
    logging.info('Start test')

    try:
        rssi_num = int(re.findall(r'signal:\s+-?(\d+)\s+dBm', rssi_info, re.S)[0])
        freq_num = int(re.findall(r'freq:\s+(\d+)\s+', rssi_info, re.S)[0])
        with open(pytest.testResult.detail_file, 'a') as f:
            f.write(f'Rssi : {rssi_num}\n')
            f.write(f'Freq : {freq_num}\n')
    except IndexError as e:
        rssi_num = -1
        freq_num = -1
    # handle iperf pair count
    logging.info(router_info)
    protocol = 'TCP' if 'TCP' in router_info.protocol_type else 'UDP'
    # iperf  打流
    if 'tx' in router_info.test_type:
        # 动态匹配 打流通道数
        pair = set_pair_count(router_info, db_set, protocol, 'tx')
        logging.info(f'rssi : {rssi_num} pair : {pair}')
        get_tx_rate(router_info, pair, freq_num, rssi_num, protocol, corner_set=corner_set, db_set=db_set)
    if 'rx' in router_info.test_type:
        pair = set_pair_count(router_info, db_set, protocol, 'rx')
        logging.info(f'rssi : {rssi_num} pair : {pair}')
        get_rx_rate(router_info, pair, freq_num, rssi_num, protocol, corner_set=corner_set, db_set=db_set)
